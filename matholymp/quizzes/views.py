from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from django.core.mail import send_mail
from django.conf import settings
from .models import Evaluacion, AdminProfile, Participantes, GrupoParticipantes, Representante, SolicitudClaveTemporal, UserProfile, IntentosParticipante, AuditLog
from .email_utils import generate_email_messages
from .scope_utils import get_user_scope, filter_queryset_by_scope
from .decorators import superuser_required, full_access_required, admin_required
from django.utils import timezone
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import authenticate, login,logout
from django.contrib.auth.models import User
from django.utils.crypto import get_random_string
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db.models import Q
from django.db import models
import re
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from openpyxl import load_workbook
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import os
from django.conf import settings
from .models import Pregunta, Opcion, Categoria
from .models import ResultadoEvaluacion
from django.db.models import Avg
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from io import BytesIO
try:
    from PIL import Image as PILImage
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False





# Función helper para verificar acceso total
def has_full_access(user):
    """
    Verifica si un usuario tiene acceso total al sistema.
    Retorna True si es superuser o si es admin con acceso_total=True
    """
    if user.is_superuser:
        return True
    try:
        admin_profile = AdminProfile.objects.get(user=user)
        return admin_profile.acceso_total
    except AdminProfile.DoesNotExist:
        return False

def check_question_modification_allowed(evaluacion):
    """
    Verifica si se pueden modificar las preguntas de una evaluación.
    Retorna una tupla (allowed: bool, message: str)
    """
    if not evaluacion.can_modify_questions():
        message = evaluacion.get_question_modification_restriction_message()
        return False, message
    return True, ""


def get_evaluacion_monitoreable_or_404(request, pk):
    """Obtiene una evaluación respetando el alcance del administrador actual."""
    evaluaciones = filter_queryset_by_scope(
        Evaluacion.objects.all(), request, model_name='Evaluacion'
    )
    return get_object_or_404(evaluaciones, pk=pk)


def get_resultado_monitoreable_or_404(request, resultado_id, evaluacion=None):
    """Obtiene un resultado de auditoría dentro de la carrera autorizada."""
    resultados = ResultadoEvaluacion.objects.select_related('evaluacion', 'participante')
    if evaluacion is not None:
        resultados = resultados.filter(evaluacion=evaluacion)
    else:
        # Un detalle identifica un resultado histórico; no debe fallar sólo porque
        # el administrador cambió luego el concurso activo en su sesión. El límite
        # de seguridad se conserva por carrera (o por carrera seleccionada por un
        # superusuario).
        is_global, carrera, _ = get_user_scope(request)
        if carrera:
            resultados = resultados.filter(evaluacion__concurso__carrera=carrera)
        elif not is_global:
            resultados = resultados.none()
    return get_object_or_404(resultados, pk=resultado_id)


def calcular_tiempo_restante_servidor(resultado, evaluacion):
    """Calcula el tiempo restante desde el inicio; nunca confía en el navegador."""
    if not resultado.fecha_inicio:
        return evaluacion.duration_minutes * 60
    tiempo_total = evaluacion.duration_minutes * 60
    transcurrido = int((timezone.now() - resultado.fecha_inicio).total_seconds())
    return max(0, tiempo_total - transcurrido)


# Función para validar contraseñas de forma robusta
def validate_password_strength(password, username):
    """
    Valida la fortaleza de una contraseña.
    
    Args:
        password (str): La contraseña a validar
        username (str): El nombre de usuario para verificar similitud
    
    Returns:
        str or None: Mensaje de error si la contraseña no es válida, None si es válida
    """
    # 1. Validar longitud mínima
    if len(password) < 6:
        return 'La contraseña debe tener al menos 6 caracteres.'
    
    # 2. Validar que no sea similar al nombre de usuario
    if password.lower() in username.lower() or username.lower() in password.lower():
        return 'La contraseña no puede ser similar al nombre de usuario.'
    
    # 3. Validar que no sea completamente numérica
    if password.isdigit():
        return 'La contraseña no puede ser completamente numérica.'
    
    # 4. Validar contraseñas comunes
    common_passwords = ['12345678', '123456','1234567', '123456789', 'password', 'contraseña', 'qwerty123', '87654321', 'abc12345']
    if password.lower() in common_passwords:
        return 'La contraseña es demasiado común. Elige una más segura.'
    
    # 5. Validar que tenga al menos una letra y un número
    if not any(c.isalpha() for c in password) or not any(c.isdigit() for c in password):
        return 'La contraseña debe contener al menos una letra y un número.'
    
    # Si llegamos aquí, la contraseña es válida
    return None


# Vista de login

def custom_login(request):
    # Si el usuario ya está autenticado, redirigir al dashboard
    if request.user.is_authenticated:
        return redirect('quizzes:dashboard')
        
    if request.GET.get('session_expired'):
        messages.warning(request, 'Tu sesión expiró por inactividad')
        return redirect(settings.LOGIN_URL)
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        # Buscar usuario por username case-insensitive
        try:
            user_obj = User.objects.get(username__iexact=username)
            # Intentar autenticar con el username real (case-sensitive) pero password
            user = authenticate(request, username=user_obj.username, password=password)
        except User.DoesNotExist:
            user = None
        
        if user is not None:
            login(request, user)
            request.session['last_activity'] = timezone.now().timestamp()
            messages.success(request, f'¡Bienvenido, {user.get_full_name() or user.username}!')
            return redirect('quizzes:dashboard')
        else:
            # Mostrar mensaje de error específico
            if not username:
                messages.error(request, 'Por favor, ingresa tu nombre de usuario.')
            elif not password:
                messages.error(request, 'Por favor, ingresa tu contraseña.')
            else:
                messages.error(request, 'Usuario o contraseña incorrectos. Por favor, verifica tus credenciales.')
    
    # Crear un formulario vacío para el template
    form = AuthenticationForm()

    return render(request, 'quizzes/login.html', {
        'form': form,
        'messages': messages.get_messages(request)
    })



def custom_logout(request):
    logout(request)  # Cierra la sesión
    return redirect('quizzes:login')  # Redirige al dashboard después de cerrar sesión


def session_check(request):
    return JsonResponse({
        'is_authenticated': request.user.is_authenticated
    }, status=200 if request.user.is_authenticated else 403)


def obtener_diccionario_respuestas(resultado):
    """
    Extrae el diccionario plano {'pregunta_102': '415'} desde respuestas_guardadas,
    soporta diccionarios planos y snapshots congelados.
    """
    if not resultado or not resultado.respuestas_guardadas:
        return {}
        
    data = resultado.respuestas_guardadas
    if isinstance(data, dict):
        if 'preguntas_snapshot' in data:
            resp_dict = {}
            for item in data['preguntas_snapshot']:
                p_id = item.get('pregunta_id')
                o_id = item.get('respuesta_estudiante_id')
                if p_id and o_id:
                    resp_dict[f'pregunta_{p_id}'] = str(o_id)
                    resp_dict[str(p_id)] = str(o_id)
            return resp_dict
        return data
    return {}


def obtener_progreso_respuestas(resultado, preguntas_mostradas):
    """Cuenta respuestas válidas del intento sin depender del modelo de monitoreo eliminado."""
    respuestas = obtener_diccionario_respuestas(resultado)
    respondidas = sum(
        bool(respuestas.get(f'pregunta_{pregunta.id}') or respuestas.get(str(pregunta.id)))
        for pregunta in preguntas_mostradas
    )
    return respondidas, len(preguntas_mostradas)


def generar_snapshot_respuestas(preguntas_mostradas, respuestas_diccionario):
    """
    Genera un snapshot JSON congelado e inmutable al entregar un examen.
    Preserva los textos completos de las preguntas, opciones, respuesta elegida y la opción correcta.
    """
    preguntas_snapshot = []
    respuestas_diccionario = respuestas_diccionario or {}
    
    for pregunta in preguntas_mostradas:
        pregunta_key = f'pregunta_{pregunta.id}'
        opcion_seleccionada_id_raw = respuestas_diccionario.get(pregunta_key)
        
        opciones_list = []
        opcion_seleccionada_dict = None
        opcion_correcta_dict = None
        es_correcta = False
        
        opciones_qs = pregunta.opciones.all()
        for opcion in opciones_qs:
            es_sel = False
            if opcion_seleccionada_id_raw is not None:
                es_sel = (str(opcion.id) == str(opcion_seleccionada_id_raw))
                
            opcion_info = {
                'id': opcion.id,
                'text': opcion.text,
                'is_correct': opcion.is_correct,
                'seleccionada': es_sel
            }
            opciones_list.append(opcion_info)
            
            if opcion.is_correct:
                opcion_correcta_dict = opcion_info
            
            if es_sel:
                opcion_seleccionada_dict = opcion_info
                if opcion.is_correct:
                    es_correcta = True
                    
        peso = getattr(pregunta, 'puntos', 1) or 1
        puntos_ganados = peso if es_correcta else 0
        
        opcion_id_int = None
        if opcion_seleccionada_id_raw is not None and str(opcion_seleccionada_id_raw).isdigit():
            opcion_id_int = int(opcion_seleccionada_id_raw)
            
        preguntas_snapshot.append({
            'pregunta_id': pregunta.id,
            'text': pregunta.text,
            'puntos_pregunta': peso,
            'puntos_ganados': puntos_ganados,
            'es_correcta': es_correcta,
            'opciones': opciones_list,
            'respuesta_estudiante_id': opcion_id_int,
            'respuesta_estudiante': opcion_seleccionada_dict,
            'opcion_correcta': opcion_correcta_dict
        })
        
    return {
        'completada': True,
        'fecha_entrega': timezone.now().isoformat(),
        'preguntas_snapshot': preguntas_snapshot
    }


@login_required
def take_quiz(request, pk):
    evaluacion = get_object_or_404(Evaluacion, pk=pk)
    
    # Verificar que sea estudiante
    if not Participantes.objects.filter(user=request.user).exists():
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('quizzes:dashboard')
    
    # Obtener el participante
    participante = Participantes.objects.get(user=request.user)
    
    # Verificar que el participante esté autorizado para esta evaluación
    participantes_autorizados = evaluacion.get_participantes_autorizados()
    if participante not in participantes_autorizados:
        messages.error(request, 'No estás autorizado para rendir esta evaluación.')
        return redirect('quizzes:quiz')

    # VERIFICAR INTENTOS DISPONIBLES antes de proceder
    intentos_disponibles = participante.get_intentos_disponibles(evaluacion)
    if intentos_disponibles <= 0:
        intentos_usados = participante.get_intentos_usados(evaluacion)
        messages.error(request, f'Has agotado tus intentos para esta evaluación. Intentos utilizados: {intentos_usados}')
        return redirect('quizzes:quiz')
    
    # Verificar si hay un intento activo (no completado)
    resultado_activo = ResultadoEvaluacion.objects.filter(
        evaluacion=evaluacion,
        participante=participante,
        completada=False
    ).first()
    
    # Si hay un intento activo, el servidor es la única fuente de verdad del tiempo.
    continuar_evaluacion = False
    if resultado_activo:
        tiempo_restante = calcular_tiempo_restante_servidor(resultado_activo, evaluacion)
        if tiempo_restante > 0:
            continuar_evaluacion = True
            resultado_activo.tiempo_restante = tiempo_restante
            resultado_activo.ultima_actividad = timezone.now()
            resultado_activo.save(update_fields=['tiempo_restante', 'ultima_actividad'])
        else:
            # El intento caducado se cierra antes de permitir cualquier nuevo ingreso.
            respuestas_guardadas = resultado_activo.respuestas_guardadas or {}
            preguntas_mostradas = evaluacion.get_preguntas_para_estudiante(
                participante.id, resultado_activo.numero_intento
            )
            puntos_ganados = 0
            puntos_posibles_totales = 0
            for pregunta in preguntas_mostradas:
                peso = getattr(pregunta, 'puntos', 1) or 1
                puntos_posibles_totales += peso
                pregunta_key = f'pregunta_{pregunta.id}'
                respuesta_id = respuestas_guardadas.get(pregunta_key)
                if respuesta_id and pregunta.opciones.filter(id=respuesta_id, is_correct=True).exists():
                    puntos_ganados += peso

            puntaje_ponderado = round(
                (puntos_ganados / puntos_posibles_totales) * 10, 3
            ) if puntos_posibles_totales else 0
            snapshot_data = generar_snapshot_respuestas(preguntas_mostradas, respuestas_guardadas)
            resultado_activo.puntos_obtenidos = puntaje_ponderado
            resultado_activo.puntos_totales = 10
            resultado_activo.tiempo_utilizado = evaluacion.duration_minutes * 60
            resultado_activo.fecha_fin = timezone.now()
            resultado_activo.completada = True
            resultado_activo.respuestas_guardadas = snapshot_data
            resultado_activo.tiempo_restante = 0
            resultado_activo.save()
            messages.warning(request, 'Se acabó el tiempo para esta evaluación.')
            return redirect('quizzes:quiz')
    
    # Si no hay intento en progreso, verificar ventana de acceso solo para nuevos ingresos
    if not continuar_evaluacion:
        if evaluacion.is_not_started():
            messages.warning(request, 'Esta evaluación aún no ha comenzado.')
            return redirect('quizzes:quiz')
        
        if evaluacion.is_finished():
            messages.warning(request, 'Esta evaluación ya ha finalizado.')
            return redirect('quizzes:quiz')
        
        if not evaluacion.is_available():
            messages.warning(request, 'Esta evaluación no está disponible en este momento.')
            return redirect('quizzes:quiz')
        
    if request.method == 'POST':
        # Verificar si la evaluación fue finalizada por cambios de pestaña
        finalizada_por_cambios_pestana = request.POST.get('finalizada_por_cambios_pestana') == 'true'
        numero_intento_actual = resultado_activo.numero_intento if resultado_activo else 1
        preguntas_mostradas = evaluacion.get_preguntas_para_estudiante(participante.id, numero_intento_actual)
        
        if finalizada_por_cambios_pestana:
            # Finalización por cambios de pestaña - asignar puntaje de 0
            score = 0
            puntos_obtenidos = 0
            puntos_totales = 10
            percentage = 0
            
            # Calcular tiempo utilizado
            tiempo_utilizado = 0
            if resultado_activo:
                tiempo_total = evaluacion.duration_minutes * 60  # en segundos
                tiempo_restante = int(request.POST.get('tiempo_restante', 0))
                tiempo_utilizado = tiempo_total - tiempo_restante
            
            # Obtener respuestas guardadas hasta el momento de la finalización
            respuestas_finales = {}
            if resultado_activo and resultado_activo.respuestas_guardadas:
                respuestas_finales = resultado_activo.respuestas_guardadas
            
        else:
            # Procesar envío normal de evaluación con nuevo sistema de puntuación
            score = 0
            puntos_ganados = 0
            puntos_posibles_totales = 0
            respuestas_finales = {}
            for pregunta in preguntas_mostradas:
                selected = request.POST.get(f'pregunta_{pregunta.id}')
                respuestas_finales[f'pregunta_{pregunta.id}'] = selected
                peso = getattr(pregunta, 'puntos', 1) or 1
                puntos_posibles_totales += peso
                
                if selected:
                    if pregunta.opciones.filter(id=selected, is_correct=True).exists():
                        puntos_ganados += peso
                   
            puntaje_ponderado = (puntos_ganados / puntos_posibles_totales) * 10 if puntos_posibles_totales > 0 else 0
            puntos_obtenidos = round(max(0, puntaje_ponderado), 3)
            puntos_totales = 10
            percentage = round((puntos_obtenidos / puntos_totales) * 100, 1)
            
            # Calcular tiempo utilizado en segundos
            if resultado_activo and resultado_activo.fecha_inicio:
                tiempo_utilizado = int((timezone.now() - resultado_activo.fecha_inicio).total_seconds())
            else:
                tiempo_total = evaluacion.duration_minutes * 60
                tiempo_restante = int(request.POST.get('tiempo_restante', 0))
                tiempo_utilizado = max(0, tiempo_total - tiempo_restante)
        
        # Generar snapshot congelado para inmutabilidad del examen entregado
        snapshot_data = generar_snapshot_respuestas(preguntas_mostradas, respuestas_finales)
        
        # Guardar resultado en la base de datos con nuevo sistema de puntuación
        if resultado_activo:
            resultado_activo.puntos_obtenidos = puntos_obtenidos
            resultado_activo.puntos_totales = puntos_totales
            resultado_activo.tiempo_utilizado = tiempo_utilizado
            resultado_activo.fecha_fin = timezone.now()
            resultado_activo.completada = True
            resultado_activo.respuestas_guardadas = snapshot_data
            resultado_activo.tiempo_restante = 0
            resultado_activo.save()
            
            if finalizada_por_cambios_pestana:
                resultado_activo.agregar_alerta(
                    'finalizacion_automatica',
                    'Evaluación finalizada automáticamente por exceso de cambios de pestaña (4/4)',
                    severidad='alta'
                )
                resultado_activo.finalizado_por_admin = request.user
                resultado_activo.motivo_finalizacion = 'Evaluación finalizada automáticamente por exceder el límite de cambios de pestaña (4/4)'
                resultado_activo.fecha_finalizacion_admin = timezone.now()
                resultado_activo.save()
                
        else:
            # Obtener el siguiente número de intento
            siguiente_intento = ResultadoEvaluacion.get_siguiente_numero_intento(evaluacion, participante)
            
            nuevo_resultado = ResultadoEvaluacion.objects.create(
                evaluacion=evaluacion,
                participante=participante,
                numero_intento=siguiente_intento,
                puntos_obtenidos=puntos_obtenidos,
                puntos_totales=puntos_totales,
                tiempo_utilizado=tiempo_utilizado,
                fecha_fin=timezone.now(),
                completada=True,
                respuestas_guardadas=snapshot_data,
                tiempo_restante=0
            )
            
            if finalizada_por_cambios_pestana:
                nuevo_resultado.agregar_alerta(
                    'finalizacion_automatica',
                    'Evaluación finalizada automáticamente por exceso de cambios de pestaña (4/4)',
                    severidad='alta'
                )
                nuevo_resultado.finalizado_por_admin = request.user
                nuevo_resultado.motivo_finalizacion = 'Evaluación finalizada automáticamente por exceder el límite de cambios de pestaña (4/4)'
                nuevo_resultado.fecha_finalizacion_admin = timezone.now()
                nuevo_resultado.save()
        
        # Agregar un mensaje específico si fue finalizada por cambios de pestaña
        if finalizada_por_cambios_pestana:
            messages.warning(request, 'Tu evaluación fue finalizada automáticamente por exceder el límite de cambios de pestaña permitidos (4/4). Tu puntaje es 0/10.')

        return render(request, 'quizzes/result.html', {
            'evaluacion': evaluacion, 
            'resultado': resultado_activo if resultado_activo else nuevo_resultado,
            'score': score,
            'total_questions': len(preguntas_mostradas) if not finalizada_por_cambios_pestana else 0,
            'percentage': percentage,
            'finalizada_por_cambios_pestana': finalizada_por_cambios_pestana,
        })
    
    # Obtener preguntas para este estudiante específico
    # Si hay un resultado activo, usar su número de intento, si no, obtener el siguiente
    if resultado_activo:
        numero_intento = resultado_activo.numero_intento
    else:
        numero_intento = ResultadoEvaluacion.get_siguiente_numero_intento(evaluacion, participante)
    
    preguntas_mostradas = evaluacion.get_preguntas_para_estudiante(participante.id, numero_intento)
    
    if not preguntas_mostradas:
        messages.error(request, 'Esta evaluación no tiene preguntas configuradas.')
        return redirect('quizzes:quiz')
    
    # Si no hay intento en progreso, crear uno nuevo
    if not continuar_evaluacion:
        tiempo_total = evaluacion.duration_minutes * 60  # en segundos
        
        # Obtener el siguiente número de intento
        siguiente_intento = ResultadoEvaluacion.get_siguiente_numero_intento(evaluacion, participante)
        
        resultado_activo = ResultadoEvaluacion.objects.create(
            evaluacion=evaluacion,
            participante=participante,
            numero_intento=siguiente_intento,
            fecha_inicio=timezone.now(),
            tiempo_restante=tiempo_total
        )
        
    context = {
        'evaluacion': evaluacion,
        'preguntas': preguntas_mostradas,
        'resultado': resultado_activo,
        'tiempo_total': evaluacion.duration_minutes * 60,  # en segundos
        'continuar_evaluacion': continuar_evaluacion
    }
    
    return render(request, 'quizzes/take_quiz.html', context)

@login_required
def guardar_respuesta_automatica(request, pk):
    """
    Vista para guardado automático de respuestas
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        evaluacion = get_object_or_404(Evaluacion, pk=pk)
        participante = Participantes.objects.get(user=request.user)
        
        # Verificar que el participante esté autorizado
        participantes_autorizados = evaluacion.get_participantes_autorizados()
        if participante not in participantes_autorizados:
            return JsonResponse({'success': False, 'error': 'No autorizado'})
        
        # Verificar si la evaluación fue finalizada administrativamente.
        resultado_finalizado_admin = ResultadoEvaluacion.objects.filter(
            evaluacion=evaluacion,
            participante=participante,
            completada=True,
            finalizado_por_admin__isnull=False
        ).order_by('-numero_intento').first()
        
        if resultado_finalizado_admin:
            return JsonResponse({
                'success': False, 
                'error': 'Evaluación finalizada administrativamente',
                'redirect': True
            })
        
        # Obtener o crear resultado activo (no completado)
        resultado, created = ResultadoEvaluacion.objects.get_or_create(
            evaluacion=evaluacion,
            participante=participante,
            completada=False,
            defaults={
                'fecha_inicio': timezone.now(),
                'tiempo_restante': evaluacion.duration_minutes * 60
            }
        )
        
        # Verificación adicional: No permitir guardado si ya está completada
        if resultado.completada:
            return JsonResponse({
                'success': False, 
                'error': 'Esta evaluación ya ha sido completada',
                'redirect': True
            })
        
        # Actualizar respuestas guardadas
        import json
        data = json.loads(request.body.decode('utf-8'))
        respuestas = data.get('respuestas', {})
        
        # Asegurar que respuestas_guardadas esté inicializado
        if resultado.respuestas_guardadas is None:
            resultado.respuestas_guardadas = {}
        
        # Hacer una copia del diccionario para evitar problemas de referencia
        respuestas_actualizadas = resultado.respuestas_guardadas.copy()
        respuestas_actualizadas.update(respuestas)
        
        resultado.respuestas_guardadas = respuestas_actualizadas
        resultado.tiempo_restante = calcular_tiempo_restante_servidor(resultado, evaluacion)
        resultado.ultima_actividad = timezone.now()
        resultado.save()
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def registrar_cambio_pestana(request, pk):
    """
    Vista para registrar cambios de pestaña durante la evaluación
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        evaluacion = get_object_or_404(Evaluacion, pk=pk)
        participante = Participantes.objects.get(user=request.user)
        
        # Verificar que el participante esté autorizado
        participantes_autorizados = evaluacion.get_participantes_autorizados()
        if participante not in participantes_autorizados:
            return JsonResponse({'success': False, 'error': 'No autorizado'})
        
        # Verificar si la evaluación fue finalizada administrativamente.
        resultado_finalizado_admin = ResultadoEvaluacion.objects.filter(
            evaluacion=evaluacion,
            participante=participante,
            completada=True,
            finalizado_por_admin__isnull=False
        ).order_by('-numero_intento').first()
        
        if resultado_finalizado_admin:
            return JsonResponse({
                'success': False, 
                'error': 'Evaluación finalizada administrativamente',
                'redirect': True
            })
        
        # Obtener resultado actual
        resultado = ResultadoEvaluacion.objects.filter(
            evaluacion=evaluacion,
            participante=participante,
            completada=False
        ).first()
        
        if not resultado:
            return JsonResponse({
                'success': False, 
                'error': 'No se encontró una evaluación activa'
            })

        # El navegador sólo informa el evento: el contador y tiempo se calculan aquí.
        with transaction.atomic():
            resultado = ResultadoEvaluacion.objects.select_for_update().get(pk=resultado.pk)
            resultado.cambios_pestana = (resultado.cambios_pestana or 0) + 1
            resultado.tiempo_restante = calcular_tiempo_restante_servidor(resultado, evaluacion)
            resultado.ultima_actividad = timezone.now()
            resultado.save(update_fields=['cambios_pestana', 'tiempo_restante', 'ultima_actividad'])

        cambios_pestana = resultado.cambios_pestana
        tiempo_restante = resultado.tiempo_restante
        alerta_texto = f"Cambio de pestaña #{cambios_pestana} - Tiempo restante: {tiempo_restante//60}:{tiempo_restante%60:02d}"
        resultado.agregar_alerta('cambio_pestana', alerta_texto, severidad='media')
        
        return JsonResponse({
            'success': True,
            'cambios_pestana': cambios_pestana,
            'mensaje': f'Cambio de pestaña #{cambios_pestana} registrado'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def registrar_evento_auditoria(request, pk):
    """Registra señales del navegador para revisión, sin penalizar al estudiante."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    try:
        evaluacion = get_object_or_404(Evaluacion, pk=pk)
        participante = Participantes.objects.get(user=request.user)
        if participante not in evaluacion.get_participantes_autorizados():
            return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)

        resultado = ResultadoEvaluacion.objects.filter(
            evaluacion=evaluacion, participante=participante, completada=False
        ).first()
        if not resultado:
            return JsonResponse({'success': False, 'error': 'No se encontró una evaluación activa'}, status=404)

        data = json.loads(request.body.decode('utf-8'))
        eventos_auditoria = {
            'perdida_foco': 'La ventana de la evaluación perdió el foco. Se registró para revisión, sin descontar cambios de pestaña.',
            'salida_pantalla_completa': 'El estudiante salió de pantalla completa. Se registró para revisión, sin descontar cambios de pestaña.',
        }
        tipo_evento = data.get('tipo_evento')
        if tipo_evento not in eventos_auditoria:
            return JsonResponse({'success': False, 'error': 'Tipo de evento no válido'}, status=400)

        resultado.agregar_alerta(tipo_evento, eventos_auditoria[tipo_evento], severidad='baja')
        return JsonResponse({'success': True, 'auditoria': True})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'}, status=400)
    except Participantes.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Participante no encontrado'}, status=403)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def verificar_estado_evaluacion(request, pk):
    """
    Endpoint para verificar si la evaluación fue finalizada administrativamente
    y obtener el estado actualizado de cambios de pestañas
    """
    try:
        evaluacion = get_object_or_404(Evaluacion, pk=pk)
        participante = Participantes.objects.get(user=request.user)
        
        # Obtener resultado activo (no completado)
        resultado_activo = ResultadoEvaluacion.objects.filter(
            evaluacion=evaluacion,
            participante=participante,
            completada=False
        ).first()
        
        # Verificar si la evaluación fue finalizada administrativamente.
        resultado_finalizado_admin = ResultadoEvaluacion.objects.filter(
            evaluacion=evaluacion,
            participante=participante,
            completada=True,
            finalizado_por_admin__isnull=False
        ).order_by('-numero_intento').first()
        
        response_data = {'finalizada_admin': False}
        
        if resultado_finalizado_admin:
            response_data.update({
                'finalizada_admin': True,
                'motivo': resultado_finalizado_admin.motivo_finalizacion,
                'admin': resultado_finalizado_admin.finalizado_por_admin.get_full_name() or resultado_finalizado_admin.finalizado_por_admin.username
            })
        
        # Incluir información actualizada de cambios de pestañas si hay resultado activo
        if resultado_activo:
            cambios_actuales = getattr(resultado_activo, 'cambios_pestana', 0) or 0
            response_data.update({
                'cambios_pestana_actuales': cambios_actuales,
                'cambios_pestana_maximo': 4
            })
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({'finalizada_admin': False, 'error': str(e)})

@login_required
def obtener_progreso_evaluacion(request, pk):
    """
    Vista para obtener progreso guardado de una evaluación
    """
    try:
        evaluacion = get_object_or_404(Evaluacion, pk=pk)
        participante = Participantes.objects.get(user=request.user)
        
        resultado = ResultadoEvaluacion.objects.filter(
            evaluacion=evaluacion,
            participante=participante,
            completada=False
        ).first()
        
        if resultado:
            return JsonResponse({
                'success': True,
                'respuestas': resultado.respuestas_guardadas,
                'tiempo_restante': resultado.tiempo_restante,
                'ultima_actividad': resultado.ultima_actividad.isoformat()
            })
        else:
            return JsonResponse({
                'success': True,
                'respuestas': {},
                'tiempo_restante': evaluacion.duration_minutes * 60,
                'ultima_actividad': None
            })
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def dashboard(request):
    user = request.user
    context = {}
    is_global, scope_carrera, scope_concurso = get_user_scope(user)
    context['is_global'] = is_global
    context['scope_carrera'] = scope_carrera
    context['scope_concurso'] = scope_concurso

    # Determinar el tipo de usuario
    if user.is_superuser:
        context['role'] = 'superadmin'
        context['has_full_access'] = True
    elif AdminProfile.objects.filter(user=user).exists():
        admin_profile = AdminProfile.objects.select_related('carrera', 'carrera__facultad').get(user=user)
        context['role'] = 'admin'
        context['has_full_access'] = admin_profile.acceso_total
        context['admin_profile'] = admin_profile
    elif Participantes.objects.filter(user=user).exists():
        context['role'] = 'participant'
        context['has_full_access'] = False
    else:
        context['role'] = 'unknown'
        context['has_full_access'] = False
    return render(request, 'quizzes/dashboard.html', context)

# Gestión de participantes
@login_required
@admin_required
def manage_participants(request):
    user = request.user

    # Eliminar participante
    delete_id = request.GET.get('delete_id')
    if delete_id:
        participante = Participantes.objects.filter(id=delete_id).first()
        if participante:
            participante.user.delete()  # Elimina el usuario relacionado al participante
            participante.delete()  # Elimina el perfil de participante
        return redirect('quizzes:manage_participants')

    # Agregar participante
    if request.method == 'POST' and request.POST.get('add_participant'):
        cedula = request.POST.get('cedula')
        NombresCompletos = request.POST.get('NombresCompletos')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        edad = request.POST.get('edad')
        
        # Validar cédula y teléfono
        is_valid_cedula, error_cedula = validate_cedula_format(cedula)
        is_valid_phone, error_phone = validate_phone_format(phone)
        
        if not is_valid_cedula:
            messages.error(request, f"Cédula: {error_cedula}")
            return redirect('quizzes:manage_participants')
        
        if not is_valid_phone:
            messages.error(request, f"Teléfono: {error_phone}")
            return redirect('quizzes:manage_participants')
        
        # Verificar si la cédula ya existe (global)
        if Participantes.objects.filter(cedula=cedula).exists():
            messages.error(request, f"La cédula {cedula} ya está registrada por otro participante.")
            return redirect('quizzes:manage_participants')
        
        # Convertir edad vacía a None
        if edad == '':
            edad = None
        elif edad:
            try:
                edad = int(edad)
            except ValueError:
                messages.error(request, "La edad debe ser un número válido.")
                return redirect('quizzes:manage_participants')
        
        # Crear el participante con manejo de validaciones
        try:
            is_global, scope_carrera, scope_concurso = get_user_scope(request)
            if not scope_concurso:
                messages.warning(request, '⚠️ Debes seleccionar un Concurso activo específico en la barra superior (Navbar) antes de registrar un participante.')
                return redirect('quizzes:manage_participants')

            participante, password = Participantes.create_participant(
                cedula, NombresCompletos, email, phone, edad,
                concurso=scope_concurso, carrera=scope_carrera
            )
            try:
                subject = 'Credenciales de Acceso - Sistema Olymp'
                system_name = 'Sistema Olymp'
                plain_message, html_message = generate_email_messages(
                    subject=subject,
                    nombre=NombresCompletos,
                    system_name=system_name,
                    username=cedula,
                    nueva_password=password,
                    email_type='credentials'
                )
                send_mail(
                    subject,
                    plain_message,
                    settings.EMAIL_HOST_USER,
                    [participante.email],
                    fail_silently=False,
                    html_message=html_message
                )
                messages.success(request, f"Participante {NombresCompletos} creado correctamente y correo de credenciales enviado a {participante.email}.")
            except Exception as email_err:
                messages.warning(request, f"Participante {NombresCompletos} creado correctamente, pero no se pudo enviar el correo de credenciales: {str(email_err)}")
            return redirect('quizzes:manage_participants')
        except ValidationError as e:
            error_message = extract_validation_error_message(e)
            messages.error(request, f"{error_message}")
            return redirect('quizzes:manage_participants')
        except Exception as e:
            messages.error(request, f"Error inesperado al crear participante: {str(e)}")
            return redirect('quizzes:manage_participants')

    # Editar participante
    if request.method == 'POST' and request.POST.get('edit_id'):
        edit_id = request.POST.get('edit_id')
        cedula = request.POST.get('cedula')
        NombresCompletos = request.POST.get('NombresCompletos')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        edad = request.POST.get('edad')
        
        # Validar cédula y teléfono
        is_valid_cedula, error_cedula = validate_cedula_format(cedula)
        is_valid_phone, error_phone = validate_phone_format(phone)
        
        if not is_valid_cedula:
            messages.error(request, f"Cédula: {error_cedula}")
            return redirect('quizzes:manage_participants')
        
        if not is_valid_phone:
            messages.error(request, f"Teléfono: {error_phone}")
            return redirect('quizzes:manage_participants')
        
        try:
            participante = Participantes.objects.get(id=edit_id)
            user_obj = participante.user
            
            # Normalizar email
            email_normalized = email.lower().strip()
            
            # Verificar si la nueva cédula ya existe en otro participante
            if cedula != participante.cedula and Participantes.objects.filter(cedula=cedula).exists():
                messages.error(request, f"La cédula {cedula} ya está registrada por otro participante.")
                return redirect('quizzes:manage_participants')
            
            # Verificar si el nuevo email ya existe en otro participante
            if email_normalized != participante.email and Participantes.objects.filter(email__iexact=email_normalized).exists():
                messages.error(request, f"El correo {email} ya está registrado por otro participante.")
                return redirect('quizzes:manage_participants')
            
            # Verificar si el nuevo email está en uso por representantes
            if email_normalized != participante.email:
                if Representante.objects.filter(
                    models.Q(CorreoInstitucional__iexact=email_normalized) | 
                    models.Q(CorreoRepresentante__iexact=email_normalized)
                ).exists():
                    messages.error(request, f"El correo {email} ya está siendo usado por un representante.")
                    return redirect('quizzes:manage_participants')
            
            # Convertir edad vacía a None
            if edad == '':
                edad = None
            elif edad:
                try:
                    edad = int(edad)
                except ValueError:
                    messages.error(request, "La edad debe ser un número válido.")
                    return redirect('quizzes:manage_participants')
            
            try:
                # Usar transacción para asegurar consistencia
                from django.db import transaction
                with transaction.atomic():
                    # Mantener username igual a la cédula
                    user_obj.username = cedula
                    user_obj.first_name = NombresCompletos
                    user_obj.email = email.lower().strip()  # Normalizar correo
                    participante.cedula = cedula
                    participante.NombresCompletos = NombresCompletos
                    participante.email = email.lower().strip()  # Normalizar correo
                    participante.phone = phone
                    participante.edad = edad
                    
                    # Validar y guardar
                    user_obj.full_clean()
                    participante.full_clean()
                    user_obj.save()
                    participante.save()
                    
                messages.success(request, f"Participante {NombresCompletos} actualizado correctamente.")
                return redirect('quizzes:manage_participants')
                
            except ValidationError as e:
                error_message = extract_validation_error_message(e)
                messages.error(request, f"Error de validación: {error_message}")
                return redirect('quizzes:manage_participants')
            except IntegrityError as e:
                if 'email' in str(e).lower():
                    messages.error(request, 'El correo electrónico ya está en uso.')
                elif 'cedula' in str(e).lower():
                    messages.error(request, 'La cédula ya está registrada.')
                else:
                    messages.error(request, 'Error de integridad en la base de datos.')
                return redirect('quizzes:manage_participants')
            except Exception as e:
                messages.error(request, f'Error inesperado al actualizar participante: {str(e)}')
                return redirect('quizzes:manage_participants')
        except Participantes.DoesNotExist:
            messages.error(request, 'Participante no encontrado.')
            return redirect('quizzes:manage_participants')

    # Búsqueda de participantes
    search_query = request.GET.get('search', '').strip()
    participantes = filter_queryset_by_scope(
        Participantes.objects.select_related('user', 'carrera', 'concurso').order_by('-id'),
        request
    )
    
    if search_query:
        participantes = participantes.filter(
            Q(NombresCompletos__icontains=search_query) |
            Q(cedula__icontains=search_query)
        )
    
    # Paginación para participantes
    paginator = Paginator(participantes, 7)  # 7 elementos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'quizzes/manage_participants.html', {
        'participantes': page_obj,
        'page_obj': page_obj,
        'search_query': search_query
    })



# Gestión de admins
@login_required
@full_access_required
def manage_admins(request):
    user = request.user

    # Eliminar admin
    delete_id = request.GET.get('delete_id')
    if delete_id:
        admin = AdminProfile.objects.filter(id=delete_id).first()
        if admin:
            if admin.user == user:
                messages.error(request, 'No puedes eliminar tu propia cuenta de administrador activa.')
                return redirect('quizzes:manage_admins')
                
            # Un admin con acceso total (no superuser) NO puede eliminar a otro de su mismo estatus
            if not user.is_superuser and admin.acceso_total:
                messages.error(request, 'Un Administrador con Acceso Total no puede eliminar a otro Administrador de su mismo estatus.')
                return redirect('quizzes:manage_admins')

            # Registrar log de auditoría
            AuditLog.registrar_accion(
                usuario_ejecutor=user,
                accion='ELIMINACION_ADMINISTRADOR',
                detalles=f'Eliminó al administrador "{admin.user.username}" ({admin.user.get_full_name() or "Sin nombre"}) - Email: {admin.user.email} - Carrera: {admin.carrera.nombre if admin.carrera else "Global"}',
                request=request
            )

            admin.user.delete()
            admin.delete()
            messages.success(request, 'Administrador eliminado correctamente.')
        return redirect('quizzes:manage_admins')

    # Agregar admin
    if request.method == 'POST' and request.POST.get('add_admin'):
        username = request.POST.get('username')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        
        # Validar campos obligatorios
        if not username or not first_name or not email:
            messages.error(request, "Todos los campos son obligatorios.")
            return redirect('quizzes:manage_admins')
        
        # Normalizar email
        email_normalized = email.lower().strip()
        
        # Verificar si el username ya existe
        if User.objects.filter(username=username).exists():
            messages.error(request, f"El nombre de usuario '{username}' ya está registrado por otro usuario.")
            return redirect('quizzes:manage_admins')
        
        # Verificar si el email ya existe
        if User.objects.filter(email__iexact=email_normalized).exists():
            messages.error(request, f"El correo electrónico '{email}' ya está registrado por otro usuario.")
            return redirect('quizzes:manage_admins')
        
        # Verificar conflictos con participantes y representantes
        if Participantes.objects.filter(email__iexact=email_normalized).exists():
            messages.error(request, f"El correo '{email}' ya está siendo usado por un participante.")
            return redirect('quizzes:manage_admins')
        
        if Representante.objects.filter(
            models.Q(CorreoInstitucional__iexact=email_normalized) | 
            models.Q(CorreoRepresentante__iexact=email_normalized)
        ).exists():
            messages.error(request, f"El correo '{email}' ya está siendo usado por un representante.")
            return redirect('quizzes:manage_admins')
        
        try:
            password = get_random_string(length=8)
            carrera_id = request.POST.get('carrera_id')
            if not user.is_superuser:
                carrera_id = user.adminprofile.carrera_id if hasattr(user, 'adminprofile') and user.adminprofile.carrera_id else None
                
            if not carrera_id:
                messages.error(request, "Debe seleccionar una carrera obligatoriamente para el nuevo administrador.")
                return redirect('quizzes:manage_admins')

            new_user = User.objects.create_user(
                username=username, 
                password=password, 
                first_name=first_name, 
                last_name=last_name, 
                email=email_normalized
            )
            admin_profile = AdminProfile.objects.create(
                user=new_user, 
                created_by=user,
                carrera_id=carrera_id
            )
            try:
                subject = 'Credenciales de Acceso - Panel de Administración Olymp'
                system_name = 'Panel de Administración Olymp'
                full_name = f"{first_name} {last_name}".strip()
                plain_message, html_message = generate_email_messages(
                    subject=subject,
                    nombre=full_name,
                    system_name=system_name,
                    username=username,
                    nueva_password=password,
                    email_type='credentials'
                )
                send_mail(
                    subject,
                    plain_message,
                    settings.EMAIL_HOST_USER,
                    [email_normalized],
                    fail_silently=False,
                    html_message=html_message
                )
                messages.success(request, f"Administrador {first_name} {last_name} creado correctamente y correo de credenciales enviado a {email_normalized}.")
            except Exception as email_err:
                messages.warning(request, f"Administrador {first_name} {last_name} creado correctamente, pero no se pudo enviar el correo de credenciales: {str(email_err)}")
            return redirect('quizzes:manage_admins')
        except IntegrityError as e:
            if 'username' in str(e).lower():
                messages.error(request, 'El nombre de usuario ya está en uso.')
            elif 'email' in str(e).lower():
                messages.error(request, 'El correo electrónico ya está en uso.')
            else:
                messages.error(request, 'Error de integridad en la base de datos.')
            return redirect('quizzes:manage_admins')
        except Exception as e:
            messages.error(request, f'Error inesperado al crear administrador: {str(e)}')
            return redirect('quizzes:manage_admins')

    # Editar admin
    if request.method == 'POST' and request.POST.get('edit_id'):
        edit_id = request.POST.get('edit_id')
        username = request.POST.get('username')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        carrera_id = request.POST.get('carrera_id')
        
        if not user.is_superuser:
            carrera_id = user.adminprofile.carrera_id if hasattr(user, 'adminprofile') and user.adminprofile.carrera_id else None
            
        if not carrera_id:
            messages.error(request, "Debe seleccionar una carrera obligatoriamente.")
            return redirect('quizzes:manage_admins')
        
        # Validar campos obligatorios
        if not username or not first_name or not email:
            messages.error(request, "Todos los campos son obligatorios.")
            return redirect('quizzes:manage_admins')
        
        try:
            admin = AdminProfile.objects.select_related('user').get(id=edit_id)
            user_obj = admin.user

            # Verificar jerarquía de edición
            if not user.is_superuser:
                # Un admin secundario no puede editar a otro admin de acceso total ni a sí mismo desde la tabla
                if admin.acceso_total or user_obj == user:
                    messages.error(request, 'No tienes permisos para editar a este administrador.')
                    return redirect('quizzes:manage_admins')
                # Preservar username e imponer la carrera del usuario secundario
                username = user_obj.username
                if hasattr(user, 'adminprofile') and user.adminprofile.carrera_id:
                    carrera_id = user.adminprofile.carrera_id
                else:
                    carrera_id = admin.carrera_id
            
            # Normalizar email
            email_normalized = email.lower().strip()
            
            # Verificar si el nuevo username ya existe en otro usuario (solo si el superuser intentó cambiarlo)
            if username != user_obj.username and User.objects.filter(username=username).exists():
                messages.error(request, f"El nombre de usuario '{username}' ya está registrado.")
                return redirect('quizzes:manage_admins')
            
            # Verificar si el nuevo email ya existe en otro usuario
            if email_normalized != user_obj.email and User.objects.filter(email__iexact=email_normalized).exclude(id=user_obj.id).exists():
                messages.error(request, f"El correo electrónico '{email}' ya está registrado.")
                return redirect('quizzes:manage_admins')
            
            # Verificar conflictos con participantes y representantes
            if email_normalized != user_obj.email:
                if Participantes.objects.filter(email__iexact=email_normalized).exists():
                    messages.error(request, f"El correo '{email}' ya está siendo usado por un participante.")
                    return redirect('quizzes:manage_admins')
                
                if Representante.objects.filter(
                    models.Q(CorreoInstitucional__iexact=email_normalized) | 
                    models.Q(CorreoRepresentante__iexact=email_normalized)
                ).exists():
                    messages.error(request, f"El correo '{email}' ya está siendo usado por un representante.")
                    return redirect('quizzes:manage_admins')
            
            # Usar transacción para asegurar consistencia
            from django.db import transaction
            with transaction.atomic():
                user_obj.username = username
                user_obj.first_name = first_name
                user_obj.last_name = last_name
                user_obj.email = email_normalized
                
                # Validar y guardar
                user_obj.full_clean()
                user_obj.save()
                
                if carrera_id:
                    admin.carrera_id = carrera_id
                admin.save()
            
            # Registrar log de auditoría
            AuditLog.registrar_accion(
                usuario_ejecutor=user,
                accion='EDICION_ADMINISTRADOR',
                detalles=f'Editó al administrador "{user_obj.username}" (Nombres: {first_name} {last_name}, Email: {email_normalized}) - Carrera: {admin.carrera.nombre if admin.carrera else "Global"}',
                request=request
            )

            messages.success(request, f"Administrador {first_name} {last_name} actualizado correctamente.")
            return redirect('quizzes:manage_admins')
            
        except AdminProfile.DoesNotExist:
            messages.error(request, 'Administrador no encontrado.')
            return redirect('quizzes:manage_admins')
        except ValidationError as e:
            error_message = extract_validation_error_message(e)
            messages.error(request, f"Error de validación: {error_message}")
            return redirect('quizzes:manage_admins')
        except IntegrityError as e:
            if 'username' in str(e).lower():
                messages.error(request, 'El nombre de usuario ya está en uso.')
            elif 'email' in str(e).lower():
                messages.error(request, 'El correo electrónico ya está en uso.')
            else:
                messages.error(request, 'Error de integridad en la base de datos.')
            return redirect('quizzes:manage_admins')
        except Exception as e:
            messages.error(request, f'Error inesperado al actualizar administrador: {str(e)}')
            return redirect('quizzes:manage_admins')

    from .models import Carrera
    admins = filter_queryset_by_scope(
        AdminProfile.objects.select_related('user', 'carrera', 'carrera__facultad').all(),
        request
    )
    carreras = Carrera.objects.select_related('facultad').filter(activa=True)
    return render(request, 'quizzes/manage_admins.html', {'admins': admins, 'carreras': carreras})


@login_required
@superuser_required
def manage_admin_permissions(request):
    user = request.user

    # Cambiar acceso total de un admin
    if request.method == 'POST':
        admin_id = request.POST.get('admin_id')
        acceso_total = request.POST.get('acceso_total') == 'on'
        
        try:
            admin_profile = AdminProfile.objects.get(id=admin_id)
            admin_profile.acceso_total = acceso_total
            admin_profile.save()
            
            status = "habilitado" if acceso_total else "deshabilitado"
            messages.success(request, f'Acceso total {status} para {admin_profile.user.get_full_name()}')
        except AdminProfile.DoesNotExist:
            messages.error(request, 'Administrador no encontrado.')
        
        return redirect('quizzes:manage_admin_permissions')

    # Obtener todos los admins (excluyendo superadmins)
    admins = AdminProfile.objects.select_related('user').all()
    return render(request, 'quizzes/manage_admin_permissions.html', {'admins': admins})


# Helper para control de acceso a gestión de representantes y grupos
def can_manage_representantes(user):
    return user.is_superuser or (hasattr(user, 'adminprofile'))

# Funciones de validación
def validate_cedula_format(cedula):
    """Valida formato de cédula: exactamente 10 dígitos numéricos"""
    if not cedula or not re.match(r'^\d{10}$', cedula):
        return False, "La cédula debe tener exactamente 10 dígitos numéricos."
    return True, ""

def validate_phone_format(phone):
    """Valida formato de teléfono: exactamente 10 dígitos numéricos"""
    if phone and not re.match(r'^\d{10}$', phone):
        return False, "El teléfono debe tener exactamente 10 dígitos numéricos."
    return True, ""

def extract_validation_error_message(validation_error):
    """
    Extrae el mensaje de error de un ValidationError de manera segura
    
    Args:
        validation_error: La excepción ValidationError
    
    Returns:
        str: El mensaje de error como cadena simple
    """
    if hasattr(validation_error, 'message_dict'):
        # Si es un diccionario de errores, tomar el primer mensaje
        if validation_error.message_dict:
            first_field = list(validation_error.message_dict.keys())[0]
            first_messages = validation_error.message_dict[first_field]
            if isinstance(first_messages, list) and first_messages:
                return first_messages[0]
            else:
                return str(first_messages)
        else:
            return str(validation_error)
    elif hasattr(validation_error, 'messages') and validation_error.messages:
        # Si es una lista de mensajes, tomar el primero
        if isinstance(validation_error.messages, list) and validation_error.messages:
            return validation_error.messages[0]
        else:
            return str(validation_error.messages)
    else:
        # Si es un mensaje simple
        return str(validation_error)

# Vista para listar y registrar representantes
@login_required
@admin_required
def manage_representantes(request):

    # Eliminar representante
    delete_id = request.GET.get('delete_id')
    if delete_id:
        representante = Representante.objects.filter(id=delete_id).first()
        if representante:
            representante.delete()
            messages.success(request, 'Representante eliminado exitosamente.')
        return redirect('quizzes:manage_representantes')

    if request.method == 'POST':
        # Registrar nuevo representante
        if request.POST.get('add_representante'):
            data = request.POST
            
            # Validar campos obligatorios
            required_fields = ['NombreColegio', 'DireccionColegio', 'TelefonoInstitucional', 
                             'CorreoInstitucional', 'NombresRepresentante', 'TelefonoRepresentante', 
                             'CorreoRepresentante']
            for field in required_fields:
                if not data.get(field):
                    messages.error(request, f"El campo {field} es obligatorio.")
                    return redirect('quizzes:manage_representantes')
            
            # Validar teléfonos
            telefono_inst = data.get('TelefonoInstitucional')
            telefono_rep = data.get('TelefonoRepresentante')
            
            is_valid_inst, error_inst = validate_phone_format(telefono_inst)
            is_valid_rep, error_rep = validate_phone_format(telefono_rep)
            
            if not is_valid_inst:
                messages.error(request, f"Teléfono Institucional: {error_inst}")
                return redirect('quizzes:manage_representantes')
            
            if not is_valid_rep:
                messages.error(request, f"Teléfono Representante: {error_rep}")
                return redirect('quizzes:manage_representantes')
            
            # Normalizar correos
            correo_inst = data.get('CorreoInstitucional').lower().strip()
            correo_rep = data.get('CorreoRepresentante').lower().strip()
            
            # Verificar conflictos con otros modelos usando la función de validación
            try:
                from .models import validate_email_across_all_models
                validate_email_across_all_models(correo_inst)
                validate_email_across_all_models(correo_rep)
            except ValidationError as e:
                error_message = extract_validation_error_message(e)
                messages.error(request, error_message)
                return redirect('quizzes:manage_representantes')
            
            try:
                is_global, scope_carrera, scope_concurso = get_user_scope(request)
                if not scope_concurso:
                    messages.warning(request, '⚠️ Debes seleccionar un Concurso activo específico en la barra superior (Navbar) antes de registrar un representante.')
                    return redirect('quizzes:manage_representantes')

                # Usar transacción para asegurar consistencia
                from django.db import transaction
                with transaction.atomic():
                    representante = Representante.objects.create(
                        NombreColegio=data.get('NombreColegio'),
                        DireccionColegio=data.get('DireccionColegio'),
                        TelefonoInstitucional=telefono_inst,
                        CorreoInstitucional=correo_inst,
                        NombresRepresentante=data.get('NombresRepresentante'),
                        TelefonoRepresentante=telefono_rep,
                        CorreoRepresentante=correo_rep,
                        concurso=scope_concurso,
                    )
                    # Validar el objeto creado
                    representante.full_clean()
                
                messages.success(request, 'Representante registrado exitosamente.')
                return redirect('quizzes:manage_representantes')
                
            except ValidationError as e:
                error_message = extract_validation_error_message(e)
                messages.error(request, f"Error de validación: {error_message}")
                return redirect('quizzes:manage_representantes')
            except IntegrityError as e:
                if 'correo' in str(e).lower():
                    messages.error(request, 'Uno de los correos ya está en uso.')
                else:
                    messages.error(request, 'Error de integridad en la base de datos.')
                return redirect('quizzes:manage_representantes')
            except Exception as e:
                messages.error(request, f'Error inesperado al crear representante: {str(e)}')
                return redirect('quizzes:manage_representantes')
        
        # Editar representante
        elif request.POST.get('edit_id'):
            edit_id = request.POST.get('edit_id')
            data = request.POST
            
            # Validar campos obligatorios
            required_fields = ['NombreColegio', 'DireccionColegio', 'TelefonoInstitucional', 
                             'CorreoInstitucional', 'NombresRepresentante', 'TelefonoRepresentante', 
                             'CorreoRepresentante']
            for field in required_fields:
                if not data.get(field):
                    messages.error(request, f"El campo {field} es obligatorio.")
                    return redirect('quizzes:manage_representantes')
            
            # Validar teléfonos
            telefono_inst = data.get('TelefonoInstitucional')
            telefono_rep = data.get('TelefonoRepresentante')
            
            is_valid_inst, error_inst = validate_phone_format(telefono_inst)
            is_valid_rep, error_rep = validate_phone_format(telefono_rep)
            
            if not is_valid_inst:
                messages.error(request, f"Teléfono Institucional: {error_inst}")
                return redirect('quizzes:manage_representantes')
            
            if not is_valid_rep:
                messages.error(request, f"Teléfono Representante: {error_rep}")
                return redirect('quizzes:manage_representantes')
            
            try:
                representante = Representante.objects.get(id=edit_id)
                
                # Normalizar correos
                correo_inst = data.get('CorreoInstitucional').lower().strip()
                correo_rep = data.get('CorreoRepresentante').lower().strip()
                
                # Verificar conflictos solo si los correos cambiaron
                if correo_inst != representante.CorreoInstitucional:
                    try:
                        from .models import validate_email_across_all_models
                        validate_email_across_all_models(correo_inst, exclude_representante_id=representante.id)
                    except ValidationError as e:
                        error_message = extract_validation_error_message(e)
                        messages.error(request, error_message)
                        return redirect('quizzes:manage_representantes')
                
                if correo_rep != representante.CorreoRepresentante:
                    try:
                        from .models import validate_email_across_all_models
                        validate_email_across_all_models(correo_rep, exclude_representante_id=representante.id)
                    except ValidationError as e:
                        error_message = extract_validation_error_message(e)
                        messages.error(request, error_message)
                        return redirect('quizzes:manage_representantes')
                
                # Usar transacción para asegurar consistencia
                from django.db import transaction
                with transaction.atomic():
                    representante.NombreColegio = data.get('NombreColegio')
                    representante.DireccionColegio = data.get('DireccionColegio')
                    representante.TelefonoInstitucional = telefono_inst
                    representante.CorreoInstitucional = correo_inst
                    representante.NombresRepresentante = data.get('NombresRepresentante')
                    representante.TelefonoRepresentante = telefono_rep
                    representante.CorreoRepresentante = correo_rep
                    
                    # Validar y guardar
                    representante.full_clean()
                    representante.save()
                
                messages.success(request, 'Representante actualizado exitosamente.')
                return redirect('quizzes:manage_representantes')
                
            except Representante.DoesNotExist:
                messages.error(request, 'Representante no encontrado.')
                return redirect('quizzes:manage_representantes')
            except ValidationError as e:
                error_message = extract_validation_error_message(e)
                messages.error(request, f"Error de validación: {error_message}")
                return redirect('quizzes:manage_representantes')
            except IntegrityError as e:
                if 'correo' in str(e).lower():
                    messages.error(request, 'Uno de los correos ya está en uso.')
                else:
                    messages.error(request, 'Error de integridad en la base de datos.')
                return redirect('quizzes:manage_representantes')
            except Exception as e:
                messages.error(request, f'Error inesperado al actualizar representante: {str(e)}')
                return redirect('quizzes:manage_representantes')

    representantes = filter_queryset_by_scope(
        Representante.objects.order_by('-id'),
        request
    )
    
    # Paginación
    paginator = Paginator(representantes, 10)  # 10 elementos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'quizzes/manage_representantes.html', {
        'representantes': page_obj,
        'page_obj': page_obj
    })

# Vista para listar y crear grupos de participantes
@login_required
@admin_required
def manage_grupos(request):

    # Eliminar grupo
    delete_id = request.GET.get('delete_id')
    if delete_id:
        grupo = GrupoParticipantes.objects.filter(id=delete_id).first()
        if grupo:
            grupo.delete()
            messages.success(request, 'Grupo eliminado exitosamente.')
        return redirect('quizzes:manage_grupos')

    if request.method == 'POST':
        # Crear nuevo grupo
        if request.POST.get('add_grupo'):
            name = request.POST.get('name')
            representante_id = request.POST.get('representante')
            participantes_ids = request.POST.getlist('participantes')
            
            # Validar que el representante no esté en otro grupo
            representante = Representante.objects.get(id=representante_id)
            if representante.grupos.exists():
                messages.error(request, f'El representante "{representante.NombresRepresentante}" ya está asignado a otro grupo.')
                return redirect('quizzes:manage_grupos')
            
            # Validar que los participantes no estén en otros grupos
            participantes_seleccionados = Participantes.objects.filter(id__in=participantes_ids)
            participantes_en_otros_grupos = participantes_seleccionados.filter(grupos__isnull=False)
            
            if participantes_en_otros_grupos.exists():
                nombres_problema = [p.NombresCompletos for p in participantes_en_otros_grupos]
                messages.error(request, f'Los siguientes participantes ya están en otros grupos: {", ".join(nombres_problema)}')
                return redirect('quizzes:manage_grupos')
            
            is_global, scope_carrera, scope_concurso = get_user_scope(request)
            if not scope_concurso:
                messages.warning(request, '⚠️ Debes seleccionar un Concurso activo específico en la barra superior (Navbar) antes de registrar un grupo.')
                return redirect('quizzes:manage_grupos')

            grupo = GrupoParticipantes.objects.create(name=name, representante=representante, concurso=scope_concurso)
            grupo.participantes.set(participantes_ids)
            messages.success(request, 'Grupo creado exitosamente.')
            return redirect('quizzes:manage_grupos')
        
        # Editar grupo
        elif request.POST.get('edit_id'):
            edit_id = request.POST.get('edit_id')
            name = request.POST.get('name')
            representante_id = request.POST.get('representante')
            participantes_ids = request.POST.getlist('participantes')
            
            # Validar que el representante no esté en otro grupo (excluyendo el grupo actual)
            representante = Representante.objects.get(id=representante_id)
            if representante.grupos.exclude(id=edit_id).exists():
                messages.error(request, f'El representante "{representante.NombresRepresentante}" ya está asignado a otro grupo.')
                return redirect('quizzes:manage_grupos')
            
            # Validar que los participantes no estén en otros grupos (excluyendo el grupo actual)
            participantes_seleccionados = Participantes.objects.filter(id__in=participantes_ids)
            participantes_en_otros_grupos = participantes_seleccionados.filter(
                grupos__isnull=False
            ).exclude(grupos=edit_id)
            
            if participantes_en_otros_grupos.exists():
                nombres_problema = [p.NombresCompletos for p in participantes_en_otros_grupos]
                messages.error(request, f'Los siguientes participantes ya están en otros grupos: {", ".join(nombres_problema)}')
                return redirect('quizzes:manage_grupos')
            
            grupo = GrupoParticipantes.objects.get(id=edit_id)
            grupo.name = name
            grupo.representante = representante
            grupo.participantes.set(participantes_ids)
            grupo.save()
            messages.success(request, 'Grupo actualizado exitosamente.')
            return redirect('quizzes:manage_grupos')

    grupos = filter_queryset_by_scope(
        GrupoParticipantes.objects.select_related('representante').prefetch_related('participantes'),
        request
    )
    
    # Obtener representantes disponibles (que no están en ningún grupo)
    representantes_disponibles = filter_queryset_by_scope(
        Representante.objects.filter(grupos__isnull=True),
        request
    )
    
    # Obtener todos los representantes para mostrar en modales de edición
    representantes_todos = filter_queryset_by_scope(
        Representante.objects.all(),
        request
    )
    
    # Obtener participantes disponibles (que no están en ningún grupo)
    participantes_disponibles = filter_queryset_by_scope(
        Participantes.objects.filter(grupos__isnull=True),
        request
    )
    
    # Obtener todos los participantes para mostrar en modales de edición
    participantes_todos = filter_queryset_by_scope(
        Participantes.objects.all(),
        request.user
    )
    
    # Paginación para grupos
    paginator = Paginator(grupos, 10)  # 10 elementos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'quizzes/manage_grupos.html', {
        'grupos': page_obj,
        'page_obj': page_obj,
        'representantes': representantes_disponibles,  # Solo disponibles para crear
        'representantes_todos': representantes_todos,  # Todos para editar
        'participantes': participantes_disponibles,  # Solo disponibles para crear
        'participantes_todos': participantes_todos   # Todos para editar
    })


# Vista para leer los headers del Excel con openpyxl (misma librería que procesa datos)
@login_required
def get_excel_headers(request):
    if not can_manage_representantes(request.user):
        return JsonResponse({'error': 'No tienes permisos para acceder a esta sección.'}, status=403)
    
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                return JsonResponse({'error': 'No se ha seleccionado ningún archivo.'}, status=400)
            
            workbook = load_workbook(excel_file, data_only=True)
            worksheet = workbook.active
            
            # Leer la primera fila (headers) incluyendo columnas vacías
            header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            if not header_row:
                return JsonResponse({'error': 'El archivo Excel está vacío.'}, status=400)
            
            headers = []
            for idx, cell_value in enumerate(header_row[0]):
                headers.append({
                    'index': idx + 1,  # 1-based, igual que en process_excel_participants
                    'name': str(cell_value).strip() if cell_value else None
                })
            
            return JsonResponse({'success': True, 'headers': headers})
            
        except Exception as e:
            return JsonResponse({'error': f'Error al leer el archivo: {str(e)}'}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


# Vista para procesar archivo Excel de participantes
@login_required
def process_excel_participants(request):
    if not can_manage_representantes(request.user):
        return JsonResponse({'error': 'No tienes permisos para acceder a esta sección.'}, status=403)
    
    is_global, scope_carrera, scope_concurso = get_user_scope(request)
    if not scope_concurso:
        return JsonResponse({'error': '⚠️ Debes seleccionar un Concurso activo específico en la barra superior (Navbar) antes de procesar un archivo Excel.'}, status=400)
    
    if request.method == 'POST':
        try:
            # Obtener el archivo y mapeo de columnas
            excel_file = request.FILES.get('excel_file')
            column_mapping = json.loads(request.POST.get('column_mapping', '{}'))
            
            if not excel_file:
                return JsonResponse({'error': 'No se ha seleccionado ningún archivo.'}, status=400)
            
            # Cargar el archivo Excel
            workbook = load_workbook(excel_file, data_only=True)
            worksheet = workbook.active
            
            # Obtener todas las filas (excluyendo la primera que son los headers)
            rows = list(worksheet.iter_rows(min_row=2, values_only=True))
            
            # Procesar cada fila
            processed_data = []
            errors = []
            
            for row_index, row in enumerate(rows, start=2):
                if not any(row):  # Fila vacía
                    continue
                    
                row_data = {}
                row_errors = []
                
                # Mapear columnas según el mapeo proporcionado
                for excel_col, model_field in column_mapping.items():
                    try:
                        col_index = int(excel_col) - 1  # Convertir a índice base 0
                        if col_index < len(row):
                            value = row[col_index]
                            if value is not None:
                                val_str = str(value).strip()
                                if val_str.endswith('.0'):
                                    val_str = val_str[:-2]
                                if model_field == 'cedula' and len(val_str) == 9 and val_str.isdigit():
                                    val_str = '0' + val_str
                                row_data[model_field] = val_str
                            else:
                                row_data[model_field] = ''
                        else:
                            row_data[model_field] = ''
                    except (ValueError, IndexError):
                        row_data[model_field] = ''
                
                # Validar datos requeridos
                if not row_data.get('cedula'):
                    row_errors.append('Cédula es requerida')
                elif not validate_cedula_format(row_data['cedula'])[0]:
                    row_errors.append(f"Cédula inválida: {row_data['cedula']}")
                
                if not row_data.get('NombresCompletos'):
                    row_errors.append('Nombres Completos es requerido')
                
                if not row_data.get('email'):
                    row_errors.append('Email es requerido')
                elif '@' not in row_data['email']:
                    row_errors.append('Email inválido')
                else:
                    # Validar que el correo no esté duplicado
                    email_normalized = row_data['email'].lower().strip()
                    
                    # Verificar si ya existe un participante con este correo
                    if Participantes.objects.filter(email__iexact=email_normalized).exists():
                        row_errors.append(f'Ya existe un participante con el correo "{row_data["email"]}"')
                    
                    # Verificar si el correo está siendo usado por un representante
                    if Representante.objects.filter(
                        Q(CorreoInstitucional__iexact=email_normalized) |
                        Q(CorreoRepresentante__iexact=email_normalized)
                    ).exists():
                        row_errors.append(f'El correo "{row_data["email"]}" ya está siendo usado por un representante')
                
                # Validar teléfono si está presente
                if row_data.get('phone') and not validate_phone_format(row_data['phone'])[0]:
                    row_errors.append(f"Teléfono inválido: {row_data['phone']}")
                
                # Validar edad si está presente
                if row_data.get('edad'):
                    try:
                        edad = int(float(row_data['edad']))
                        if edad < 0 or edad > 120:
                            row_errors.append('Edad debe estar entre 0 y 120')
                    except (ValueError, TypeError):
                        row_errors.append('Edad debe ser un número válido')
                
                if row_errors:
                    errors.append(f"Fila {row_index}: {', '.join(row_errors)}")
                else:
                    processed_data.append({
                        'row_index': row_index,
                        'data': row_data
                    })
            
            return JsonResponse({
                'success': True,
                'data': processed_data,
                'errors': errors,
                'total_rows': len(rows),
                'valid_rows': len(processed_data),
                'error_rows': len(errors)
            })
            
        except Exception as e:
            return JsonResponse({'error': f'Error al procesar el archivo: {str(e)}'}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

# Vista para guardar participantes desde Excel
@login_required
def save_excel_participants(request):
    if not can_manage_representantes(request.user):
        return JsonResponse({'error': 'No tienes permisos para acceder a esta sección.'}, status=403)

    is_global, scope_carrera, scope_concurso = get_user_scope(request)
    if not scope_concurso:
        return JsonResponse({'error': '⚠️ Debes seleccionar un Concurso activo específico en la barra superior (Navbar) antes de guardar participantes.'}, status=400)

    scope_carrera_final = scope_carrera or (scope_concurso.carrera if scope_concurso else None)

    if request.method == 'POST':
        try:
            participants_data = json.loads(request.POST.get('participants_data', '[]'))
            
            created_count = 0
            errors = []
            
            for participant_info in participants_data:
                try:
                    data = participant_info['data']
                    
                    cedula = str(data.get('cedula', '')).strip()
                    if cedula.endswith('.0'):
                        cedula = cedula[:-2]
                    if len(cedula) == 9 and cedula.isdigit():
                        cedula = '0' + cedula
                    
                    # Verificar si la cédula ya existe
                    if Participantes.objects.filter(cedula=cedula).exists():
                        errors.append(f"Cédula {cedula} ya existe")
                        continue
                    
                    # Crear el participante asignando concurso y carrera del scope activo
                    participante, password = Participantes.create_participant(
                        cedula=cedula,
                        NombresCompletos=data['NombresCompletos'],
                        email=data['email'],
                        phone=data.get('phone', ''),
                        edad=int(float(data['edad'])) if data.get('edad') else None,
                        concurso=scope_concurso,
                        carrera=scope_carrera_final
                    )
                    
                    created_count += 1
                    
                except Exception as e:
                    errors.append(f"Error al crear participante con cédula {data.get('cedula', 'N/A')}: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'created_count': created_count,
                'errors': errors
            })
            
        except Exception as e:
            return JsonResponse({'error': f'Error al guardar participantes: {str(e)}'}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def quiz_view(request):
    """
    Vista que maneja las evaluaciones según el rol del usuario:
    - Superuser: muestra manage_quizs.html para gestionar evaluaciones
    - Admin (con o sin acceso_total): muestra manage_quizs.html para gestionar evaluaciones
    - Estudiante: muestra quiz.html para ver evaluaciones disponibles
    """
    user = request.user
    
    # Determinar el tipo de usuario
    is_admin = user.is_superuser or AdminProfile.objects.filter(user=user).exists()
    
    if is_admin:
        # Es admin (superuser, admin con acceso_total, o admin sin acceso_total) - mostrar gestión de evaluaciones
        return manage_quizs(request)
    else:
        # Es estudiante - mostrar evaluaciones disponibles
        return student_quizs(request)

@login_required
def manage_quizs(request):
    """
    Vista para que los administradores gestionen las evaluaciones
    Acceso permitido para:
    - Superuser
    - Admin con acceso_total = True
    - Admin con acceso_total = False
    """
    user = request.user
    # Verificar que sea admin (superuser o admin con perfil)
    is_superuser = user.is_superuser
    is_admin = AdminProfile.objects.filter(user=user).exists()
    
    if not (is_superuser or is_admin):
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('quizzes:dashboard')
    
    # Obtener evaluaciones del ámbito
    evaluaciones = filter_queryset_by_scope(
        Evaluacion.objects.order_by('-start_time'),
        request
    )
    
    is_global, scope_carrera, scope_concurso = get_user_scope(request)
    num_etapas = scope_concurso.num_etapas if scope_concurso else 2

    # Obtener las Unidades Temáticas de la carrera actual para la configuración dinámica de cuotas
    from .models import UnidadTematica
    unidades = UnidadTematica.objects.none()
    if scope_carrera:
        unidades = UnidadTematica.objects.filter(carrera=scope_carrera).order_by('numero')
    elif scope_concurso and scope_concurso.carrera:
        unidades = UnidadTematica.objects.filter(carrera=scope_concurso.carrera).order_by('numero')

    # Determinar el tipo específico de admin para el contexto
    if is_superuser:
        admin_type = 'superuser'
        has_full_access = True
    elif is_admin:
        admin_profile = AdminProfile.objects.get(user=user)
        admin_type = 'admin_full' if admin_profile.acceso_total else 'admin_limited'
        has_full_access = True  # Todos los administradores tienen acceso completo
    else:
        admin_type = 'unknown'
        has_full_access = False
    
    context = {
        'evaluaciones': evaluaciones,
        'unidades': unidades,
        'role': 'admin',
        'admin_type': admin_type,
        'has_full_access': has_full_access,
        'num_etapas': num_etapas,
        'now': timezone.now(),
        'user': user
    }
    
    return render(request, 'quizzes/manage_quizs.html', context)

@login_required
def student_quizs(request):
    """
    Vista para que los estudiantes vean las evaluaciones disponibles
    """
    # Verificar que sea estudiante
    if not Participantes.objects.filter(user=request.user).exists():
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('quizzes:dashboard')
    
    participante = Participantes.objects.get(user=request.user)
    
    # Obtener evaluaciones correspondientes a la carrera del estudiante
    if participante.carrera:
        todas_evaluaciones = Evaluacion.objects.filter(concurso__carrera=participante.carrera).order_by('etapa', 'start_time')
    else:
        todas_evaluaciones = Evaluacion.objects.order_by('etapa', 'start_time')
    
    # Filtrar evaluaciones autorizadas para este participante y agregar información de estado
    evaluaciones_autorizadas = []
    for evaluacion in todas_evaluaciones:
        participantes_autorizados = evaluacion.get_participantes_autorizados()
        if participante in participantes_autorizados:
            # Verificar si hay un intento en progreso con tiempo restante
            resultado_activo = evaluacion.resultados.filter(
                participante=participante, 
                completada=False
            ).first()
            
            puede_continuar = False
            puede_iniciar = False
            
            if resultado_activo:
                # Calcular tiempo transcurrido desde el inicio
                tiempo_transcurrido = (timezone.now() - resultado_activo.fecha_inicio).total_seconds()
                tiempo_total = evaluacion.duration_minutes * 60  # en segundos
                tiempo_restante = max(0, tiempo_total - tiempo_transcurrido)
                
                # Si aún hay tiempo restante, puede continuar (sin importar la ventana de acceso)
                if tiempo_restante > 0:
                    puede_continuar = True
            else:
                # Si no hay intento activo, verificar si puede iniciar uno nuevo
                # Verificar ventana de acceso Y intentos disponibles
                if evaluacion.is_available():
                    # Verificar si tiene intentos disponibles
                    intentos_disponibles = participante.get_intentos_disponibles(evaluacion)
                    if intentos_disponibles > 0:
                        puede_iniciar = True
            
            # Obtener el mejor resultado completado (si existe) para mostrar información
            mejor_resultado = ResultadoEvaluacion.get_mejor_resultado(evaluacion, participante)
            
            # Obtener información de intentos
            intentos_disponibles = participante.get_intentos_disponibles(evaluacion)
            intentos_usados = participante.get_intentos_usados(evaluacion)
            intentos_totales = intentos_disponibles + intentos_usados
            
            evaluaciones_autorizadas.append({
                'evaluacion': evaluacion,
                'resultado_activo': resultado_activo,  # Intento en progreso
                'mejor_resultado': mejor_resultado,    # Mejor intento completado
                'puede_continuar': puede_continuar,
                'puede_iniciar': puede_iniciar,
                'intentos_disponibles': intentos_disponibles,
                'intentos_usados': intentos_usados,
                'intentos_totales': intentos_totales
            })
    
    context = {
        'evaluaciones': evaluaciones_autorizadas,
        'role': 'student',
        'current_time': timezone.now(),
        'now': timezone.now(),
        'participante': participante
    }
    return render(request, 'quizzes/quiz.html', context)

@login_required
def student_results(request):
    """
    Vista para que los estudiantes vean sus resultados
    """
    # Verificar que sea estudiante
    if not Participantes.objects.filter(user=request.user).exists():
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('quizzes:dashboard')
    
    participante = Participantes.objects.get(user=request.user)
    
    # Obtener resultados del participante
    resultados = ResultadoEvaluacion.objects.filter(
        participante=participante,
        completada=True
    ).select_related('evaluacion').order_by('evaluacion__etapa', '-fecha_fin')
    
    context = {
        'participante': participante,
        'resultados': resultados,
        'role': 'student'
    }
    return render(request, 'quizzes/student_results.html', context)

@login_required
def revisar_intento_evaluacion(request, pk):
    """
    Vista para revisar el intento de una evaluación específica del estudiante
    """
    # Verificar que sea estudiante
    if not Participantes.objects.filter(user=request.user).exists():
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('quizzes:dashboard')
    
    participante = Participantes.objects.get(user=request.user)
    
    # Obtener el resultado específico
    resultado = get_object_or_404(ResultadoEvaluacion, pk=pk, participante=participante, completada=True)
    evaluacion = resultado.evaluacion

    # No revelar preguntas ni respuestas correctas mientras la ventana de la
    # evaluación siga abierta, incluso si el estudiante accede a la URL directa.
    if not evaluacion.is_finished():
        return redirect('quizzes:student_results')
    
    snapshot = resultado.get_snapshot_respuestas()
    preguntas_con_respuestas = []
    
    if snapshot:
        for s in snapshot:
            preguntas_con_respuestas.append({
                'pregunta': {'id': s.get('pregunta_id'), 'text': s.get('text')},
                'opciones': s.get('opciones', []),
                'respuesta_estudiante': s.get('respuesta_estudiante'),
                'opcion_correcta': s.get('opcion_correcta'),
                'es_correcta': s.get('es_correcta', False)
            })
    else:
        # Fallback para exámenes legacy anteriores al snapshot
        preguntas_del_intento = evaluacion.get_preguntas_para_estudiante(participante.id, resultado.numero_intento)
        respuestas_guardadas = resultado.respuestas_guardadas or {}
        
        for pregunta in preguntas_del_intento:
            pregunta_key = f"pregunta_{pregunta.id}"
            respuesta_estudiante_id = respuestas_guardadas.get(pregunta_key)
            
            opciones_data = []
            opcion_correcta = None
            respuesta_estudiante = None
            
            for opcion in pregunta.opciones.all():
                opciones_data.append({
                    'id': opcion.id,
                    'text': opcion.text,
                    'is_correct': opcion.is_correct,
                    'seleccionada': str(opcion.id) == str(respuesta_estudiante_id)
                })
                
                if opcion.is_correct:
                    opcion_correcta = opcion
                
                if str(opcion.id) == str(respuesta_estudiante_id):
                    respuesta_estudiante = opcion
            
            preguntas_con_respuestas.append({
                'pregunta': pregunta,
                'opciones': opciones_data,
                'respuesta_estudiante': respuesta_estudiante,
                'opcion_correcta': opcion_correcta,
                'es_correcta': respuesta_estudiante and respuesta_estudiante.is_correct if respuesta_estudiante else False
            })
    
    context = {
        'participante': participante,
        'resultado': resultado,
        'evaluacion': evaluacion,
        'preguntas_con_respuestas': preguntas_con_respuestas,
        'role': 'student'
    }
    return render(request, 'quizzes/revisar_intento.html', context)

@login_required
def manage_questions(request, eval_id):
    """
    Vista para gestionar las preguntas de una evaluación específica
    """
    evaluacion = get_object_or_404(Evaluacion, pk=eval_id)
    
    # Verificar si se pueden modificar las preguntas
    can_modify, restriction_message = check_question_modification_allowed(evaluacion)
    
    # Optimizar consulta para incluir opciones y categorías, evitar N+1
    preguntas_qs = evaluacion.preguntas.prefetch_related('opciones').select_related('categoria', 'categoria__unidad').order_by('id')
    
    # Paginación — 20 preguntas por página
    paginator = Paginator(preguntas_qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    from .models import UnidadTematica, Categoria, EvaluacionCuotaUnidad
    from django.db.models import Count
    
    # Obtener carrera de la evaluación para cargar sus Unidades Temáticas
    carrera = evaluacion.concurso.carrera if (evaluacion.concurso and evaluacion.concurso.carrera) else None
    unidades = UnidadTematica.objects.filter(carrera=carrera).order_by('numero') if carrera else UnidadTematica.objects.none()
    
    # Contar preguntas creadas en la evaluación agrupadas por unidad temática
    preguntas_por_unidad = dict(
        evaluacion.preguntas.values_list('categoria__unidad_id').annotate(total=Count('id'))
    )
    
    # Obtener cuotas actuales configuradas
    cuotas_map = {c.unidad_id: c.cantidad_preguntas for c in evaluacion.cuotas_unidades.all()}
    
    unidades_info = []
    for idx, u in enumerate(unidades):
        disponibles = preguntas_por_unidad.get(u.id, 0)
        cantidad_cuota = cuotas_map.get(u.id, 4 if idx == 0 else 2)
        unidades_info.append({
            'unidad': u,
            'disponibles': disponibles,
            'cantidad': cantidad_cuota
        })
    
    # Obtener todas las categorías (temas) activas con su unidad relacionada
    categorias = Categoria.objects.filter(activa=True).select_related('unidad').order_by('unidad__numero', 'nombre')
    if carrera:
        categorias = categorias.filter(unidad__carrera=carrera)
    
    context = {
        'evaluacion': evaluacion,
        'preguntas': page_obj,          # ahora es la página actual
        'total_preguntas': paginator.count,  # total global para el badge
        'page_obj': page_obj,
        'categorias': categorias,
        'unidades_info': unidades_info,
        'can_modify_questions': can_modify,
        'restriction_message': restriction_message,
    }
    return render(request, 'quizzes/manage_questions.html', context)


@csrf_exempt
@login_required
def update_evaluacion_cuotas(request, eval_id):
    """
    Vista AJAX para guardar las cuotas de preguntas por Unidad Temática en la vista de gestionar preguntas
    """
    evaluacion = get_object_or_404(Evaluacion, pk=eval_id)
    can_modify, restriction_message = check_question_modification_allowed(evaluacion)
    if not can_modify:
        return JsonResponse({'success': False, 'error': restriction_message}, status=403)
        
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            cuotas_input = data.get('cuotas_unidades', [])
            
            if not isinstance(cuotas_input, list):
                return JsonResponse({'success': False, 'error': 'Formato de cuotas inválido.'}, status=400)
                
            from .models import EvaluacionCuotaUnidad
            evaluacion.cuotas_unidades.all().delete()
            
            for item in cuotas_input:
                u_id = item.get('unidad_id')
                cant = int(item.get('cantidad', 0))
                if u_id and cant >= 0:
                    EvaluacionCuotaUnidad.objects.create(
                        evaluacion=evaluacion,
                        unidad_id=u_id,
                        cantidad_preguntas=cant
                    )
                    
            evaluacion.save()  # Recalcula preguntas_a_mostrar
            
            return JsonResponse({
                'success': True,
                'message': 'Configuración de cuotas de preguntas guardada exitosamente.',
                'preguntas_a_mostrar': evaluacion.preguntas_a_mostrar
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
            
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)

@csrf_exempt
@login_required
def create_evaluacion(request):
    if request.method == 'POST':
        import json
        data = json.loads(request.body.decode('utf-8'))
        title = data.get('title')
        etapa = data.get('etapa')
        start_date = data.get('start_date')
        start_time = data.get('start_time')
        end_date = data.get('end_date')
        end_time = data.get('end_time')
        duration = data.get('duration')
        description = data.get('description', '')
        
        from datetime import datetime
        from django.utils import timezone
        try:
            start_dt = timezone.make_aware(datetime.strptime(f"{start_date} {start_time}", "%Y-%m-%d %H:%M"))
            end_dt = timezone.make_aware(datetime.strptime(f"{end_date} {end_time}", "%Y-%m-%d %H:%M"))
        except Exception as e:
            return JsonResponse({'success': False, 'error': 'Formato de fecha/hora inválido.'}, status=400)
        
        if not title or not etapa or not duration or not start_date or not start_time or not end_date or not end_time:
            return JsonResponse({'success': False, 'error': 'Faltan campos obligatorios.'}, status=400)
        
        is_global, scope_carrera, scope_concurso = get_user_scope(request)
        if not scope_concurso:
            return JsonResponse({
                'success': False,
                'error': '⚠️ Debes seleccionar un Concurso activo específico en la barra superior (Navbar) antes de crear una evaluación.'
            }, status=400)

        # Validar etapa según las etapas configuradas en el concurso activo
        try:
            etapa = int(etapa)
            num_etapas = scope_concurso.num_etapas
            allowed_etapas = [1, 2] if num_etapas == 2 else [1, 2, 3]
            if etapa not in allowed_etapas:
                return JsonResponse({'success': False, 'error': f'Etapa inválida. El concurso activo "{scope_concurso.nombre}" tiene configuradas únicamente {num_etapas} etapas.'}, status=400)
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Etapa debe ser un número.'}, status=400)
        
        try:
            evaluacion = Evaluacion.objects.create(
                title=title,
                etapa=etapa,
                start_time=start_dt,
                end_time=end_dt,
                duration_minutes=int(duration),
                concurso=scope_concurso
            )
            
            return JsonResponse({
                'success': True, 
                'id': evaluacion.id, 
                'title': evaluacion.title,
                'etapa': evaluacion.etapa,
                'etapa_display': evaluacion.get_etapa_display()
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)

@csrf_exempt
@login_required
def delete_evaluacion(request, pk):
    evaluacion = get_object_or_404(Evaluacion, pk=pk)
    
    if request.method == 'DELETE':
        evaluacion.delete()
        return JsonResponse({'success': True, 'message': 'Evaluación eliminada correctamente.'})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)

@login_required
def edit_evaluacion(request, pk):
    """
    Vista para editar una evaluación
    """
    evaluacion = get_object_or_404(Evaluacion, pk=pk)
    
    # Verificar permisos básicos (solo admins pueden editar)
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('quizzes:dashboard')
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))
            
            # Validar datos
            title = data.get('title', '').strip()
            start_date = data.get('start_date')
            start_time = data.get('start_time')
            end_date = data.get('end_date')
            end_time = data.get('end_time')
            duration = data.get('duration')
            cuotas_unidades_input = data.get('cuotas_unidades', None)
            
            if not all([title, start_date, start_time, end_date, end_time, duration]):
                return JsonResponse({
                    'success': False, 
                    'error': 'Todos los campos son obligatorios'
                }, status=400)
            
            # Convertir fechas
            from datetime import datetime
            start_dt = timezone.make_aware(datetime.strptime(f"{start_date} {start_time}", "%Y-%m-%d %H:%M"))
            end_dt = timezone.make_aware(datetime.strptime(f"{end_date} {end_time}", "%Y-%m-%d %H:%M"))
            
            # Validar que la fecha de inicio sea anterior a la de fin
            if start_dt >= end_dt:
                return JsonResponse({
                    'success': False, 
                    'error': 'La fecha de inicio debe ser anterior a la fecha de finalización'
                }, status=400)
            
            # Actualizar evaluación
            evaluacion.title = title
            evaluacion.etapa = int(data.get('etapa', evaluacion.etapa))
            evaluacion.start_time = start_dt
            evaluacion.end_time = end_dt
            evaluacion.duration_minutes = int(duration)
            evaluacion.save()
            
            if cuotas_unidades_input is not None and isinstance(cuotas_unidades_input, list):
                from .models import EvaluacionCuotaUnidad
                evaluacion.cuotas_unidades.all().delete()
                for item in cuotas_unidades_input:
                    u_id = item.get('unidad_id')
                    cant = int(item.get('cantidad', 0))
                    if u_id and cant > 0:
                        EvaluacionCuotaUnidad.objects.create(
                            evaluacion=evaluacion,
                            unidad_id=u_id,
                            cantidad_preguntas=cant
                        )
                evaluacion.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Evaluación actualizada exitosamente',
                'evaluacion': {
                    'id': evaluacion.id,
                    'title': evaluacion.title,
                    'start_time': evaluacion.start_time.strftime("%d/%m/%Y %H:%M"),
                    'end_time': evaluacion.end_time.strftime("%d/%m/%Y %H:%M"),
                    'duration_minutes': evaluacion.duration_minutes,
                    'preguntas_a_mostrar': evaluacion.preguntas_a_mostrar,
                    'status': evaluacion.get_status(),
                    'status_display': evaluacion.get_status_display()
                }
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False, 
                'error': 'Datos JSON inválidos'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'error': f'Error al actualizar la evaluación: {str(e)}'
            }, status=500)
    
    # GET request - mostrar formulario de edición con Unidades y sus cuotas asociadas
    from .models import UnidadTematica
    carrera = evaluacion.concurso.carrera if (evaluacion.concurso and evaluacion.concurso.carrera) else None
    unidades = UnidadTematica.objects.filter(carrera=carrera).order_by('numero') if carrera else UnidadTematica.objects.none()
    
    cuotas_map = {c.unidad_id: c.cantidad_preguntas for c in evaluacion.cuotas_unidades.all()}
    unidades_con_cuotas = []
    for idx, u in enumerate(unidades):
        unidades_con_cuotas.append({
            'unidad': u,
            'cantidad': cuotas_map.get(u.id, 4 if idx == 0 else 2)
        })

    context = {
        'evaluacion': evaluacion,
        'unidades_con_cuotas': unidades_con_cuotas
    }
    return render(request, 'quizzes/edit_evaluacion.html', context)

@csrf_exempt
@login_required
def upload_image(request):
    """
    Vista para subir imágenes desde CKEditor
    """
    if request.method == 'POST':
        if request.FILES.get('upload'):
            uploaded_file = request.FILES['upload']
            
            # Validar que sea una imagen (permitir más formatos)
            allowed_types = [
                'image/jpeg', 'image/jpg', 'image/png', 'image/gif', 
                'image/bmp', 'image/webp', 'image/tiff', 'image/svg+xml'
            ]
            
            # Si el content_type no está en la lista, verificar la extensión del archivo
            if uploaded_file.content_type not in allowed_types:
                # Verificar extensión del archivo como respaldo
                file_extension = os.path.splitext(uploaded_file.name)[1].lower()
                allowed_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.tiff', '.svg']
                
                if file_extension not in allowed_extensions:
                    return JsonResponse({
                        'error': {
                            'message': f'Formato de imagen no soportado. Formatos permitidos: JPEG, PNG, GIF, BMP, WebP, TIFF, SVG'
                        }
                    }, status=400)
            
            # Crear directorio si no existe en media/ckeditor_uploads
            upload_dir = os.path.join(settings.MEDIA_ROOT, 'ckeditor_uploads')
            os.makedirs(upload_dir, exist_ok=True)
            
            # Generar nombre único para el archivo
            import uuid
            file_extension = os.path.splitext(uploaded_file.name)[1]
            filename = f"{uuid.uuid4()}{file_extension}"
            filepath = os.path.join(upload_dir, filename)
            
            # Guardar el archivo
            with open(filepath, 'wb+') as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)
            
            # Retornar URL para CKEditor (usando media URL)
            file_url = f"{settings.MEDIA_URL}ckeditor_uploads/{filename}"
            
            return JsonResponse({
                'url': file_url,
                'uploaded': 1,
                'fileName': filename
            })
        else:
            return JsonResponse({
                'error': {
                    'message': 'No se recibió ningún archivo'
                }
            }, status=400)
    
    return JsonResponse({
        'error': {
            'message': 'Método no permitido'
        }
    }, status=405)

@csrf_exempt
@login_required
def save_question(request, eval_id):
    """
    Vista para guardar una pregunta y sus opciones
    """
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body.decode('utf-8'))
            
            # Obtener la evaluación
            evaluacion = get_object_or_404(Evaluacion, pk=eval_id)
            
            # Verificar si se pueden modificar las preguntas
            can_modify, restriction_message = check_question_modification_allowed(evaluacion)
            if not can_modify:
                return JsonResponse({
                    'success': False, 
                    'error': restriction_message
                }, status=403)
            
            # Obtener datos del formulario
            pregunta_texto = data.get('pregunta', '').strip()
            opciones = data.get('opciones', [])
            opcion_correcta = data.get('opcion_correcta')
            puntos = data.get('puntos', 1)
            categoria_id = data.get('categoria')
            
            # Validaciones
            if not pregunta_texto:
                return JsonResponse({
                    'success': False, 
                    'error': 'El enunciado de la pregunta es obligatorio'
                }, status=400)
            
            if not categoria_id:
                return JsonResponse({
                    'success': False, 
                    'error': 'La categoría (Tema) es obligatoria'
                }, status=400)
            
            if len(opciones) != 4:
                return JsonResponse({
                    'success': False, 
                    'error': 'Debe proporcionar exactamente 4 opciones'
                }, status=400)
            
            if not opcion_correcta or not str(opcion_correcta).isdigit() or int(opcion_correcta) < 1 or int(opcion_correcta) > 4:
                return JsonResponse({
                    'success': False, 
                    'error': 'Debe seleccionar una opción correcta válida'
                }, status=400)
            
            # Validar puntos
            try:
                puntos = int(puntos)
                if puntos < 1 or puntos > 10:
                    return JsonResponse({
                        'success': False, 
                        'error': 'Los puntos deben estar entre 1 y 10'
                    }, status=400)
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False, 
                    'error': 'Los puntos deben ser un número válido'
                }, status=400)
            
            # Validar categoría
            from .models import Categoria
            try:
                categoria = Categoria.objects.get(id=categoria_id, activa=True)
            except Categoria.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'error': 'La categoría seleccionada no es válida'
                }, status=400)
            
            # Crear la pregunta
            pregunta = Pregunta.objects.create(
                evaluacion=evaluacion,
                categoria=categoria,
                text=pregunta_texto,
                puntos=puntos
            )
            
            # Validar que todas las opciones tengan contenido
            for i, opcion_texto in enumerate(opciones):
                if not opcion_texto.strip():
                    return JsonResponse({
                        'success': False, 
                        'error': f'La opción {chr(65 + i)} ({"ABCD"[i]}) es obligatoria'
                    }, status=400)
            
            # Crear las opciones
            opcion_correcta_index = int(opcion_correcta) - 1  # Convertir a índice 0-based
            
            for i, opcion_texto in enumerate(opciones):
                Opcion.objects.create(
                    pregunta=pregunta,
                    text=opcion_texto.strip(),
                    is_correct=(i == opcion_correcta_index)
                )
            
            return JsonResponse({
                'success': True,
                'message': 'Pregunta guardada exitosamente',
                'pregunta_id': pregunta.id,
                'total_preguntas': evaluacion.preguntas.count()
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False, 
                'error': 'Datos JSON inválidos'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'error': f'Error al guardar la pregunta: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False, 
        'error': 'Método no permitido'
    }, status=405)

@csrf_exempt
@login_required
def delete_question(request, pk):
    """
    Vista para eliminar una pregunta y sus opciones
    """
    if request.method == 'POST':
        try:
            pregunta = get_object_or_404(Pregunta, pk=pk)
            evaluacion = pregunta.evaluacion
            
            # Verificar si se pueden modificar las preguntas
            can_modify, restriction_message = check_question_modification_allowed(evaluacion)
            if not can_modify:
                return JsonResponse({
                    'success': False, 
                    'error': restriction_message
                }, status=403)
            
            # Eliminar la pregunta (esto también eliminará las opciones por CASCADE)
            pregunta.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Pregunta eliminada exitosamente',
                'total_preguntas': evaluacion.preguntas.count()
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'error': f'Error al eliminar la pregunta: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False, 
        'error': 'Método no permitido'
    }, status=405)


@csrf_exempt
@login_required
def get_question_data(request, pk):
    """
    Vista para obtener los datos de una pregunta para edición
    """
    if request.method == 'GET':
        try:
            pregunta = get_object_or_404(Pregunta, pk=pk)
            opciones = pregunta.opciones.all().order_by('id')
            
            # Encontrar la opción correcta
            opcion_correcta = None
            for i, opcion in enumerate(opciones):
                if opcion.is_correct:
                    opcion_correcta = i + 1
                    break
            
            return JsonResponse({
                'success': True,
                'data': {
                    'pregunta': pregunta.text,
                    'opciones': [opcion.text for opcion in opciones],
                    'opcion_correcta': opcion_correcta,
                    'puntos': pregunta.puntos,
                    'categoria': pregunta.categoria.id if pregunta.categoria else None
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'error': f'Error al obtener datos de la pregunta: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False, 
        'error': 'Método no permitido'
    }, status=405)


@csrf_exempt
@login_required
def update_question(request, pk):
    """
    Vista para actualizar una pregunta existente
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            pregunta = get_object_or_404(Pregunta, pk=pk)
            evaluacion = pregunta.evaluacion
            
            # Verificar si se pueden modificar las preguntas
            can_modify, restriction_message = check_question_modification_allowed(evaluacion)
            if not can_modify:
                return JsonResponse({
                    'success': False, 
                    'error': restriction_message
                }, status=403)
            
            # Validar datos
            pregunta_texto = data.get('pregunta', '').strip()
            opciones = data.get('opciones', [])
            opcion_correcta = data.get('opcion_correcta')
            puntos = data.get('puntos', 1)
            categoria_id = data.get('categoria')
            
            if not pregunta_texto:
                return JsonResponse({
                    'success': False, 
                    'error': 'El enunciado de la pregunta es obligatorio'
                }, status=400)
            
            if not categoria_id:
                return JsonResponse({
                    'success': False, 
                    'error': 'La categoría (Tema) es obligatoria'
                }, status=400)
            
            if not opcion_correcta:
                return JsonResponse({
                    'success': False, 
                    'error': 'Debe seleccionar una opción correcta'
                }, status=400)
            
            # Validar puntos
            try:
                puntos = int(puntos)
                if puntos < 1 or puntos > 10:
                    return JsonResponse({
                        'success': False, 
                        'error': 'Los puntos deben estar entre 1 y 10'
                    }, status=400)
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False, 
                    'error': 'Los puntos deben ser un número válido'
                }, status=400)
            
            # Validar que todas las opciones tengan contenido
            for i, opcion_texto in enumerate(opciones):
                if not opcion_texto.strip():
                    return JsonResponse({
                        'success': False, 
                        'error': f'La opción {chr(65 + i)} ({"ABCD"[i]}) es obligatoria'
                    }, status=400)
            
            # Validar categoría
            from .models import Categoria
            try:
                categoria = Categoria.objects.get(id=categoria_id, activa=True)
            except Categoria.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'error': 'La categoría seleccionada no es válida'
                }, status=400)
            
            # Actualizar la pregunta
            pregunta.text = pregunta_texto
            pregunta.puntos = puntos
            pregunta.categoria = categoria
            pregunta.save()
            
            # Eliminar opciones existentes y crear nuevas
            pregunta.opciones.all().delete()
            
            # Crear las nuevas opciones
            opcion_correcta_index = int(opcion_correcta) - 1  # Convertir a índice 0-based
            
            for i, opcion_texto in enumerate(opciones):
                Opcion.objects.create(
                    pregunta=pregunta,
                    text=opcion_texto.strip(),
                    is_correct=(i == opcion_correcta_index)
                )
            
            return JsonResponse({
                'success': True,
                'message': 'Pregunta actualizada exitosamente'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False, 
                'error': 'Datos JSON inválidos'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'error': f'Error al actualizar la pregunta: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False, 
        'error': 'Método no permitido'
    }, status=405)

# Nuevas vistas para las opciones del dropdown de evaluaciones

@login_required
def view_evaluacion(request, pk):
    """
    Vista para ver los detalles de una evaluación
    """
    evaluacion = get_object_or_404(Evaluacion, pk=pk)
    
    # Verificar permisos (solo admins pueden ver detalles)
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('quizzes:dashboard')
    
    context = {
        'evaluacion': evaluacion,
        'preguntas': evaluacion.preguntas.prefetch_related('opciones').all(),
        'total_preguntas': evaluacion.preguntas.count(),
        'participantes_count': Participantes.objects.count()
    }
    
    return render(request, 'quizzes/view_evaluacion.html', context)

@login_required
def edit_evaluacion(request, pk):
    """
    Vista para editar una evaluación
    """
    evaluacion = get_object_or_404(Evaluacion, pk=pk)
    
    # Verificar permisos básicos (solo admins pueden editar)
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('quizzes:dashboard')
    
    # Todos los administradores pueden editar evaluaciones de cualquier etapa
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))
            
            # Validar datos
            title = data.get('title', '').strip()
            start_date = data.get('start_date')
            start_time = data.get('start_time')
            end_date = data.get('end_date')
            end_time = data.get('end_time')
            duration = data.get('duration')
            cuotas_unidades_input = data.get('cuotas_unidades', None)
            
            if not all([title, start_date, start_time, end_date, end_time, duration]):
                return JsonResponse({
                    'success': False, 
                    'error': 'Todos los campos son obligatorios'
                }, status=400)
            
            # Convertir fechas
            from datetime import datetime
            start_dt = timezone.make_aware(datetime.strptime(f"{start_date} {start_time}", "%Y-%m-%d %H:%M"))
            end_dt = timezone.make_aware(datetime.strptime(f"{end_date} {end_time}", "%Y-%m-%d %H:%M"))
            
            # Validar que la fecha de inicio sea anterior a la de fin
            if start_dt >= end_dt:
                return JsonResponse({
                    'success': False, 
                    'error': 'La fecha de inicio debe ser anterior a la fecha de finalización'
                }, status=400)
            
            # Actualizar evaluación
            evaluacion.title = title
            evaluacion.etapa = int(data.get('etapa', evaluacion.etapa))
            evaluacion.start_time = start_dt
            evaluacion.end_time = end_dt
            evaluacion.duration_minutes = int(duration)
            evaluacion.save()
            
            if cuotas_unidades_input is not None and isinstance(cuotas_unidades_input, list):
                from .models import EvaluacionCuotaUnidad
                evaluacion.cuotas_unidades.all().delete()
                for item in cuotas_unidades_input:
                    u_id = item.get('unidad_id')
                    cant = int(item.get('cantidad', 0))
                    if u_id and cant > 0:
                        EvaluacionCuotaUnidad.objects.create(
                            evaluacion=evaluacion,
                            unidad_id=u_id,
                            cantidad_preguntas=cant
                        )
                evaluacion.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Evaluación actualizada exitosamente',
                'evaluacion': {
                    'id': evaluacion.id,
                    'title': evaluacion.title,
                    'start_time': evaluacion.start_time.strftime("%d/%m/%Y %H:%M"),
                    'end_time': evaluacion.end_time.strftime("%d/%m/%Y %H:%M"),
                    'duration_minutes': evaluacion.duration_minutes,
                    'preguntas_a_mostrar': evaluacion.preguntas_a_mostrar,
                    'status': evaluacion.get_status(),
                    'status_display': evaluacion.get_status_display()
                }
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False, 
                'error': 'Datos JSON inválidos'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'error': f'Error al actualizar la evaluación: {str(e)}'
            }, status=500)
    
    # GET request - mostrar formulario de edición
    context = {
        'evaluacion': evaluacion
    }
    return render(request, 'quizzes/edit_evaluacion.html', context)

@login_required
def evaluacion_results(request, pk):
    """
    Vista para ver los resultados de una evaluación
    """
    evaluacion = get_object_or_404(Evaluacion, pk=pk)
    
    # Verificar permisos (solo admins pueden ver resultados)
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('quizzes:dashboard')
    
    # Obtener filtros desde los parámetros GET
    grupo_filtro = request.GET.get('grupo', 'todos')
    categoria_filtro = request.GET.get('categoria', 'todas')
    
    # Obtener todos los grupos disponibles para esta evaluación
    grupos_disponibles = GrupoParticipantes.objects.filter(
        participantes__resultados__evaluacion=evaluacion
    ).distinct().order_by('name')
    
    # Obtener todas las categorías disponibles para esta evaluación
    from .models import Categoria
    categorias_disponibles = Categoria.objects.filter(
        preguntas__in=evaluacion.preguntas.all()
    ).distinct().order_by('nombre')
    
    # Filtrar resultados según el grupo y categoría seleccionados
    if grupo_filtro != 'todos' and grupo_filtro.isdigit():
        grupo_seleccionado = get_object_or_404(GrupoParticipantes, id=int(grupo_filtro))
        participantes_filtrados = grupo_seleccionado.participantes.all()
        
        # Filtrar resultados por participantes del grupo
        resultados_completados = ResultadoEvaluacion.objects.filter(
            evaluacion=evaluacion, 
            completada=True,
            participante__in=participantes_filtrados
        )
        todos_resultados = ResultadoEvaluacion.objects.filter(
            evaluacion=evaluacion,
            participante__in=participantes_filtrados
        )
    else:
        grupo_seleccionado = None
        # Obtener todos los resultados
        resultados_completados = ResultadoEvaluacion.objects.filter(
            evaluacion=evaluacion, 
            completada=True
        )
        todos_resultados = ResultadoEvaluacion.objects.filter(evaluacion=evaluacion)
    
    # Manejar filtro por categoría
    if categoria_filtro != 'todas':
        if categoria_filtro == 'sin_categoria':
            categoria_seleccionada = None
            # Filtrar preguntas sin categoría
            preguntas_filtradas = evaluacion.preguntas.filter(categoria__isnull=True)
        elif categoria_filtro.isdigit():
            categoria_seleccionada = get_object_or_404(Categoria, id=int(categoria_filtro))
            # Filtrar preguntas de la categoría seleccionada
            preguntas_filtradas = evaluacion.preguntas.filter(categoria=categoria_seleccionada)
        else:
            categoria_seleccionada = None
            preguntas_filtradas = evaluacion.preguntas.all()
    else:
        categoria_seleccionada = None
        preguntas_filtradas = evaluacion.preguntas.all()
    
    # Estadísticas de participación
    participantes_con_resultados = todos_resultados.values_list('participante', flat=True).distinct().count()
    participantes_completaron = resultados_completados.values_list('participante', flat=True).distinct().count()
    participantes_en_progreso = todos_resultados.filter(completada=False).values_list('participante', flat=True).distinct().count()
    
    # Obtener participantes elegibles para esta evaluación
    if grupo_seleccionado:
        participantes_elegibles = grupo_seleccionado.participantes.count()
    elif evaluacion.etapa == 1:
        participantes_elegibles = len(evaluacion.get_participantes_etapa1())
    else:
        # Para etapas superiores, usar todos los participantes registrados como referencia
        participantes_elegibles = Participantes.objects.count()
    
    participantes_no_iniciaron = max(0, participantes_elegibles - participantes_con_resultados)
    
    # Calcular tasa de participación
    tasa_participacion = (participantes_con_resultados / participantes_elegibles * 100) if participantes_elegibles > 0 else 0
    
    # Estadísticas de rendimiento
    estadisticas_rendimiento = {}
    if resultados_completados.exists():
        from django.db.models import Avg, Max, Min
        
        stats = resultados_completados.aggregate(
            promedio_puntaje=Avg('puntos_obtenidos'),
            mejor_puntaje=Max('puntos_obtenidos'),
            peor_puntaje=Min('puntos_obtenidos'),
            tiempo_promedio=Avg('tiempo_utilizado')
        )
        
        estadisticas_rendimiento = {
            'promedio_porcentaje': (stats['promedio_puntaje'] / 10 * 100) if stats['promedio_puntaje'] else 0,
            'mejor_puntaje': stats['mejor_puntaje'] if stats['mejor_puntaje'] else 0,
            'peor_puntaje': stats['peor_puntaje'] if stats['peor_puntaje'] else 0,
            'tiempo_promedio': int(stats['tiempo_promedio']) if stats['tiempo_promedio'] else 0,
        }
    else:
        estadisticas_rendimiento = {
            'promedio_porcentaje': 0,
            'mejor_puntaje': 0,
            'peor_puntaje': 0,
            'tiempo_promedio': 0,
        }
    
    # Análisis por pregunta (preparar datos para gráficos) - Optimizado
    analisis_preguntas = []
    
    # Obtener todos los resultados con respuestas guardadas una sola vez
    resultados_con_respuestas = todos_resultados.exclude(
        respuestas_guardadas__isnull=True
    ).exclude(
        respuestas_guardadas={}
    ).select_related('participante').prefetch_related('evaluacion__preguntas__opciones')
    
    # Crear cache de opciones correctas por pregunta para evitar consultas repetidas
    preguntas_opciones_correctas = {}
    for pregunta in evaluacion.preguntas.prefetch_related('opciones'):
        opciones_correctas = {}
        for opcion in pregunta.opciones.all():
            opciones_correctas[opcion.id] = opcion.is_correct
        preguntas_opciones_correctas[pregunta.id] = opciones_correctas
    
    # Procesar cada pregunta filtrada
    for pregunta in preguntas_filtradas:
        correctas = 0
        incorrectas = 0
        sin_responder = 0
        
        # Obtener las opciones correctas para esta pregunta desde el cache
        opciones_correctas_pregunta = preguntas_opciones_correctas.get(pregunta.id, {})
        
        for resultado in resultados_con_respuestas:
            respuestas = obtener_diccionario_respuestas(resultado)
            
            # Validar que respuestas no sea None o vacío
            if not respuestas or not isinstance(respuestas, dict):
                sin_responder += 1
                continue
            
            # Buscar la respuesta para esta pregunta con el formato correcto
            # Las claves se guardan como "pregunta_407", "pregunta_410", etc.
            opcion_id = None
            pregunta_key = f"pregunta_{pregunta.id}"
            
            if pregunta_key in respuestas:
                opcion_id = respuestas[pregunta_key]
            
            if opcion_id is not None:
                try:
                    # Normalizar opcion_id a entero
                    if isinstance(opcion_id, str) and opcion_id.isdigit():
                        opcion_id = int(opcion_id)
                    elif not isinstance(opcion_id, int):
                        sin_responder += 1
                        continue
                    
                    # Verificar si la respuesta es correcta usando el cache
                    if opcion_id in opciones_correctas_pregunta:
                        if opciones_correctas_pregunta[opcion_id]:
                            correctas += 1
                        else:
                            incorrectas += 1
                    else:
                        # La opción no existe, contar como sin responder
                        sin_responder += 1
                        
                except (ValueError, TypeError):
                    sin_responder += 1
            else:
                sin_responder += 1
        
        # Calcular totales y porcentajes
        total_participantes_que_respondieron = correctas + incorrectas
        total_participantes_considerados = resultados_con_respuestas.count()
        
        # Para el cálculo de porcentaje de acierto, usar solo quienes respondieron
        porcentaje_correctas = (correctas / total_participantes_que_respondieron * 100) if total_participantes_que_respondieron > 0 else 0
        
        analisis_preguntas.append({
            'pregunta': pregunta,
            'correctas': correctas,
            'incorrectas': incorrectas,
            'sin_responder': sin_responder,
            'porcentaje_correctas': porcentaje_correctas,
            'dificultad': 'Fácil' if porcentaje_correctas > 70 else 'Media' if porcentaje_correctas > 40 else 'Difícil'
        })
    
    # Crear versión serializable para JavaScript
    analisis_preguntas_json = []
    for analisis in analisis_preguntas:
        analisis_preguntas_json.append({
            'pregunta_id': analisis['pregunta'].id,
            'pregunta_text': analisis['pregunta'].text,
            'correctas': analisis['correctas'],
            'incorrectas': analisis['incorrectas'],
            'sin_responder': analisis['sin_responder'],
            'porcentaje_correctas': analisis['porcentaje_correctas'],
            'dificultad': analisis['dificultad']
        })
    
    # Análisis por categorías
    from collections import defaultdict
    analisis_categorias = []
    categorias_stats = defaultdict(lambda: {
        'preguntas': [],
        'correctas': 0,
        'incorrectas': 0,
        'sin_responder': 0,
        'total_preguntas': 0
    })
    
    # Agrupar análisis de preguntas por categoría
    for analisis in analisis_preguntas:
        pregunta = analisis['pregunta']
        categoria = pregunta.categoria
        
        if categoria:
            categoria_key = categoria.id
            categoria_nombre = categoria.nombre
        else:
            categoria_key = 'sin_categoria'
            categoria_nombre = 'Sin categoría'
        
        categorias_stats[categoria_key]['categoria_nombre'] = categoria_nombre
        categorias_stats[categoria_key]['categoria_obj'] = categoria
        categorias_stats[categoria_key]['preguntas'].append(analisis)
        categorias_stats[categoria_key]['correctas'] += analisis['correctas']
        categorias_stats[categoria_key]['incorrectas'] += analisis['incorrectas']
        categorias_stats[categoria_key]['sin_responder'] += analisis['sin_responder']
        categorias_stats[categoria_key]['total_preguntas'] += 1
    
    # Calcular estadísticas finales por categoría
    for categoria_key, stats in categorias_stats.items():
        total_respuestas = stats['correctas'] + stats['incorrectas']
        porcentaje_acierto = (stats['correctas'] / total_respuestas * 100) if total_respuestas > 0 else 0
        
        analisis_categorias.append({
            'categoria_nombre': stats['categoria_nombre'],
            'categoria_obj': stats['categoria_obj'],
            'total_preguntas': stats['total_preguntas'],
            'correctas': stats['correctas'],
            'incorrectas': stats['incorrectas'],
            'sin_responder': stats['sin_responder'],
            'porcentaje_acierto': porcentaje_acierto,
            'dificultad': 'Fácil' if porcentaje_acierto > 70 else 'Media' if porcentaje_acierto > 40 else 'Difícil',
            'preguntas_detalle': stats['preguntas']
        })
    
    # Ordenar por porcentaje de acierto descendente
    analisis_categorias.sort(key=lambda x: x['porcentaje_acierto'], reverse=True)

    # Distribución de puntajes (para gráficos)
    distribucion_puntajes = []
    if resultados_completados.exists():
        # Crear rangos de puntajes
        rangos = [(0, 2), (2, 4), (4, 6), (6, 8), (8, 10)]
        for rango_min, rango_max in rangos:
            count = resultados_completados.filter(
                puntos_obtenidos__gte=rango_min,
                puntos_obtenidos__lt=rango_max if rango_max < 10 else 11
            ).count()
            distribucion_puntajes.append({
                'rango': f'{rango_min}-{rango_max}',
                'count': count
            })
    
    # Serializar distribucion_puntajes para json_script
    distribucion_puntajes_json = [
        {'rango': item['rango'], 'count': item['count']}
        for item in distribucion_puntajes
    ]

    # Serializar analisis_categorias para json_script
    analisis_categorias_json = [
        {
            'nombre': cat['categoria_nombre'],
            'porcentaje': float(cat['porcentaje_acierto']),
            'preguntas': cat['total_preguntas'],
            'dificultad': cat['dificultad'],
        }
        for cat in analisis_categorias
    ]

    context = {
        'evaluacion': evaluacion,
        'total_preguntas': evaluacion.preguntas.count(),
        'participantes_count': participantes_elegibles,
        'evaluacion_status': evaluacion.get_status(),
        'evaluacion_status_display': evaluacion.get_status_display(),

        # Filtros y grupos
        'grupos_disponibles': grupos_disponibles,
        'grupo_seleccionado': grupo_seleccionado,
        'grupo_filtro': grupo_filtro,
        'categorias_disponibles': categorias_disponibles,
        'categoria_seleccionada': categoria_seleccionada,
        'categoria_filtro': categoria_filtro,

        # Estadísticas de participación
        'participantes_con_resultados': participantes_con_resultados,
        'participantes_completaron': participantes_completaron,
        'participantes_en_progreso': participantes_en_progreso,
        'participantes_no_iniciaron': participantes_no_iniciaron,
        'tasa_participacion': round(tasa_participacion, 1),

        # Estadísticas de rendimiento
        'estadisticas_rendimiento': estadisticas_rendimiento,

        # Análisis detallado
        'analisis_preguntas': analisis_preguntas,
        'analisis_preguntas_json': analisis_preguntas_json,
        'analisis_categorias': analisis_categorias,
        'distribucion_puntajes': distribucion_puntajes,

        # Versiones JSON para json_script (sin template-tags en <script>)
        'distribucion_puntajes_json': distribucion_puntajes_json,
        'analisis_categorias_json': analisis_categorias_json,

        # Top 5 resultados para mostrar
        'top_resultados': resultados_completados.order_by('-puntos_obtenidos', 'tiempo_utilizado')[:5],
    }
    
    return render(request, 'quizzes/evaluacion_results.html', context)

@csrf_exempt
@login_required
def delete_evaluacion(request, pk):
    """
    Vista para eliminar una evaluación
    """
    evaluacion = get_object_or_404(Evaluacion, pk=pk)
    
    # Verificar permisos básicos (solo admins pueden eliminar)
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('quizzes:dashboard')
    
    # Todos los administradores pueden eliminar evaluaciones de cualquier etapa
    
    if request.method == 'POST':
        try:
            evaluacion.delete()
            return JsonResponse({
                'success': True,
                'message': 'Evaluación eliminada exitosamente'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al eliminar la evaluación: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'error': 'Método no permitido'
    }, status=405)

@login_required
def ranking_evaluacion(request, pk):
    """
    Vista para mostrar el ranking de una evaluación
    """
    evaluacion = get_object_or_404(Evaluacion, pk=pk)
    
    # Verificar permisos
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('quizzes:dashboard')
    
    # Obtener el filtro de estado
    filtro_estado = request.GET.get('estado', 'todos')
    
    # Obtener el mejor resultado por participante ordenado por nota (desc) y tiempo en segundos (asc)
    todos_completados = ResultadoEvaluacion.objects.filter(
        evaluacion=evaluacion,
        completada=True
    ).order_by('-puntos_obtenidos', 'tiempo_utilizado').select_related('participante')
    
    resultados = []
    vistos = set()
    for r in todos_completados:
        if r.participante_id not in vistos:
            vistos.add(r.participante_id)
            resultados.append(r)
    
    # Calcular estadísticas
    total_participantes = len(resultados)
    
    # Calcular promedio de puntos obtenidos (número, no porcentaje)
    promedio_puntos = sum(r.puntos_obtenidos for r in resultados) / total_participantes if total_participantes > 0 else 0
    
    # Calcular tiempo promedio real (usando fechas de inicio y fin)
    tiempo_total_minutos = 0
    resultados_con_tiempo_real = 0
    
    for resultado in resultados:
        if resultado.fecha_inicio and resultado.fecha_fin:
            tiempo_segundos = (resultado.fecha_fin - resultado.fecha_inicio).total_seconds()
            tiempo_minutos = tiempo_segundos / 60
            tiempo_total_minutos += tiempo_minutos
            resultados_con_tiempo_real += 1
    
    promedio_tiempo = tiempo_total_minutos / resultados_con_tiempo_real if resultados_con_tiempo_real > 0 else 0
    
    # Determinar ganadores según la etapa y configuración
    ganadores = []
    num_etapas = evaluacion.concurso.num_etapas if evaluacion.concurso else 3
    if evaluacion.etapa == 1:
        top_n = 15 if num_etapas == 3 else 5
        if total_participantes >= top_n:
            ganadores = resultados[:top_n]
    elif evaluacion.etapa == 2:
        # En flujo de 3 etapas, etapa 2 muestra top 5; en flujo de 2 etapas no debería usarse
        if num_etapas == 3 and total_participantes >= 5:
            ganadores = resultados[:5]
    elif evaluacion.etapa == 3 and total_participantes >= 5:
        ganadores = resultados[:5]
    
    # Aplicar filtro de estado y agregar posición real
    # Primero pre-cargar la información de colegios para optimizar consultas
    participante_ids = [resultado.participante.id for resultado in resultados]
    from .models import GrupoParticipantes
    grupos_info = {}
    grupos_queryset = GrupoParticipantes.objects.filter(
        participantes__id__in=participante_ids
    ).select_related('representante').prefetch_related('participantes')
    
    for grupo in grupos_queryset:
        for participante in grupo.participantes.all():
            if participante.id not in grupos_info:
                grupos_info[participante.id] = grupo.representante.NombreColegio if grupo.representante else None
    
    resultados_con_posicion = []
    for i, resultado in enumerate(resultados, 1):
        resultado.posicion_real = i
        
        # Asignar el colegio desde la información pre-cargada
        resultado.participante.colegio_nombre = grupos_info.get(resultado.participante.id, None)
        
        resultados_con_posicion.append(resultado)
    
    resultados_filtrados = resultados_con_posicion
    if filtro_estado == 'clasificados':
        if evaluacion.etapa == 1:
            top_n = 15 if num_etapas == 3 else 5
            resultados_filtrados = resultados_con_posicion[:top_n]
        elif evaluacion.etapa == 2:
            if num_etapas == 3:
                resultados_filtrados = resultados_con_posicion[:5]
        elif evaluacion.etapa == 3:
            resultados_filtrados = resultados_con_posicion[:5]
    elif filtro_estado == 'no_clasificados':
        if evaluacion.etapa == 1:
            top_n = 15 if num_etapas == 3 else 5
            resultados_filtrados = resultados_con_posicion[top_n:] if len(resultados_con_posicion) > top_n else []
        elif evaluacion.etapa == 2:
            if num_etapas == 3:
                resultados_filtrados = resultados_con_posicion[5:] if len(resultados_con_posicion) > 5 else []
            else:
                resultados_filtrados = []  # En flujo de 2 etapas, etapa 2 no debería existir
        elif evaluacion.etapa == 3:
            resultados_filtrados = resultados_con_posicion[5:] if len(resultados_con_posicion) > 5 else []
    
    context = {
        'evaluacion': evaluacion,
        'resultados': resultados_filtrados,
        'resultados_totales': resultados,  # Guardamos los resultados completos para estadísticas
        'total_participantes': total_participantes,
        'promedio_puntaje': round(promedio_puntos, 3),  # Mostrar como número con 3 decimales
        'promedio_tiempo': round(promedio_tiempo, 1),   # Mostrar tiempo en minutos con 1 decimal
        'ganadores': ganadores,
        'num_etapas': num_etapas,
        'filtro_estado': filtro_estado
    }
    
    return render(request, 'quizzes/ranking_evaluacion.html', context)

@csrf_exempt
@login_required
def gestionar_participantes_evaluacion(request, pk):
    """
    Vista para gestionar participantes de una evaluación
    """
    evaluacion = get_object_or_404(Evaluacion, pk=pk)
    
    # Verificar permisos
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        return JsonResponse({
            'success': False, 
            'error': 'No tienes permisos para acceder a esta funcionalidad.'
        }, status=403)
    
    if request.method == 'GET':
        try:
            # Obtener grupos con conteo de participantes
            grupos = []
            for grupo in GrupoParticipantes.objects.all():
                grupos.append({
                    'id': grupo.id,
                    'name': grupo.name,
                    'participantes_count': grupo.participantes.count()
                })
            
            # Obtener participantes individuales según la etapa y permisos del usuario
            participantes_individuales = []
            
            if evaluacion.etapa == 1:
                # Para etapa 1: todos los participantes individuales
                for participante in Participantes.objects.all():
                    if not participante.grupos.exists():
                        participantes_individuales.append({
                            'id': participante.id,
                            'NombresCompletos': participante.NombresCompletos,
                            'cedula': participante.cedula
                        })
            elif evaluacion.etapa == 2:
                # Para etapa 2: solo aplica si el concurso está en 3 etapas
                num_etapas = evaluacion.concurso.num_etapas if evaluacion.concurso else 3
                if num_etapas == 3:
                    if has_full_access(request.user):
                        participantes_automaticos = evaluacion.get_participantes_etapa2()
                        participantes_automaticos_ids = set(p.id for p in participantes_automaticos)
                        # Primero automáticos
                        for participante in participantes_automaticos:
                            participantes_individuales.append({
                                'id': participante.id,
                                'NombresCompletos': participante.NombresCompletos,
                                'cedula': participante.cedula
                            })
                        # Luego el resto
                        for participante in Participantes.objects.all():
                            if participante.id not in participantes_automaticos_ids:
                                participantes_individuales.append({
                                    'id': participante.id,
                                    'NombresCompletos': participante.NombresCompletos,
                                    'cedula': participante.cedula
                                })
                    else:
                        if evaluacion.participantes_individuales.exists():
                            participantes_actuales = evaluacion.participantes_individuales.all()
                        else:
                            participantes_actuales = evaluacion.get_participantes_etapa2()
                        for participante in participantes_actuales:
                            participantes_individuales.append({
                                'id': participante.id,
                                'NombresCompletos': participante.NombresCompletos,
                                'cedula': participante.cedula
                            })
            elif evaluacion.etapa == 3:
                # Para etapa 3: soportar flujo de 3 etapas y de 2 etapas (saltando etapa 2)
                if has_full_access(request.user):
                    participantes_automaticos = evaluacion.get_participantes_etapa3()
                    participantes_automaticos_ids = set(p.id for p in participantes_automaticos)
                    for participante in participantes_automaticos:
                        participantes_individuales.append({
                            'id': participante.id,
                            'NombresCompletos': participante.NombresCompletos,
                            'cedula': participante.cedula
                        })
                    for participante in Participantes.objects.all():
                        if participante.id not in participantes_automaticos_ids:
                            participantes_individuales.append({
                                'id': participante.id,
                                'NombresCompletos': participante.NombresCompletos,
                                'cedula': participante.cedula
                            })
                else:
                    if evaluacion.participantes_individuales.exists():
                        participantes_actuales = evaluacion.participantes_individuales.all()
                    else:
                        participantes_actuales = evaluacion.get_participantes_etapa3()
                    for participante in participantes_actuales:
                        participantes_individuales.append({
                            'id': participante.id,
                            'NombresCompletos': participante.NombresCompletos,
                            'cedula': participante.cedula
                        })
            

            
            # Obtener grupos y participantes asignados actualmente
            grupos_asignados = []
            for grupo in evaluacion.grupos_participantes.all():
                grupos_asignados.append({
                    'id': grupo.id,
                    'name': grupo.name,
                    'participantes_count': grupo.participantes.count()
                })
            
            participantes_asignados = []
            for participante in evaluacion.participantes_individuales.all():
                participantes_asignados.append({
                    'id': participante.id,
                    'NombresCompletos': participante.NombresCompletos,
                    'cedula': participante.cedula
                })
            
            # Para superusuarios en etapas avanzadas, incluir automáticos si no hay asignados manualmente
            participantes_automaticos = []
            if has_full_access(request.user) and evaluacion.etapa in [2, 3] and not evaluacion.participantes_individuales.exists():
                num_etapas = evaluacion.concurso.num_etapas if evaluacion.concurso else 3
                if evaluacion.etapa == 2 and num_etapas == 3:
                    participantes_automaticos = evaluacion.get_participantes_etapa2()
                elif evaluacion.etapa == 3:
                    participantes_automaticos = evaluacion.get_participantes_etapa3()
                
                # Agregar información de que son automáticos
                for participante in participantes_automaticos:
                    participantes_asignados.append({
                        'id': participante.id,
                        'NombresCompletos': participante.NombresCompletos,
                        'cedula': participante.cedula,
                        'automatico': True  # Marcar como automático
                    })
            
            # Para otros casos, marcar como manual
            if not participantes_automaticos:
                for item in participantes_asignados:
                    item['automatico'] = False
            
            return JsonResponse({
                'success': True,
                'grupos': grupos,
                'participantes_individuales': participantes_individuales,
                'grupos_asignados': grupos_asignados,
                'participantes_asignados': participantes_asignados,
                'etapa': evaluacion.etapa
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al cargar datos: {str(e)}'
            }, status=500)
    
    elif request.method == 'POST':
        try:
            import json
            data = json.loads(request.body.decode('utf-8'))
            
            # Verificar permisos para etapas avanzadas
            if evaluacion.etapa != 1 and not has_full_access(request.user):
                return JsonResponse({
                    'success': False,
                    'error': 'Solo los superusuarios y administradores con acceso total pueden modificar participantes en etapas avanzadas.'
                }, status=403)
            
            grupos_ids = data.get('grupos', [])
            participantes_individuales_ids = data.get('participantes_individuales', [])
            
            # Para etapa 1: actualizar grupos y participantes individuales
            if evaluacion.etapa == 1:
                from django.db import transaction
                with transaction.atomic():
                    # Actualizar grupos asignados
                    evaluacion.grupos_participantes.clear()
                    if grupos_ids:
                        grupos = GrupoParticipantes.objects.filter(id__in=grupos_ids)
                        evaluacion.grupos_participantes.add(*grupos)
                    
                    # Actualizar participantes individuales asignados
                    evaluacion.participantes_individuales.clear()
                    if participantes_individuales_ids:
                        participantes = Participantes.objects.filter(id__in=participantes_individuales_ids)
                        evaluacion.participantes_individuales.add(*participantes)
            
            # Para etapas 2 y 3: solo superusuarios y admins con acceso total pueden modificar
            elif evaluacion.etapa in [2, 3] and has_full_access(request.user):
                # Limpiar participantes individuales actuales
                evaluacion.participantes_individuales.clear()
                
                # Agregar solo los participantes seleccionados por el usuario autorizado
                if participantes_individuales_ids:
                    participantes = Participantes.objects.filter(id__in=participantes_individuales_ids)
                    evaluacion.participantes_individuales.add(*participantes)
            
            return JsonResponse({
                'success': True,
                'message': 'Participantes asignados correctamente'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al guardar datos: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'error': 'Método no permitido'
    }, status=405)

@login_required
def exportar_resultado_pdf(request, pk):
    """
    Vista para exportar resultado de evaluación a PDF con diseño moderno y profesional
    """
    try:
        evaluacion = get_object_or_404(Evaluacion, pk=pk)
        participante = Participantes.objects.get(user=request.user)
        
        # Verificar que el participante tenga resultado para esta evaluación
        resultado = ResultadoEvaluacion.objects.filter(
            evaluacion=evaluacion,
            participante=participante,
            completada=True
        ).first()
        
        if not resultado:
            messages.error(request, 'No tienes resultados para esta evaluación.')
            return redirect('quizzes:student_results')
        
        # Crear el PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="resultado_{evaluacion.title}_{participante.cedula}.pdf"'
        
        # Crear el documento PDF con márgenes optimizados
        doc = SimpleDocTemplate(response, pagesize=A4, 
                              leftMargin=0.6*inch, rightMargin=0.6*inch,
                              topMargin=0.6*inch, bottomMargin=0.6*inch)
        elements = []
        
        # Definir paleta de colores moderna y elegante
        primary_color = colors.Color(0.15, 0.35, 0.75)  # Azul corporativo
        secondary_color = colors.Color(0.95, 0.95, 0.98)  # Gris muy claro
        accent_color = colors.Color(0.12, 0.55, 0.35)  # Verde moderno
        header_bg = colors.Color(0.08, 0.25, 0.55)  # Azul más oscuro para encabezado
        success_color = colors.Color(0.1, 0.6, 0.1)  # Verde para éxito
        warning_color = colors.Color(0.8, 0.6, 0.1)  # Amarillo para advertencia
        danger_color = colors.Color(0.8, 0.1, 0.1)  # Rojo para peligro
        
        # Estilos mejorados
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'ModernTitle',
            parent=styles['Heading1'],
            fontSize=22,
            spaceAfter=18,
            alignment=1,  # Centrado
            textColor=primary_color,
            fontName='Helvetica-Bold',
            spaceBefore=8,
            leading=26
        )
        
        subtitle_style = ParagraphStyle(
            'ModernSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            alignment=1,  # Centrado
            textColor=colors.Color(0.3, 0.3, 0.3),
            fontName='Helvetica',
            leading=18
        )
        
        etapa_style = ParagraphStyle(
            'EtapaStyle',
            parent=styles['Heading2'],
            fontSize=18,
            spaceAfter=20,
            alignment=1,  # Centrado
            textColor=accent_color,
            fontName='Helvetica-Bold',
            spaceBefore=10,
            leading=22
        )
        
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=8,
            textColor=colors.Color(0.2, 0.2, 0.2),
            fontName='Helvetica',
            leading=16
        )
        
        # Título principal con mejor formato
        title = Paragraph("OLIMPIADAS DE MATEMÁTICAS<br/>CARRERA MECÁNICA", title_style)
        elements.append(title)
        
        # Manejo mejorado de logos
        logo_mecanica = None
        logo_uteq = None
        
        # Cargar Logo Mecánica
        logo_mecanica_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logoMecanica.png')
        if os.path.exists(logo_mecanica_path):
            try:
                logo_mecanica = Image(logo_mecanica_path, width=1.6*inch, height=1.2*inch)
            except Exception as e:
                print(f"Error al cargar logo Mecánica: {e}")
        
        # Cargar Logo UTEQ
        logo_uteq_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo-uteq.png')
        if os.path.exists(logo_uteq_path):
            try:
                logo_uteq = Image(logo_uteq_path, width=1.6*inch, height=1.2*inch)
            except Exception as e:
                print(f"Error al cargar logo UTEQ: {e}")
        
        # Tabla de logos mejorada
        if logo_mecanica or logo_uteq:
            logo_row = []
            
            if logo_mecanica:
                logo_row.append(logo_mecanica)
            else:
                logo_row.append(Paragraph("", styles['Normal']))
            
            logo_row.append(Paragraph("", styles['Normal']))  # Espacio central
            
            if logo_uteq:
                logo_row.append(logo_uteq)
            else:
                logo_row.append(Paragraph("", styles['Normal']))
            
            logo_table = Table([logo_row], colWidths=[2.4*inch, 1.2*inch, 2.4*inch])
            logo_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('ALIGN', (2, 0), (2, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, 0), 0),
                ('RIGHTPADDING', (0, 0), (-1, 0), 0),
                ('TOPPADDING', (0, 0), (-1, 0), 5),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
            ]))
            
            elements.append(logo_table)
            elements.append(Spacer(1, 20))
        
        # Información de la etapa
        etapa_text = f"CERTIFICADO DE PARTICIPACIÓN - ETAPA {evaluacion.etapa}"
        elements.append(Paragraph(etapa_text, etapa_style))
        
        # Información adicional
        evaluacion_info = f"Evaluación: {evaluacion.title}"
        elements.append(Paragraph(evaluacion_info, subtitle_style))
        elements.append(Spacer(1, 20))
        
        # Información del participante en un recuadro elegante
        participante_info = [
            ['Campo', 'Información'],
            ['Participante', participante.NombresCompletos],
            ['Cédula de Identidad', participante.cedula],
            ['Correo Electrónico', participante.email if participante.email else 'No registrado'],
        ]
        
        info_table = Table(participante_info, colWidths=[2.2*inch, 3.8*inch])
        info_table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Cuerpo
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.7, 0.7, 0.7)),
            ('LINEBELOW', (0, 0), (-1, 0), 2, header_bg),
            
            # Alineación
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Filas alternadas
            ('BACKGROUND', (0, 1), (-1, 1), colors.white),
            ('BACKGROUND', (0, 2), (-1, 2), secondary_color),
            ('BACKGROUND', (0, 3), (-1, 3), colors.white),
        ]))
        
        elements.append(info_table)
        elements.append(Spacer(1, 25))
        
        # Resultados de la evaluación
        results_title_style = ParagraphStyle(
            'ResultsTitle',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=15,
            alignment=1,
            textColor=primary_color,
            fontName='Helvetica-Bold'
        )
        
        elements.append(Paragraph("RESULTADOS DE LA EVALUACIÓN", results_title_style))
        
        # Calcular puntuación en formato numérico sobre 10
        puntaje_numerico = float(resultado.puntos_obtenidos) if hasattr(resultado, 'puntos_obtenidos') else 0.0
        puntos_totales = resultado.puntos_totales if hasattr(resultado, 'puntos_totales') else 10
        
        # Tabla de resultados mejorada
        resultados_data = [
            ['Métrica', 'Resultado'],
            ['Nota Final', f"{puntaje_numerico:.3f}"],
            ['Tiempo Utilizado', resultado.get_tiempo_formateado()],
            ['Fecha de Finalización', resultado.fecha_fin.strftime("%d/%m/%Y a las %H:%M") if resultado.fecha_fin else 'No disponible'],
        ]
        
        results_table = Table(resultados_data, colWidths=[2.5*inch, 3.5*inch])
        results_table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
            ('TOPPADDING', (0, 0), (-1, 0), 15),
            
            # Cuerpo
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 11),
            ('TOPPADDING', (0, 1), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 12),
            ('LEFTPADDING', (0, 0), (-1, -1), 15),
            ('RIGHTPADDING', (0, 0), (-1, -1), 15),
            
            # Bordes elegantes
            ('GRID', (0, 0), (-1, -1), 1, colors.Color(0.6, 0.6, 0.6)),
            ('LINEBELOW', (0, 0), (-1, 0), 2, header_bg),
            
            # Alineación
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Filas alternadas
            ('BACKGROUND', (0, 1), (-1, 1), secondary_color),
            ('BACKGROUND', (0, 2), (-1, 2), colors.white),
            ('BACKGROUND', (0, 3), (-1, 3), secondary_color),
        ]))
        
        # Destacar la puntuación final con color según el rendimiento
        if puntaje_numerico >= 8.0:
            results_table.setStyle(TableStyle([
                ('BACKGROUND', (1, 1), (1, 1), success_color),
                ('TEXTCOLOR', (1, 1), (1, 1), colors.white),
                ('FONTNAME', (1, 1), (1, 1), 'Helvetica-Bold'),
            ]))
        elif puntaje_numerico >= 6.0:
            results_table.setStyle(TableStyle([
                ('BACKGROUND', (1, 1), (1, 1), warning_color),
                ('TEXTCOLOR', (1, 1), (1, 1), colors.white),
                ('FONTNAME', (1, 1), (1, 1), 'Helvetica-Bold'),
            ]))
        else:
            results_table.setStyle(TableStyle([
                ('BACKGROUND', (1, 1), (1, 1), danger_color),
                ('TEXTCOLOR', (1, 1), (1, 1), colors.white),
                ('FONTNAME', (1, 1), (1, 1), 'Helvetica-Bold'),
            ]))
        
        elements.append(results_table)
        elements.append(Spacer(1, 25))
        
        # Comentario personalizado según puntuación
        comment_style = ParagraphStyle(
            'CommentStyle',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=10,
            alignment=1,  # Centrado
            fontName='Helvetica',
            leading=16
        )
        
        if puntaje_numerico >= 9.0:
            mensaje = "<b>¡EXCELENTE DESEMPEÑO!</b><br/>"
        elif puntaje_numerico >= 8.0:
            mensaje = "<b>¡MUY BUEN TRABAJO!</b><br/>"
        elif puntaje_numerico >= 7.0:
            mensaje = "<b>BUEN DESEMPEÑO</b><br/>"
        elif puntaje_numerico >= 6.0:
            mensaje = "<b>DESEMPEÑO ACEPTABLE</b><br/>"
        else:
            mensaje = "<b>OPORTUNIDAD DE MEJORA</b><br/>"

        
        comment_paragraph = Paragraph(mensaje, comment_style)
        elements.append(comment_paragraph)
        elements.append(Spacer(1, 20))
        
        # Pie de página con información institucional
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,  # Centrado
            textColor=colors.Color(0.4, 0.4, 0.4),
            fontName='Helvetica-Oblique',
            leading=12
        )
        
        fecha_emision = timezone.now().strftime("%d de %B de %Y")
        footer_text = f"<i>Documento generado automáticamente el {fecha_emision}<br/>Universidad Técnica Estatal de Quevedo - Carrera de Ingeniería Mecánica</i>"
        elements.append(Paragraph(footer_text, footer_style))
        
        # Construir el PDF
        doc.build(elements)
        
        return response
        
    except Participantes.DoesNotExist:
        messages.error(request, 'No se encontró información del participante.')
        return redirect('quizzes:student_results')
    except Exception as e:
        messages.error(request, f'Error generando PDF: {str(e)}')
        return redirect('quizzes:student_results')

@csrf_exempt
@login_required
def actualizar_puntos_pregunta(request, pk):
    """
    Vista para actualizar los puntos de una pregunta individual
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            pregunta = get_object_or_404(Pregunta, pk=pk)
            evaluacion = pregunta.evaluacion
            
            # Verificar si se pueden modificar las preguntas
            can_modify, restriction_message = check_question_modification_allowed(evaluacion)
            if not can_modify:
                return JsonResponse({
                    'success': False, 
                    'error': restriction_message
                }, status=403)
            
            puntos = data.get('puntos', 1)
            
            # Validar puntos
            try:
                puntos = int(puntos)
                if puntos < 1 or puntos > 10:
                    return JsonResponse({
                        'success': False, 
                        'error': 'Los puntos deben estar entre 1 y 10'
                    }, status=400)
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False, 
                    'error': 'Los puntos deben ser un número válido'
                }, status=400)
            
            # Actualizar puntos
            pregunta.puntos = puntos
            pregunta.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Puntos actualizados exitosamente',
                'puntos': puntos
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False, 
                'error': 'Datos JSON inválidos'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'error': f'Error al actualizar puntos: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False, 
        'error': 'Método no permitido'
    }, status=405)

@login_required
def send_participants_email(request, grupo_id):
    """
    Envía un correo al representante con la lista de participantes del grupo
    """
    if not can_manage_representantes(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('quizzes:dashboard')
    
    try:
        grupo = GrupoParticipantes.objects.select_related('representante').prefetch_related('participantes').get(id=grupo_id)
        
        if not grupo.representante:
            messages.error(request, 'Este grupo no tiene un representante asignado.')
            return redirect('quizzes:manage_grupos')
        
        if not grupo.participantes.exists():
            messages.error(request, 'Este grupo no tiene participantes asignados.')
            return redirect('quizzes:manage_grupos')
        
        # Crear tabla HTML moderna con los datos de los participantes
        participantes_html = """
        <table style="width: 100%; border-collapse: collapse; background: white;">
            <thead>
                <tr>
                    <th style="background: linear-gradient(135deg, #34495e 0%, #2c3e50 100%); color: white; padding: 15px 10px; text-align: left; font-weight: 600; font-size: 14px;">Cédula</th>
                    <th style="background: linear-gradient(135deg, #34495e 0%, #2c3e50 100%); color: white; padding: 15px 10px; text-align: left; font-weight: 600; font-size: 14px;">Nombres Completos</th>
                    <th style="background: linear-gradient(135deg, #34495e 0%, #2c3e50 100%); color: white; padding: 15px 10px; text-align: left; font-weight: 600; font-size: 14px;">Email</th>
                    <!--<th style="background: linear-gradient(135deg, #34495e 0%, #2c3e50 100%); color: white; padding: 15px 10px; text-align: left; font-weight: 600; font-size: 14px;">Teléfono</th>-->
                    <!--<th style="background: linear-gradient(135deg, #34495e 0%, #2c3e50 100%); color: white; padding: 15px 10px; text-align: left; font-weight: 600; font-size: 14px;">Edad</th>-->
                    <th style="background: linear-gradient(135deg, #34495e 0%, #2c3e50 100%); color: white; padding: 15px 10px; text-align: left; font-weight: 600; font-size: 14px;">Usuario</th>
                    <th style="background: linear-gradient(135deg, #34495e 0%, #2c3e50 100%); color: white; padding: 15px 10px; text-align: left; font-weight: 600; font-size: 14px;">Contraseña</th>
                </tr>
            </thead>
            <tbody>
        """
        
        for participante in grupo.participantes.all():
            nueva_password = get_random_string(length=6)
            participante.user.set_password(nueva_password)
            participante.user.save()
            
            participantes_html += f"""
                <tr style="border-bottom: 1px solid #e0e0e0;">
                    <td style="padding: 12px 10px; font-size: 13px; font-weight: 600; color: #333;">{participante.cedula}</td>
                    <td style="padding: 12px 10px; font-size: 13px; color: #555;">{participante.NombresCompletos}</td>
                    <td style="padding: 12px 10px; font-size: 13px; color: #667eea;">{participante.email}</td>
                    <!--<td style="padding: 12px 10px; font-size: 13px; color: #666;">{participante.phone or 'No registrado'}</td>-->
                    <!--<td style="padding: 12px 10px; font-size: 13px; color: #666;">{participante.edad or 'No registrado'}</td>-->
                    <td style="padding: 12px 10px; font-size: 13px; font-family: 'Courier New', monospace; font-weight: 600; color: #667eea;">{participante.user.username}</td>
                    <td style="padding: 12px 10px; font-size: 13px; font-family: 'Courier New', monospace; font-weight: 600; color: #e74c3c; background: #fdf2f2; border-radius: 4px;">{nueva_password}</td>
                </tr>
            """
        
        participantes_html += """
            </tbody>
        </table>
        """
        
        # Crear el mensaje del correo usando la función global
        subject = f'Lista de Participantes - Grupo: {grupo.name}'
        
        # Preparar contenido adicional para la función global
        additional_content = {
            'participantes_html': participantes_html,
            'total_participantes': grupo.participantes.count()
        }
        
        # Generar mensajes usando la función global
        plain_message, html_message = generate_email_messages(
            subject=subject,
            nombre=grupo.representante.NombresRepresentante,
            system_name=grupo.name,
            username='',  # No aplica para lista de participantes
            nueva_password='',  # No aplica para lista de participantes
            email_type='participants_list',
            additional_content=additional_content
        )
        
        # Enviar el correo
        send_mail(
            subject,
            plain_message,
            settings.EMAIL_HOST_USER,
            [grupo.representante.CorreoRepresentante],
            fail_silently=False,
            html_message=html_message
        )
        
        # Verificar si es una petición AJAX
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if is_ajax:
            # Para peticiones AJAX, devolver JSON
            return JsonResponse({
                'success': True,
                'message': f'Correo enviado exitosamente al representante {grupo.representante.NombresRepresentante}.'
            })
        else:
            # Para peticiones normales, usar mensajes y redirecciones
            messages.success(request, f'Correo enviado exitosamente al representante {grupo.representante.NombresRepresentante}.')
            return redirect('quizzes:manage_grupos')
        
    except GrupoParticipantes.DoesNotExist:
        error_msg = 'El grupo especificado no existe.'
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': error_msg
            })
        else:
            messages.error(request, error_msg)
            return redirect('quizzes:manage_grupos')
    except Exception as e:
        error_msg = f'Error al enviar el correo: {str(e)}'
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': error_msg
            })
        else:
            messages.error(request, error_msg)
            return redirect('quizzes:manage_grupos')

@login_required
def send_credentials_email(request, user_type, user_id):
    """Enviar correo con credenciales a un participante o administrador"""
    try:
        if user_type == 'participante':
            user_obj = Participantes.objects.get(id=user_id)
            nombre = user_obj.NombresCompletos
            email = user_obj.email
            username = user_obj.user.username
            
            # Siempre generar una nueva contraseña temporal
            nueva_password = get_random_string(length=6)
            user_obj.user.set_password(nueva_password)
            user_obj.user.save()
            
            subject = f'Credenciales de Acceso - Sistema Olymp'
            system_name = 'Sistema Olymp'
            
        elif user_type == 'admin':
            if not request.user.is_superuser:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': 'Solo el Superadministrador global puede enviar credenciales a administradores.'}, status=403)
                messages.error(request, 'Solo el Superadministrador global puede enviar credenciales a administradores.')
                return redirect('quizzes:manage_admins')

            user_obj = AdminProfile.objects.get(id=user_id)
            nombre = user_obj.user.get_full_name()
            email = user_obj.user.email
            username = user_obj.user.username
            
            # Siempre generar una nueva contraseña
            nueva_password = get_random_string(length=8)
            user_obj.user.set_password(nueva_password)
            user_obj.user.save()
            
            subject = f'Credenciales de Acceso - Panel de Administración Olymp'
            system_name = 'Panel de Administración Olymp'
            
        else:
            messages.error(request, 'Tipo de usuario no válido.')
            return redirect('quizzes:dashboard')
        
        # Generar mensajes usando la función global
        plain_message, html_message = generate_email_messages(
            subject=subject,
            nombre=nombre,
            system_name=system_name,
            username=username,
            nueva_password=nueva_password,
            email_type='credentials'
        )
        
        # Enviar el correo
        send_mail(
            subject,
            plain_message,
            settings.EMAIL_HOST_USER,
            [email],
            fail_silently=False,
            html_message=html_message
        )
        
        # Verificar si es una petición AJAX
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if is_ajax:
            # Para peticiones AJAX, devolver JSON
            return JsonResponse({
                'success': True,
                'message': f'Correo enviado exitosamente al {"participante" if user_type == "participante" else "administrador"} {nombre}.'
            })
        else:
            # Para peticiones normales, usar mensajes y redirecciones
            if user_type == 'participante':
                messages.success(request, f'Correo enviado exitosamente al participante {nombre}.')
                return redirect('quizzes:manage_participants')
            else:
                messages.success(request, f'Correo enviado exitosamente al administrador {nombre}.')
                return redirect('quizzes:manage_admins')
        
    except (Participantes.DoesNotExist, AdminProfile.DoesNotExist):
        error_msg = 'El usuario especificado no existe.'
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': error_msg
            })
        else:
            messages.error(request, error_msg)
            return redirect('quizzes:dashboard')
    except Exception as e:
        error_msg = f'Error al enviar el correo: {str(e)}'
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': error_msg
            })
        else:
            messages.error(request, error_msg)
            return redirect('quizzes:dashboard')


@login_required
def profile_view(request):
    """Vista para mostrar y editar el perfil del usuario"""
    # Obtener o crear el perfil del usuario
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    # Verificar si el usuario es un participante
    try:
        participante = Participantes.objects.get(user=request.user)
        is_participante = True
        # Para participantes, usar el teléfono del modelo Participantes
        phone_value = participante.phone
    except Participantes.DoesNotExist:
        is_participante = False
        # Para otros usuarios, usar el teléfono del UserProfile
        phone_value = profile.phone
    
    # Verificar si es una petición AJAX
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if request.method == 'POST':
        # Procesar el formulario de actualización de perfil
        if 'update_profile' in request.POST:
            try:
                # Actualizar información básica
                full_name = request.POST.get('full_name', '').strip()
                email = request.POST.get('email', '').strip()
                phone = request.POST.get('phone', '').strip()
                bio = request.POST.get('bio', '').strip()
                
                # Validar que se proporcione el nombre completo
                if not full_name:
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': 'El campo Nombres Completos es obligatorio.'
                        })
                    else:
                        messages.error(request, 'El campo Nombres Completos es obligatorio.')
                        return redirect('quizzes:profile')
                
                # Normalizar email siempre
                email_normalized = email.lower().strip()
                
                # Validar email único solo si cambió
                if email != request.user.email:
                    # Verificar si ya existe otro usuario con este correo
                    if User.objects.filter(email__iexact=email_normalized).exclude(id=request.user.id).exists():
                        if is_ajax:
                            return JsonResponse({
                                'success': False,
                                'message': 'El correo electrónico ya está en uso por otro usuario.'
                            })
                        else:
                            messages.error(request, 'El correo electrónico ya está en uso por otro usuario.')
                            return redirect('quizzes:profile')
                    
                    # Si es participante, verificar conflictos con representantes
                    if is_participante:
                        if Representante.objects.filter(
                            models.Q(CorreoInstitucional__iexact=email_normalized) | 
                            models.Q(CorreoRepresentante__iexact=email_normalized)
                        ).exists():
                            if is_ajax:
                                return JsonResponse({
                                    'success': False,
                                    'message': 'El correo electrónico ya está siendo usado por un representante.'
                                })
                            else:
                                messages.error(request, 'El correo electrónico ya está siendo usado por un representante.')
                                return redirect('quizzes:profile')
                
                # Validar formato de teléfono si se proporciona
                if phone and not re.match(r'^\d{10}$', phone):
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': 'El teléfono debe tener exactamente 10 dígitos numéricos.'
                        })
                    else:
                        messages.error(request, 'El teléfono debe tener exactamente 10 dígitos numéricos.')
                        return redirect('quizzes:profile')
                
                # Actualizar usuario - dividir el nombre completo en first_name y last_name
                # Si hay espacios, el primer espacio separa nombres de apellidos
                name_parts = full_name.split(' ', 1)
                if len(name_parts) > 1:
                    first_name = name_parts[0]
                    last_name = name_parts[1]
                else:
                    first_name = full_name
                    last_name = ''
                
                try:
                    # Usar transacción para asegurar consistencia
                    from django.db import transaction
                    with transaction.atomic():
                        request.user.first_name = first_name
                        request.user.last_name = last_name
                        request.user.email = email_normalized
                        
                        # Validar y guardar usuario
                        request.user.full_clean()
                        request.user.save()
                        
                        # Actualizar perfil según el tipo de usuario
                        if is_participante:
                            # Para participantes, actualizar el teléfono en el modelo Participantes
                            participante.phone = phone
                            participante.email = email_normalized  # Actualizar también el correo del participante
                            participante.full_clean()
                            participante.save()
                            # También actualizar el UserProfile para bio y avatar
                            profile.bio = bio
                        else:
                            # Para otros usuarios, actualizar el UserProfile
                            profile.phone = phone
                            profile.bio = bio
                        
                        # Procesar nueva foto si se subió
                        if 'avatar' in request.FILES:
                            # Eliminar foto anterior si existe
                            if profile.avatar:
                                try:
                                    os.remove(profile.avatar.path)
                                except:
                                    pass
                            
                            profile.avatar = request.FILES['avatar']
                        
                        profile.full_clean()
                        profile.save()
                        
                except ValidationError as e:
                    error_message = extract_validation_error_message(e)
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': f'Error de validación: {error_message}'
                        })
                    else:
                        messages.error(request, f'Error de validación: {error_message}')
                        return redirect('quizzes:profile')
                except IntegrityError as e:
                    if 'email' in str(e).lower():
                        error_msg = 'El correo electrónico ya está en uso.'
                    elif 'phone' in str(e).lower():
                        error_msg = 'El teléfono ya está registrado.'
                    else:
                        error_msg = 'Error de integridad en la base de datos.'
                    
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': error_msg
                        })
                    else:
                        messages.error(request, error_msg)
                        return redirect('quizzes:profile')
                
                # Si es AJAX, devolver JSON
                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'message': 'Perfil actualizado exitosamente.',
                        'avatar_url': profile.avatar.url if profile.avatar else None
                    })
                
                messages.success(request, 'Perfil actualizado exitosamente.')
                return redirect('quizzes:profile')
                
            except Exception as e:
                # Para peticiones AJAX, siempre devolver JSON
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'message': f'Error al actualizar el perfil: {str(e)}'
                    })
                else:
                    messages.error(request, f'Error al actualizar el perfil: {str(e)}')
                    return redirect('quizzes:profile')
        
        # Procesar cambio de contraseña
        elif 'change_password' in request.POST:
            try:
                current_password = request.POST.get('current_password')
                new_password = request.POST.get('new_password')
                confirm_password = request.POST.get('confirm_password')
                
                # Validar contraseña actual
                if not request.user.check_password(current_password):
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': 'La contraseña actual es incorrecta.'
                        })
                    else:
                        messages.error(request, 'La contraseña actual es incorrecta.')
                        return redirect('quizzes:profile')
                
                # Validar que las nuevas contraseñas coincidan
                if new_password != confirm_password:
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': 'Las nuevas contraseñas no coinciden.'
                        })
                    else:
                        messages.error(request, 'Las nuevas contraseñas no coinciden.')
                        return redirect('quizzes:profile')
                
                # Validar contraseña usando función reutilizable
                error_message = validate_password_strength(new_password, request.user.username)
                
                if error_message:
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': error_message
                        })
                    else:
                        messages.error(request, error_message)
                        return redirect('quizzes:profile')
                
                # Cambiar contraseña
                request.user.set_password(new_password)
                request.user.save()
                
                # Re-autenticar al usuario
                login(request, request.user)
                
                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'message': 'Contraseña cambiada exitosamente.'
                    })
                else:
                    messages.success(request, 'Contraseña cambiada exitosamente.')
                    return redirect('quizzes:profile')
                    
            except Exception as e:
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'message': f'Error al cambiar la contraseña: {str(e)}'
                    })
                else:
                    messages.error(request, f'Error al cambiar la contraseña: {str(e)}')
                    return redirect('quizzes:profile')
        else:
            # Si es AJAX pero no se reconoce el tipo de formulario
            if is_ajax:
                return JsonResponse({
                    'success': False,
                    'message': 'Tipo de formulario no reconocido.'
                })
    
    # Si es una petición AJAX GET, devolver error
    if is_ajax:
        return JsonResponse({
            'success': False,
            'message': 'Método no permitido para peticiones AJAX.'
        })
    
    context = {
        'profile': profile,
        'user': request.user,
        'phone_value': phone_value,
    }
    return render(request, 'quizzes/profile.html', context)


@login_required
def exportar_ranking_pdf(request, pk):
    """
    Genera un PDF del ranking de una evaluación específica con diseño moderno
    Muestra solo el mejor resultado de cada participante
    Respeta los filtros de estado aplicados
    """
    evaluacion = get_object_or_404(Evaluacion, pk=pk)
    
    # Verificar permisos
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        messages.error(request, 'No tienes permisos para acceder a esta funcionalidad.')
        return redirect('quizzes:dashboard')
    
    # Obtener el filtro de estado desde los parámetros GET
    filtro_estado = request.GET.get('estado', 'todos')
    
    # Obtener el mejor resultado por participante ordenado por nota (desc) y tiempo en segundos (asc)
    todos_completados = ResultadoEvaluacion.objects.filter(
        evaluacion=evaluacion,
        completada=True
    ).order_by('-puntos_obtenidos', 'tiempo_utilizado').select_related('participante')
    
    resultados = []
    vistos = set()
    for r in todos_completados:
        if r.participante_id not in vistos:
            vistos.add(r.participante_id)
            resultados.append(r)
    
    # Aplicar filtro de estado y agregar información del colegio
    from .models import GrupoParticipantes
    num_etapas = evaluacion.concurso.num_etapas if evaluacion.concurso else 3
    
    # Pre-cargar información de colegios para optimizar consultas
    participante_ids = [resultado.participante.id for resultado in resultados]
    grupos_info = {}
    grupos_queryset = GrupoParticipantes.objects.filter(
        participantes__id__in=participante_ids
    ).select_related('representante').prefetch_related('participantes')
    
    for grupo in grupos_queryset:
        for participante in grupo.participantes.all():
            if participante.id not in grupos_info:
                grupos_info[participante.id] = grupo.representante.NombreColegio if grupo.representante else None
    
    # Agregar posición real y colegio a cada resultado
    resultados_con_info = []
    for i, resultado in enumerate(resultados, 1):
        resultado.posicion_real = i
        resultado.participante.colegio_nombre = grupos_info.get(resultado.participante.id, None)
        resultados_con_info.append(resultado)
    
    # Aplicar filtro de estado
    resultados_filtrados = resultados_con_info
    titulo_filtro = "Ranking Completo"
    
    if filtro_estado == 'clasificados':
        titulo_filtro = "Participantes Clasificados"
        if evaluacion.etapa == 1:
            top_n = 15 if num_etapas == 3 else 5
            resultados_filtrados = resultados_con_info[:top_n]
        elif evaluacion.etapa == 2:
            if num_etapas == 3:
                resultados_filtrados = resultados_con_info[:5]
        elif evaluacion.etapa == 3:
            resultados_filtrados = resultados_con_info[:5]
    elif filtro_estado == 'no_clasificados':
        titulo_filtro = "Participantes No Clasificados"
        if evaluacion.etapa == 1:
            top_n = 15 if num_etapas == 3 else 5
            resultados_filtrados = resultados_con_info[top_n:] if len(resultados_con_info) > top_n else []
        elif evaluacion.etapa == 2:
            if num_etapas == 3:
                resultados_filtrados = resultados_con_info[5:] if len(resultados_con_info) > 5 else []
            else:
                resultados_filtrados = []
        elif evaluacion.etapa == 3:
            resultados_filtrados = resultados_con_info[5:] if len(resultados_con_info) > 5 else []
    
    # Crear el PDF
    response = HttpResponse(content_type='application/pdf')
    filename_suffix = f"_{filtro_estado}" if filtro_estado != 'todos' else ""
    response['Content-Disposition'] = f'attachment; filename="ranking_etapa_{evaluacion.etapa}_{evaluacion.title}{filename_suffix}.pdf"'
    
    # Crear el documento PDF con márgenes optimizados
    doc = SimpleDocTemplate(response, pagesize=A4, 
                          leftMargin=0.6*inch, rightMargin=0.6*inch,
                          topMargin=0.6*inch, bottomMargin=0.6*inch)
    elements = []
    
    # Definir paleta de colores moderna y elegante
    primary_color = colors.Color(0.15, 0.35, 0.75)  # Azul corporativo
    secondary_color = colors.Color(0.95, 0.95, 0.98)  # Gris muy claro
    accent_color = colors.Color(0.12, 0.55, 0.35)  # Verde moderno
    header_bg = colors.Color(0.08, 0.25, 0.55)  # Azul más oscuro para encabezado
    gold_color = colors.Color(1, 0.84, 0)  # Oro
    silver_color = colors.Color(0.75, 0.75, 0.75)  # Plata
    bronze_color = colors.Color(0.80, 0.50, 0.20)  # Bronce
    
    # Estilos mejorados
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'ModernTitle',
        parent=styles['Heading1'],
        fontSize=22,
        spaceAfter=18,
        alignment=1,  # Centrado
        textColor=primary_color,
        fontName='Helvetica-Bold',
        spaceBefore=8,
        leading=26
    )
    
    subtitle_style = ParagraphStyle(
        'ModernSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        alignment=1,  # Centrado
        textColor=colors.Color(0.3, 0.3, 0.3),
        fontName='Helvetica',
        leading=18
    )
    
    etapa_style = ParagraphStyle(
        'EtapaStyle',
        parent=styles['Heading2'],
        fontSize=18,
        spaceAfter=20,
        alignment=1,  # Centrado
        textColor=accent_color,
        fontName='Helvetica-Bold',
        spaceBefore=10,
        leading=22
    )
    
    # Título principal con mejor formato
    title = Paragraph("OLIMPIADAS DE MATEMÁTICAS<br/>CARRERA MECÁNICA", title_style)
    elements.append(title)
    
    # Manejo mejorado de logos
    logo_mecanica = None
    logo_uteq = None
    
    # Cargar Logo Mecánica
    logo_mecanica_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logoMecanica.png')
    if os.path.exists(logo_mecanica_path):
        try:
            logo_mecanica = Image(logo_mecanica_path, width=1.6*inch, height=1.2*inch)
        except Exception as e:
            print(f"Error al cargar logo Mecánica: {e}")
    
    # Cargar Logo UTEQ
    logo_uteq_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo-uteq.png')
    if os.path.exists(logo_uteq_path):
        try:
            logo_uteq = Image(logo_uteq_path, width=1.6*inch, height=1.2*inch)
        except Exception as e:
            print(f"Error al cargar logo UTEQ: {e}")
    
    # Tabla de logos mejorada
    if logo_mecanica or logo_uteq:
        logo_row = []
        
        if logo_mecanica:
            logo_row.append(logo_mecanica)
        else:
            logo_row.append(Paragraph("", styles['Normal']))
        
        logo_row.append(Paragraph("", styles['Normal']))  # Espacio central
        
        if logo_uteq:
            logo_row.append(logo_uteq)
        else:
            logo_row.append(Paragraph("", styles['Normal']))
        
        logo_table = Table([logo_row], colWidths=[2.4*inch, 1.2*inch, 2.4*inch])
        logo_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, 0), 0),
            ('RIGHTPADDING', (0, 0), (-1, 0), 0),
            ('TOPPADDING', (0, 0), (-1, 0), 5),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
        ]))
        
        elements.append(logo_table)
        elements.append(Spacer(1, 20))
    
    # Información de la etapa
    etapa_text = f"RESULTADOS OFICIALES - ETAPA {evaluacion.etapa}"
    elements.append(Paragraph(etapa_text, etapa_style))
    
    # Información adicional
    fecha_info = f"Evaluación: {evaluacion.title}"
    elements.append(Paragraph(fecha_info, subtitle_style))
    elements.append(Spacer(1, 15))
    
    # Subtítulo del ranking con filtro aplicado
    ranking_subtitle = Paragraph(titulo_filtro, subtitle_style)
    elements.append(ranking_subtitle)
    elements.append(Spacer(1, 15))
    
    # Preparar datos de la tabla con columna de colegio
    table_data = [['Pos.', 'Participante', 'Cédula', 'Colegio', 'Puntaje', 'Tiempo', 'Estado']]
    
    for resultado in resultados_filtrados:
        # Determinar estado según la etapa y configuración
        pos = resultado.posicion_real
        if evaluacion.etapa == 1 and pos <= (15 if num_etapas == 3 else 5):
            estado = "Clasificado"
        elif evaluacion.etapa == 2 and num_etapas == 3 and pos <= 5:
            estado = "Finalista"
        elif evaluacion.etapa == 3:
            if pos == 1:
                estado = "Oro"
            elif pos == 2:
                estado = "Plata"
            elif pos == 3:
                estado = "Bronce"
            else:
                estado = "Participante"
        else:
            estado = "Participante"
        
        # Formatear puntaje - usar puntos_obtenidos directamente
        puntaje_str = f"{resultado.puntos_obtenidos:.3f}"
        
        # Formatear nombre del colegio para el PDF (máximo 30 caracteres)
        colegio_nombre = resultado.participante.colegio_nombre or "Sin colegio"
        if len(colegio_nombre) > 30:
            colegio_nombre = colegio_nombre[:27] + "..."
        
        # Agregar fila a la tabla
        table_data.append([
            str(pos),
            resultado.participante.NombresCompletos,
            resultado.participante.cedula,
            colegio_nombre,
            puntaje_str,
            resultado.get_tiempo_formateado(),
            estado
        ])
    
    # Crear tabla con anchos optimizados para incluir columna de colegio
    table = Table(table_data, colWidths=[0.5*inch, 2.2*inch, 1.0*inch, 1.8*inch, 0.7*inch, 0.9*inch, 1.1*inch])
    
    # Estilo moderno y elegante de la tabla
    table_style = TableStyle([
        # Encabezado principal
        ('BACKGROUND', (0, 0), (-1, 0), header_bg),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Cuerpo de la tabla
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        
        # Bordes elegantes
        ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.7, 0.7, 0.7)),
        ('LINEBELOW', (0, 0), (-1, 0), 2, header_bg),
        
        # Alineación actualizada para 7 columnas
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Nombres a la izquierda
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Posición centrada
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # Cédula centrada
        ('ALIGN', (3, 1), (3, -1), 'LEFT'),    # Colegio a la izquierda
        ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # Puntaje centrado
        ('ALIGN', (5, 1), (5, -1), 'CENTER'),  # Tiempo centrado
        ('ALIGN', (6, 1), (6, -1), 'CENTER'),  # Estado centrado
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    
    # Aplicar filas alternadas correctamente
    for i in range(1, len(table_data)):
        if i % 2 == 0:  # Filas pares (índice par, pero es fila impar visualmente)
            table_style.add('BACKGROUND', (0, i), (-1, i), secondary_color)
        else:  # Filas impares (índice impar, pero es fila par visualmente)
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.white)
    
    # Aplicar colores especiales para estados de medalleros (columna 6 ahora)
    for resultado in resultados_filtrados:
        # Encontrar la fila correspondiente en la tabla
        pos = resultado.posicion_real
        row_index = None
        for i, row in enumerate(table_data[1:], 1):  # Empezar desde 1 para saltar header
            if int(row[0]) == pos:  # Comparar posición
                row_index = i
                break
        
        if row_index is not None:
            if evaluacion.etapa == 3:
                if pos == 1:  # Oro
                    table_style.add('BACKGROUND', (6, row_index), (6, row_index), gold_color)
                    table_style.add('TEXTCOLOR', (6, row_index), (6, row_index), colors.black)
                    table_style.add('FONTNAME', (6, row_index), (6, row_index), 'Helvetica-Bold')
                elif pos == 2:  # Plata
                    table_style.add('BACKGROUND', (6, row_index), (6, row_index), silver_color)
                    table_style.add('TEXTCOLOR', (6, row_index), (6, row_index), colors.black)
                    table_style.add('FONTNAME', (6, row_index), (6, row_index), 'Helvetica-Bold')
                elif pos == 3:  # Bronce
                    table_style.add('BACKGROUND', (6, row_index), (6, row_index), bronze_color)
                    table_style.add('TEXTCOLOR', (6, row_index), (6, row_index), colors.white)
                    table_style.add('FONTNAME', (6, row_index), (6, row_index), 'Helvetica-Bold')
            elif (evaluacion.etapa == 1 and pos <= (15 if num_etapas == 3 else 5)) or \
                 (evaluacion.etapa == 2 and num_etapas == 3 and pos <= 5):
                table_style.add('BACKGROUND', (6, row_index), (6, row_index), accent_color)
                table_style.add('TEXTCOLOR', (6, row_index), (6, row_index), colors.white)
                table_style.add('FONTNAME', (6, row_index), (6, row_index), 'Helvetica-Bold')
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Agregar información de estadísticas al final con datos filtrados
    if resultados_filtrados:
        elements.append(Spacer(1, 20))
        
        stats_style = ParagraphStyle(
            'StatsStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,  # Centrado
            textColor=colors.Color(0.4, 0.4, 0.4),
            fontName='Helvetica'
        )
        
        total_mostrados = len(resultados_filtrados)
        total_general = len(resultados_con_info)
        promedio_puntos = sum(r.puntos_obtenidos for r in resultados_filtrados) / total_mostrados
        
        if filtro_estado == 'todos':
            stats_text = f"Total de participantes: {total_mostrados} | Promedio general: {promedio_puntos:.1f} puntos"
        else:
            stats_text = f"Mostrando {total_mostrados} de {total_general} participantes | Promedio mostrado: {promedio_puntos:.1f} puntos"
        
        elements.append(Paragraph(stats_text, stats_style))
    
    # Construir el PDF
    doc.build(elements)
    
    return response

@login_required 
def exportar_resultados(request, pk):
    """
    Exporta los resultados de una evaluación a un archivo Excel
    Respeta los filtros por grupo si se especifican
    """
    evaluacion = get_object_or_404(Evaluacion, pk=pk)
    
    # Verificar permisos
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        return JsonResponse({'error': 'No tienes permisos para esta acción'}, status=403)
    
    # Obtener filtro de grupo desde los parámetros GET
    grupo_filtro = request.GET.get('grupo', 'todos')
    grupo_seleccionado = None
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from django.utils.html import strip_tags
        from datetime import datetime
        import html
        
        # Filtrar resultados según el grupo seleccionado
        if grupo_filtro != 'todos' and grupo_filtro.isdigit():
            grupo_seleccionado = get_object_or_404(GrupoParticipantes, id=int(grupo_filtro))
            participantes_filtrados = grupo_seleccionado.participantes.all()
            
            # Filtrar resultados por participantes del grupo
            resultados_completados = ResultadoEvaluacion.objects.filter(
                evaluacion=evaluacion, 
                completada=True,
                participante__in=participantes_filtrados
            )
            todos_resultados = ResultadoEvaluacion.objects.filter(
                evaluacion=evaluacion,
                participante__in=participantes_filtrados
            )
        else:
            # Obtener todos los resultados
            resultados_completados = ResultadoEvaluacion.objects.filter(evaluacion=evaluacion, completada=True)
            todos_resultados = ResultadoEvaluacion.objects.filter(evaluacion=evaluacion)
        
        # Crear un nuevo workbook
        wb = Workbook()
        
        # === HOJA 1: RESUMEN GENERAL ===
        ws_resumen = wb.active
        ws_resumen.title = "Resumen General"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=14)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Título principal
        titulo_reporte = f"REPORTE DE RESULTADOS - {evaluacion.title}"
        if grupo_seleccionado:
            titulo_reporte += f" - {grupo_seleccionado.name}"
            
        ws_resumen.merge_cells('A1:F1')
        ws_resumen['A1'] = titulo_reporte
        ws_resumen['A1'].font = title_font
        ws_resumen['A1'].alignment = Alignment(horizontal='center')
        
        # Información general
        ws_resumen['A3'] = "Información General"
        ws_resumen['A3'].font = header_font
        ws_resumen['A3'].fill = header_fill
        
        ws_resumen['A4'] = "Título:"
        ws_resumen['B4'] = evaluacion.title
        ws_resumen['A5'] = "Etapa:"
        ws_resumen['B5'] = f"Etapa {evaluacion.etapa}"
        ws_resumen['A6'] = "Fecha de inicio:"
        ws_resumen['B6'] = evaluacion.start_time.strftime("%d/%m/%Y %H:%M")
        ws_resumen['A7'] = "Fecha de fin:"
        ws_resumen['B7'] = evaluacion.end_time.strftime("%d/%m/%Y %H:%M")
        ws_resumen['A8'] = "Duración:"
        ws_resumen['B8'] = f"{evaluacion.duration_minutes} minutos"
        
        if grupo_seleccionado:
            ws_resumen['A9'] = "Filtro aplicado:"
            ws_resumen['B9'] = f"Grupo: {grupo_seleccionado.name}"
        
        # Estadísticas de participación
        participantes_completaron = resultados_completados.values_list('participante', flat=True).distinct().count()
        participantes_con_resultados = todos_resultados.values_list('participante', flat=True).distinct().count()
        participantes_en_progreso = todos_resultados.filter(completada=False).values_list('participante', flat=True).distinct().count()
        
        ws_resumen['D3'] = "Estadísticas de Participación"
        ws_resumen['D3'].font = header_font
        ws_resumen['D3'].fill = header_fill
        
        ws_resumen['D4'] = "Participantes que completaron:"
        ws_resumen['E4'] = participantes_completaron
        ws_resumen['D5'] = "Participantes en progreso:"
        ws_resumen['E5'] = participantes_en_progreso
        ws_resumen['D6'] = "Total con intentos:"
        ws_resumen['E6'] = participantes_con_resultados
        
        # Estadísticas de rendimiento
        if resultados_completados.exists():
            from django.db.models import Avg, Max, Min
            stats = resultados_completados.aggregate(
                promedio=Avg('puntos_obtenidos'),
                maximo=Max('puntos_obtenidos'),
                minimo=Min('puntos_obtenidos'),
                tiempo_promedio=Avg('tiempo_utilizado')
            )
            
            row_start = 11 if grupo_seleccionado else 10
            ws_resumen[f'A{row_start}'] = "Estadísticas de Rendimiento"
            ws_resumen[f'A{row_start}'].font = header_font
            ws_resumen[f'A{row_start}'].fill = header_fill
            
            ws_resumen[f'A{row_start + 1}'] = "Promedio:"
            ws_resumen[f'B{row_start + 1}'] = f"{stats['promedio']:.2f}/10" if stats['promedio'] else "N/A"
            ws_resumen[f'A{row_start + 2}'] = "Mejor puntaje:"
            ws_resumen[f'B{row_start + 2}'] = f"{stats['maximo']:.2f}/10" if stats['maximo'] else "N/A"
            ws_resumen[f'A{row_start + 3}'] = "Peor puntaje:"
            ws_resumen[f'B{row_start + 3}'] = f"{stats['minimo']:.2f}/10" if stats['minimo'] else "N/A"
            ws_resumen[f'A{row_start + 4}'] = "Tiempo promedio:"
            ws_resumen[f'B{row_start + 4}'] = f"{int(stats['tiempo_promedio'])} min" if stats['tiempo_promedio'] else "N/A"
        
        # === HOJA 2: RESULTADOS DETALLADOS ===
        ws_resultados = wb.create_sheet("Resultados Detallados")
        
        # Encabezados
        headers = ['#', 'Participante', 'Cédula', 'Puntaje', 'Porcentaje', 'Tiempo (min)', 
                  'Intento', 'Fecha Inicio', 'Fecha Fin', 'Estado']
        
        if grupo_seleccionado:
            headers.insert(3, 'Grupo')
        
        for col, header in enumerate(headers, 1):
            cell = ws_resultados.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos de resultados
        resultados = todos_resultados.order_by('-puntos_obtenidos', 'tiempo_utilizado')
        
        for row, resultado in enumerate(resultados, 2):
            col = 1
            ws_resultados.cell(row=row, column=col, value=row-1)
            col += 1
            ws_resultados.cell(row=row, column=col, value=resultado.participante.NombresCompletos)
            col += 1
            ws_resultados.cell(row=row, column=col, value=resultado.participante.cedula)
            col += 1
            
            if grupo_seleccionado:
                ws_resultados.cell(row=row, column=col, value=grupo_seleccionado.name)
                col += 1
            
            ws_resultados.cell(row=row, column=col, value=f"{resultado.puntos_obtenidos:.2f}")
            col += 1
            ws_resultados.cell(row=row, column=col, value=f"{resultado.get_puntaje_porcentaje():.1f}%")
            col += 1
            ws_resultados.cell(row=row, column=col, value=resultado.tiempo_utilizado)
            col += 1
            ws_resultados.cell(row=row, column=col, value=resultado.numero_intento)
            col += 1
            ws_resultados.cell(row=row, column=col, value=resultado.fecha_inicio.strftime("%d/%m/%Y %H:%M") if resultado.fecha_inicio else "N/A")
            col += 1
            ws_resultados.cell(row=row, column=col, value=resultado.fecha_fin.strftime("%d/%m/%Y %H:%M") if resultado.fecha_fin else "N/A")
            col += 1
            ws_resultados.cell(row=row, column=col, value="Completado" if resultado.completada else "En progreso")
            
            # Aplicar bordes
            for col_idx in range(1, len(headers) + 1):
                ws_resultados.cell(row=row, column=col_idx).border = border
        
        # === HOJA 3: ANÁLISIS POR PREGUNTA ===
        ws_preguntas = wb.create_sheet("Análisis por Pregunta")
        
        # Encabezados
        headers_preguntas = ['Pregunta', 'Texto', 'Correctas', 'Incorrectas', 'Sin Responder', '% Acierto', 'Dificultad']
        
        for col, header in enumerate(headers_preguntas, 1):
            cell = ws_preguntas.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos por pregunta
        for row, pregunta in enumerate(evaluacion.preguntas.all(), 2):
            correctas = 0
            incorrectas = 0
            sin_responder = 0
            
            for resultado in resultados_completados:
                respuestas = resultado.respuestas_guardadas
                pregunta_id = str(pregunta.id)
                
                if pregunta_id in respuestas:
                    try:
                        opcion = pregunta.opciones.get(id=respuestas[pregunta_id])
                        if opcion.is_correct:
                            correctas += 1
                        else:
                            incorrectas += 1
                    except:
                        sin_responder += 1
                else:
                    sin_responder += 1
            
            total = correctas + incorrectas + sin_responder
            porcentaje = (correctas / total * 100) if total > 0 else 0
            dificultad = 'Fácil' if porcentaje > 70 else 'Media' if porcentaje > 40 else 'Difícil'
            
            # Limpiar el texto HTML de la pregunta y decodificar entidades HTML
            texto_sin_html = strip_tags(pregunta.text)
            texto_limpio = html.unescape(texto_sin_html).strip()
            # Limitar el largo del texto para que no desborde la celda
            texto_pregunta = texto_limpio[:150] + "..." if len(texto_limpio) > 150 else texto_limpio
            
            ws_preguntas.cell(row=row, column=1, value=f"P{row-1}")
            ws_preguntas.cell(row=row, column=2, value=texto_pregunta)
            ws_preguntas.cell(row=row, column=3, value=correctas)
            ws_preguntas.cell(row=row, column=4, value=incorrectas)
            ws_preguntas.cell(row=row, column=5, value=sin_responder)
            ws_preguntas.cell(row=row, column=6, value=f"{porcentaje:.1f}%")
            ws_preguntas.cell(row=row, column=7, value=dificultad)
            
            # Aplicar bordes
            for col in range(1, len(headers_preguntas) + 1):
                ws_preguntas.cell(row=row, column=col).border = border
        
        # Ajustar ancho de columnas
        for ws in [ws_resumen, ws_resultados, ws_preguntas]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Nombre del archivo
        filename = f'resultados_{evaluacion.title.replace(" ", "_")}'
        if grupo_seleccionado:
            filename += f'_{grupo_seleccionado.name.replace(" ", "_")}'
        filename += f'_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({'error': f'Error al generar el archivo: {str(e)}'}, status=500)

# ============================================================================
# VISTAS PARA MONITOREO EN TIEMPO REAL
# ============================================================================

@login_required
def monitoreo_evaluacion(request, pk):
    """
    Vista principal para el monitoreo en tiempo real de una evaluación
    """
    # Verificar permisos básicos (solo admins pueden acceder al monitoreo)
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        messages.error(request, 'No tienes permisos para acceder al monitoreo en tiempo real.')
        return redirect('quizzes:dashboard')
    
    evaluacion = get_evaluacion_monitoreable_or_404(request, pk)
    participantes_autorizados = evaluacion.get_participantes_autorizados()
    total_participantes = len(participantes_autorizados)
    
    resultados = ResultadoEvaluacion.objects.filter(evaluacion=evaluacion).select_related('participante')
    ultimos_resultados = {}
    for resultado in resultados.order_by('participante_id', '-numero_intento'):
        ultimos_resultados.setdefault(resultado.participante_id, resultado)
    participantes_activos = sum(1 for r in ultimos_resultados.values() if r.esta_activo())
    participantes_finalizados = sum(1 for r in ultimos_resultados.values() if r.completada)
    
    context = {
        'evaluacion': evaluacion,
        'resultados': resultados,
        'total_participantes': total_participantes,
        'participantes_activos': participantes_activos,
        'participantes_finalizados': participantes_finalizados,
        'participantes_pendientes': max(0, total_participantes - participantes_activos - participantes_finalizados),
    }
    
    return render(request, 'quizzes/monitoreo_evaluacion.html', context)


@login_required
def actualizar_monitoreo(request, pk):
    """
    Endpoint de compatibilidad HTTP para actualizar actividad desde el frontend
    """
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        data = json.loads(request.body)
        participante_id = data.get('participante_id')
        evaluacion_id = data.get('evaluacion_id') or pk
        
        participante = get_object_or_404(Participantes, pk=participante_id)
        if str(evaluacion_id) != str(pk):
            return JsonResponse({'error': 'La evaluación no coincide con la URL'}, status=400)
        evaluacion = get_evaluacion_monitoreable_or_404(request, pk)
        
        resultado = ResultadoEvaluacion.objects.filter(
            evaluacion=evaluacion,
            participante=participante,
            completada=False
        ).first()
        
        if resultado:
            resultado.ultima_actividad = timezone.now()
            resultado.save()
            return JsonResponse({
                'success': True,
                'resultado_id': resultado.id,
                'ultima_actividad': resultado.ultima_actividad.isoformat()
            })
            
        return JsonResponse({'success': True, 'message': 'No hay examen activo en progreso'})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def obtener_estado_monitoreo(request, pk):
    """
    Endpoint HTTP AJAX Polling para obtener el estado actual del monitoreo en tiempo real
    """
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    evaluacion = get_evaluacion_monitoreable_or_404(request, pk)
    participantes_autorizados = evaluacion.get_participantes_autorizados()
    datos_monitoreo = []
    participante_ids = [participante.id for participante in participantes_autorizados]
    resultados = ResultadoEvaluacion.objects.filter(
        evaluacion=evaluacion, participante_id__in=participante_ids
    ).order_by('participante_id', '-numero_intento')
    configuraciones = {
        configuracion.participante_id: configuracion.intentos_maximos
        for configuracion in IntentosParticipante.objects.filter(
            evaluacion=evaluacion, participante_id__in=participante_ids
        )
    }
    ultimos_resultados = {}
    intentos_usados_por_participante = {}
    for resultado in resultados:
        ultimos_resultados.setdefault(resultado.participante_id, resultado)
        if resultado.completada:
            intentos_usados_por_participante[resultado.participante_id] = (
                intentos_usados_por_participante.get(resultado.participante_id, 0) + 1
            )

    total_banco_preguntas = evaluacion.preguntas.count()
    cantidad_configurada = sum(
        cuota.cantidad_preguntas for cuota in evaluacion.cuotas_unidades.all()
    ) or evaluacion.preguntas_a_mostrar or 10
    total_preguntas_mostradas = min(total_banco_preguntas, cantidad_configurada)

    for participante in participantes_autorizados:
        resultado = ultimos_resultados.get(participante.id)
        intentos_usados = intentos_usados_por_participante.get(participante.id, 0)
        intentos_maximos = configuraciones.get(
            participante.id, participante.intentos_maximos_default
        )
        intentos_disponibles = max(0, intentos_maximos - intentos_usados)
        ha_iniciado = resultado is not None
        
        if not resultado:
            estado = 'pendiente'
        elif resultado.completada:
            estado = 'finalizado' if intentos_disponibles == 0 else 'inactivo'
        elif resultado.esta_activo():
            estado = 'activo'
        else:
            estado = 'inactivo'
            
        respuestas = obtener_diccionario_respuestas(resultado)
        preguntas_respondidas = min(total_preguntas_mostradas, sum(
            1 for clave, respuesta in respuestas.items()
            if clave.startswith('pregunta_') and respuesta
        ))
        porcentaje_avance = round(
            (preguntas_respondidas / total_preguntas_mostradas * 100), 1
        ) if total_preguntas_mostradas else 0
        
        alertas = resultado.alertas_detectadas if (resultado and resultado.alertas_detectadas) else []
        
        datos_monitoreo.append({
            # La clave visual debe ser estable por participante, no por intento.
            'id': participante.id,
            'resultado_id': resultado.id if resultado else None,
            'participante_id': participante.id,
            'participante_nombre': participante.NombresCompletos,
            'participante_cedula': participante.cedula,
            'estado': estado,
            'esta_activo': resultado.esta_activo() if resultado else False,
            'ha_iniciado': ha_iniciado,
            'preguntas_respondidas': preguntas_respondidas,
            'preguntas_revisadas': total_preguntas_mostradas,
            'porcentaje_avance': porcentaje_avance,
            'ultima_actividad': resultado.ultima_actividad.isoformat() if (resultado and resultado.ultima_actividad) else None,
            'alertas_count': len(alertas),
            'alertas_recientes': alertas[-3:],
            'tiene_resultado_completado': (resultado.completada) if resultado else False,
            'puntos_obtenidos': float(resultado.puntos_obtenidos) if (resultado and resultado.completada) else None,
            'puntaje_numerico': resultado.get_puntaje_numerico() if (resultado and resultado.completada) else None,
            'intentos_disponibles': intentos_disponibles,
            'intentos_usados': intentos_usados,
            'cambios_pestana_actuales': resultado.cambios_pestana if resultado else 0,
            'cambios_pestana_maximo': 4
        })
    
    return JsonResponse({
        'monitoreos': datos_monitoreo,
        'timestamp': timezone.now().isoformat()
    })


@login_required
def finalizar_evaluacion_admin(request, pk):
    """
    Endpoint HTTP POST para finalizar una evaluación por decisión administrativa
    """
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        data = json.loads(request.body)
        monitoreo_id = data.get('monitoreo_id') or data.get('resultado_id')
        motivo = data.get('motivo', 'Finalización administrativa')
        
        evaluacion = get_evaluacion_monitoreable_or_404(request, pk)
        
        # Buscar por ID de ResultadoEvaluacion o por participante
        resultado = ResultadoEvaluacion.objects.filter(pk=monitoreo_id, evaluacion=evaluacion).first()
        if not resultado:
            resultado = ResultadoEvaluacion.objects.filter(
                evaluacion=evaluacion,
                participante_id=data.get('participante_id')
            ).order_by('-numero_intento').first()
            
        if not resultado:
            return JsonResponse({'error': 'No se encontró un resultado de evaluación para finalizar'}, status=404)
            
        resultado.finalizar_por_admin(request.user, motivo)
        
        return JsonResponse({
            'success': True,
            'message': f'Evaluación de {resultado.participante.NombresCompletos} finalizada exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def detalle_monitoreo(request, monitoreo_id):
    """
    Vista para ver el detalle completo de auditoría de un resultado de evaluación
    """
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        messages.error(request, 'No tienes permisos para acceder a esta funcionalidad.')
        return redirect('quizzes:dashboard')
    
    resultado_seleccionado = get_resultado_monitoreable_or_404(request, monitoreo_id)
    intentos = ResultadoEvaluacion.objects.filter(
        evaluacion=resultado_seleccionado.evaluacion,
        participante=resultado_seleccionado.participante,
    ).select_related('finalizado_por_admin').order_by('-numero_intento')

    # El detalle siempre representa el intento más reciente como estado actual,
    # incluso si se llega mediante la URL de un intento histórico.
    resultado = intentos.first()
    
    context = {
        'resultado': resultado,
        'evaluacion': resultado.evaluacion,
        'participante': resultado.participante,
        'alertas': resultado.alertas_detectadas or [],
        'intentos': intentos,
        'total_intentos': intentos.count(),
    }
    
    return render(request, 'quizzes/detalle_monitoreo.html', context)


@login_required
def agregar_alerta_manual(request, monitoreo_id):
    """
    Endpoint para agregar alertas manuales desde el panel de administración
    """
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        data = json.loads(request.body)
        tipo_alerta = data.get('tipo_alerta', 'manual')
        descripcion = data.get('descripcion', '')
        severidad = data.get('severidad', 'baja')
        
        resultado = get_resultado_monitoreable_or_404(request, monitoreo_id)
        resultado.agregar_alerta(tipo_alerta, descripcion, severidad)
        
        return JsonResponse({
            'success': True,
            'message': 'Alerta agregada exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def send_credentials_for_clave_temporal(tipo_usuario, user_id):
    """
    Función auxiliar para generar nueva contraseña y enviar email de credenciales
    para la funcionalidad de solicitar clave temporal.
    
    Args:
        tipo_usuario (str): 'participante' o 'admin'
        user_id (int): ID del usuario
    
    Returns:
        tuple: (success, message, email) donde success es bool, message es str, email es str
    """
    try:
        if tipo_usuario == 'participante':
            user_obj = Participantes.objects.get(id=user_id)
            nombre = user_obj.NombresCompletos
            email = user_obj.email
            username = user_obj.user.username
            
            # Generar nueva contraseña temporal
            nueva_password = get_random_string(length=6)
            user_obj.user.set_password(nueva_password)
            user_obj.user.save()
            
            subject = f'Credenciales de Acceso - Sistema Olymp'
            system_name = 'Sistema Olymp'
            
        else:  # admin
            user_obj = AdminProfile.objects.get(id=user_id)
            nombre = user_obj.user.get_full_name()
            email = user_obj.user.email
            username = user_obj.user.username
            
            # Generar nueva contraseña
            nueva_password = get_random_string(length=8)
            user_obj.user.set_password(nueva_password)
            user_obj.user.save()
            
            subject = f'Credenciales de Acceso - Panel de Administración Olymp'
            system_name = 'Panel de Administración Olymp'
        
        # Generar mensajes usando la función global
        plain_message, html_message = generate_email_messages(
            subject=subject,
            nombre=nombre,
            system_name=system_name,
            username=username,
            nueva_password=nueva_password,
            email_type='credentials'
        )
        
        # Enviar el correo
        send_mail(
            subject,
            plain_message,
            settings.EMAIL_HOST_USER,
            [email],
            fail_silently=False,
            html_message=html_message
        )
        
        return True, f'Se ha enviado una nueva contraseña temporal a su correo electrónico ({email}).', email
        
    except Exception as email_error:
        return False, f'Error al enviar el correo: {str(email_error)}', email if 'email' in locals() else None


@csrf_exempt
def solicitar_clave_temporal(request):
    """
    Endpoint para solicitar una nueva clave temporal
    No requiere login ya que es para usuarios que olvidaron su contraseña
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)
    
    try:
        data = json.loads(request.body)
        username = data.get('username', '').strip()
        
        if not username:
            return JsonResponse({
                'success': False, 
                'message': 'El nombre de usuario es requerido.'
            })
        
        # Buscar el usuario (puede ser admin o participante)
        tipo_usuario = None
        email = None
        user_id = None
        
        # Primero buscar como administrador
        try:
            admin_profile = AdminProfile.objects.get(user__username__iexact=username)
            tipo_usuario = 'admin'
            email = admin_profile.user.email
            user_id = admin_profile.id
        except AdminProfile.DoesNotExist:
            # Si no es admin, buscar como participante
            try:
                participante = Participantes.objects.get(user__username__iexact=username)
                tipo_usuario = 'participante'
                email = participante.email
                user_id = participante.id
            except Participantes.DoesNotExist:
                # Intentar buscar por cédula (solo para participantes)
                try:
                    participante = Participantes.objects.get(cedula=username)
                    tipo_usuario = 'participante'
                    email = participante.email
                    user_id = participante.id
                except Participantes.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'message': f'El usuario "{username}" no existe en el sistema.'
                    })
        
        # Verificar si puede solicitar (máximo 3 por semana)
        solicitudes_semana = SolicitudClaveTemporal.contar_solicitudes_semana(username, tipo_usuario)
        if solicitudes_semana >= 3:
            return JsonResponse({
                'success': False,
                'message': f'Ha excedido el límite de solicitudes. Ya ha solicitado {solicitudes_semana} veces esta semana. El límite es de 3 solicitudes por semana.'
            })
        
        # Crear registro de solicitud
        solicitud = SolicitudClaveTemporal.objects.create(
            username=username,
            tipo_usuario=tipo_usuario,
            email=email
        )
        
        # Generar nueva contraseña y enviar email usando la función auxiliar
        success, message, email_sent = send_credentials_for_clave_temporal(tipo_usuario, user_id)
        
        if success:
            # Marcar solicitud como procesada
            solicitud.procesada = True
            solicitud.save()
            
            return JsonResponse({
                'success': True,
                'message': message
            })
        else:
            # Si hay error en el envío, marcar como fallida
            solicitud.procesada = False
            solicitud.mensaje_error = message
            solicitud.save()
            
            return JsonResponse({
                'success': False,
                'message': message
            })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Error en el formato de datos enviados.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })


@login_required
@full_access_required
def settings_view(request):
    """Centro de Configuración General, Organizacional y Curricular del Sistema."""
    from .models import Facultad, Carrera, Concurso, UnidadTematica, Tema
    
    if request.method == 'POST':
        action = request.POST.get('action')
        try:
            if action in ['crear_facultad', 'crear_carrera'] and not request.user.is_superuser:
                messages.error(request, 'Solo el Superadministrador global tiene permisos para gestionar Facultades y Carreras.')
                return redirect('quizzes:settings')

            if action == 'crear_facultad':
                nombre = request.POST.get('nombre', '').strip()
                siglas = request.POST.get('siglas', '').strip()
                if nombre:
                    facultad, created = Facultad.objects.get_or_create(nombre=nombre, defaults={'siglas': siglas})
                    if not created:
                        facultad.siglas = siglas
                        facultad.save()
                    messages.success(request, f'Facultad "{nombre}" guardada correctamente.')
                return redirect('/configuracion/?tab=facultades')
                
            elif action == 'crear_carrera':
                facultad_id = request.POST.get('facultad_id')
                nombre = request.POST.get('nombre', '').strip()
                codigo = request.POST.get('codigo', '').strip()
                if facultad_id and nombre:
                    facultad = Facultad.objects.get(id=facultad_id)
                    Carrera.objects.create(facultad=facultad, nombre=nombre, codigo=codigo)
                    messages.success(request, f'Carrera "{nombre}" creada correctamente.')
                return redirect('/configuracion/?tab=carreras')
                
            elif action == 'crear_concurso':
                carrera_id = request.POST.get('carrera_id')
                nombre = request.POST.get('nombre', '').strip()
                num_etapas = int(request.POST.get('num_etapas', 2))
                fecha_inicio = request.POST.get('fecha_inicio')
                fecha_fin = request.POST.get('fecha_fin')
                estado = request.POST.get('estado', 'BORRADOR')
                if carrera_id and nombre and fecha_inicio and fecha_fin:
                    carrera = Carrera.objects.get(id=carrera_id)
                    concurso = Concurso.objects.create(
                        carrera=carrera, nombre=nombre, num_etapas=num_etapas,
                        fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, estado=estado
                    )
                    request.session['active_concurso_id'] = concurso.id
                    messages.success(request, f'Concurso "{nombre}" creado exitosamente.')
                return redirect('/configuracion/?tab=concursos')
                
            elif action == 'crear_unidad':
                numero = int(request.POST.get('numero', 1))
                nombre = request.POST.get('nombre', '').strip()
                desc = request.POST.get('descripcion', '').strip()
                
                carrera = None
                if request.user.is_superuser:
                    carrera_id = request.POST.get('carrera_id')
                    if carrera_id:
                        carrera = Carrera.objects.filter(id=carrera_id).first()
                else:
                    admin_profile = getattr(request.user, 'adminprofile', None)
                    carrera = admin_profile.carrera if admin_profile else None

                if nombre and carrera:
                    u_obj, created = UnidadTematica.objects.get_or_create(
                        carrera=carrera, numero=numero, defaults={'nombre': nombre, 'descripcion': desc}
                    )
                    if not created:
                        u_obj.nombre = nombre
                        u_obj.descripcion = desc
                        u_obj.save()
                    messages.success(request, f'Unidad Temática {numero} guardada para la carrera {carrera.nombre}.')
                elif not carrera:
                    messages.error(request, 'Debes seleccionar una carrera para asociar la Unidad Temática.')
                return redirect('/configuracion/?tab=unidades')
                
            elif action == 'crear_tema':
                unidad_id = request.POST.get('unidad_id')
                nombre = request.POST.get('nombre', '').strip()
                desc = request.POST.get('descripcion', '').strip()
                if unidad_id and nombre:
                    u_obj = UnidadTematica.objects.get(id=unidad_id)
                    Tema.objects.get_or_create(unidad=u_obj, nombre=nombre, defaults={'descripcion': desc})
                    messages.success(request, f'Tema "{nombre}" guardado en Unidad {u_obj.numero}.')
                return redirect('/configuracion/?tab=unidades')

        except Exception as e:
            messages.error(request, f'Error al guardar: {str(e)}')
            return redirect('quizzes:settings')

    # GET
    facultades = Facultad.objects.prefetch_related('carreras').all()
    carreras = Carrera.objects.select_related('facultad').filter(activa=True)
    admin_profile = getattr(request.user, 'adminprofile', None)
    admin_carrera = admin_profile.carrera if admin_profile else None
    
    if request.user.is_superuser:
        unidades = UnidadTematica.objects.select_related('carrera').prefetch_related('temas').all()
        concursos = Concurso.objects.select_related('carrera', 'carrera__facultad').all()
    elif admin_carrera:
        unidades = UnidadTematica.objects.filter(carrera=admin_carrera).select_related('carrera').prefetch_related('temas')
        concursos = Concurso.objects.filter(carrera=admin_carrera).select_related('carrera', 'carrera__facultad')
    else:
        unidades = UnidadTematica.objects.none()
        concursos = Concurso.objects.none()

    concurso_activo_id = request.session.get('active_concurso_id')
    default_tab = 'facultades' if request.user.is_superuser else 'concursos'
    tab_activa = request.GET.get('tab', default_tab)

    context = {
        'now': timezone.now(),
        'facultades': facultades,
        'carreras': carreras,
        'concursos': concursos,
        'unidades': unidades,
        'concurso_activo_id': concurso_activo_id,
        'tab_activa': tab_activa,
    }
    return render(request, 'quizzes/settings.html', context)


@login_required
@require_http_methods(["POST"])
def crear_categoria(request):
    """Crear una nueva categoría."""
    if not (request.user.is_superuser or has_full_access(request.user)):
        return JsonResponse({'success': False, 'error': 'Sin permisos'})
    
    try:
        data = json.loads(request.body)
        nombre = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', '').strip()
        activa = data.get('activa', True)
        
        if not nombre:
            return JsonResponse({'success': False, 'error': 'El nombre es obligatorio'})
        
        # Verificar que no exista otra categoría con el mismo nombre
        if Categoria.objects.filter(nombre=nombre).exists():
            return JsonResponse({'success': False, 'error': 'Ya existe una categoría con este nombre'})
        
        categoria = Categoria.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            activa=activa
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'Categoría creada exitosamente',
            'categoria_id': categoria.id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["GET"])
def obtener_categoria(request, categoria_id):
    """Obtener datos de una categoría para edición."""
    if not (request.user.is_superuser or has_full_access(request.user)):
        return JsonResponse({'success': False, 'error': 'Sin permisos'})
    
    try:
        categoria = get_object_or_404(Categoria, id=categoria_id)
        return JsonResponse({
            'success': True,
            'categoria': {
                'id': categoria.id,
                'nombre': categoria.nombre,
                'descripcion': categoria.descripcion,
                'activa': categoria.activa,
                'fecha_creacion': categoria.fecha_creacion.isoformat()
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["PUT"])
def editar_categoria(request, categoria_id):
    """Editar una categoría existente."""
    if not (request.user.is_superuser or has_full_access(request.user)):
        return JsonResponse({'success': False, 'error': 'Sin permisos'})
    
    try:
        data = json.loads(request.body)
        categoria = get_object_or_404(Categoria, id=categoria_id)
        
        nombre = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', '').strip()
        activa = data.get('activa', True)
        
        if not nombre:
            return JsonResponse({'success': False, 'error': 'El nombre es obligatorio'})
        
        # Verificar que no exista otra categoría con el mismo nombre (excluyendo la actual)
        if Categoria.objects.filter(nombre=nombre).exclude(id=categoria_id).exists():
            return JsonResponse({'success': False, 'error': 'Ya existe una categoría con este nombre'})
        
        categoria.nombre = nombre
        categoria.descripcion = descripcion
        categoria.activa = activa
        categoria.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Categoría actualizada exitosamente'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def toggle_categoria(request, categoria_id):
    """Activar o desactivar una categoría."""
    if not (request.user.is_superuser or has_full_access(request.user)):
        return JsonResponse({'success': False, 'error': 'Sin permisos'})
    
    try:
        data = json.loads(request.body)
        categoria = get_object_or_404(Categoria, id=categoria_id)
        activa = data.get('activa', not categoria.activa)
        
        categoria.activa = activa
        categoria.save()
        
        accion = "activada" if activa else "desactivada"
        return JsonResponse({
            'success': True, 
            'message': f'Categoría {accion} exitosamente'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def custom_404_view(request, exception=None, path=None):
    """Vista personalizada para errores 404"""
    from django.conf import settings
    from django.http import Http404
    from django.shortcuts import render
    
    # Si está deshabilitado, usar comportamiento estándar de Django
    if not getattr(settings, 'ENABLE_CUSTOM_ERROR_PAGES', True):
        raise Http404("Página no encontrada")
    
    # Si viene de una URL catch-all, path será el parámetro
    # Si viene del handler404, exception será el parámetro
    requested_url = path if path else request.path
    
    context = {
        'error_message': 'La página que buscas no existe',
        'error_code': '404',
        'error_description': 'El endpoint o URL solicitado no fue encontrado en nuestro servidor.',
        'requested_url': requested_url,
        'suggested_actions': [
            'Verifica que la URL esté escrita correctamente',
            'Regresa a la página principal',
            'Contacta al administrador si crees que esto es un error'
        ]
    }
    return render(request, 'errors/404.html', context, status=404)


@login_required
def dar_nuevo_intento_evaluacion(request, pk):
    """
    Vista para dar un nuevo intento a un participante en una evaluación específica
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    # Verificar permisos básicos (solo admins pueden dar nuevos intentos)
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)
    
    evaluacion = get_evaluacion_monitoreable_or_404(request, pk)
    
    try:
        data = json.loads(request.body)
        participante_id = data.get('participante_id')
        cantidad_intentos = int(data.get('cantidad_intentos', 1))
        
        if not participante_id:
            return JsonResponse({'success': False, 'error': 'ID del participante requerido'})
        if cantidad_intentos < 1 or cantidad_intentos > 10:
            return JsonResponse({'success': False, 'error': 'La cantidad de intentos debe estar entre 1 y 10'})
        
        participante = get_object_or_404(Participantes, id=participante_id)
        
        # Verificar que el participante esté autorizado para esta evaluación
        participantes_autorizados = evaluacion.get_participantes_autorizados()
        if participante not in participantes_autorizados:
            return JsonResponse({'success': False, 'error': 'Participante no autorizado para esta evaluación'})
        
        # Crear o actualizar el registro de intentos para otorgar uno adicional
        intento_config, created = IntentosParticipante.objects.get_or_create(
            participante=participante,
            evaluacion=evaluacion,
            defaults={
                'intentos_maximos': participante.intentos_maximos_default + cantidad_intentos,
                'creado_por': request.user,
                'motivo': f'{cantidad_intentos} intento(s) adicional(es) otorgado(s) por administrador'
            }
        )
        
        if not created:
            intento_config.intentos_maximos += cantidad_intentos
            intento_config.motivo = f'{cantidad_intentos} intento(s) adicional(es) otorgado(s) por administrador'
            intento_config.save(update_fields=['intentos_maximos', 'motivo'])
        
        # Recalcular los intentos disponibles después del otorgamiento
        nuevos_intentos_disponibles = participante.get_intentos_disponibles(evaluacion)
        
        return JsonResponse({
            'success': True,
            'message': f'Se otorgaron {cantidad_intentos} intento(s) a {participante.NombresCompletos}. Intentos disponibles: {nuevos_intentos_disponibles}',
            'intentos_disponibles': nuevos_intentos_disponibles,
            'intentos_usados': participante.get_intentos_usados(evaluacion)
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'})
    except ValueError:
        return JsonResponse({'success': False, 'error': 'La cantidad de intentos debe ser un número válido'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def reducir_cambios_pestana(request, pk):
    """
    Vista para reducir la cantidad de cambios de pestañas a un participante
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    # Verificar permisos básicos (solo admins pueden reducir cambios de pestañas)
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)
    
    evaluacion = get_evaluacion_monitoreable_or_404(request, pk)
    
    try:
        data = json.loads(request.body)
        participante_id = data.get('participante_id')
        cantidad_reduccion = int(data.get('cantidad_reduccion', 1))
        
        if not participante_id:
            return JsonResponse({'success': False, 'error': 'ID del participante requerido'})
        
        if cantidad_reduccion <= 0:
            return JsonResponse({'success': False, 'error': 'La cantidad de reducción debe ser mayor a 0'})
        
        participante = get_object_or_404(Participantes, id=participante_id)
        
        # Verificar que el participante esté autorizado para esta evaluación
        participantes_autorizados = evaluacion.get_participantes_autorizados()
        if participante not in participantes_autorizados:
            return JsonResponse({'success': False, 'error': 'Participante no autorizado para esta evaluación'})
        
        # Obtener el resultado activo (no completado) del participante
        resultado_activo = ResultadoEvaluacion.objects.filter(
            evaluacion=evaluacion,
            participante=participante,
            completada=False
        ).first()
        
        if not resultado_activo:
            return JsonResponse({'success': False, 'error': 'El participante no tiene una evaluación activa'})
        
        # Verificar que el participante tenga cambios de pestañas registrados
        cambios_actuales = resultado_activo.cambios_pestana or 0
        
        if cambios_actuales == 0:
            return JsonResponse({'success': False, 'error': 'El participante no ha realizado cambios de pestañas'})
        
        # Calcular nueva cantidad (no puede ser menor a 0)
        nuevos_cambios = max(0, cambios_actuales - cantidad_reduccion)
        
        # Actualizar el resultado
        resultado_activo.cambios_pestana = nuevos_cambios
        resultado_activo.agregar_alerta(
            'admin_reduccion_pestanas',
            f'Cambios de pestaña reducidos por administrador ({request.user.username}): {cambios_actuales} → {nuevos_cambios}',
            severidad='baja'
        )
        resultado_activo.save(update_fields=['cambios_pestana'])
        
        return JsonResponse({
            'success': True,
            'message': f'Se redujeron {cantidad_reduccion} cambios de pestaña a {participante.NombresCompletos}',
            'cambios_anteriores': cambios_actuales,
            'cambios_nuevos': nuevos_cambios,
            'cantidad_reducida': cantidad_reduccion
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'})
    except ValueError:
        return JsonResponse({'success': False, 'error': 'La cantidad de reducción debe ser un número válido'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def enviar_retroalimentacion(request, pk):
    """
    Envía retroalimentación por correo al representante del grupo con:
    - Retroalimentación personalizada del admin (opcional)
    - Análisis de categorías a reforzar
    - Excel con resultados detallados adjunto
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    # Verificar permisos
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        return JsonResponse({'success': False, 'error': 'No tienes permisos para esta acción'}, status=403)
    
    try:
        import json
        from django.core.mail import EmailMessage
        from django.template.loader import render_to_string
        from django.utils.html import strip_tags
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime
        
        # Obtener evaluación
        evaluacion = get_object_or_404(Evaluacion, pk=pk)
        
        # Parsear datos del request
        data = json.loads(request.body)
        grupo_id = data.get('grupo_id')
        retroalimentacion_personalizada = data.get('retroalimentacion_personalizada', '')
        categorias_dificiles = data.get('categorias_dificiles', [])
        categorias_medias = data.get('categorias_medias', [])
        
        if not grupo_id:
            return JsonResponse({'success': False, 'error': 'No se especificó el grupo'}, status=400)
        
        # Obtener grupo y representante
        grupo = get_object_or_404(GrupoParticipantes, pk=grupo_id)
        
        if not grupo.representante:
            return JsonResponse({
                'success': False, 
                'error': 'El grupo no tiene un representante asignado'
            }, status=400)
        
        representante = grupo.representante
        correo_destino = representante.CorreoRepresentante
        
        if not correo_destino:
            return JsonResponse({
                'success': False,
                'error': 'El representante no tiene un correo registrado'
            }, status=400)
        
        # Generar Excel con resultados del grupo
        excel_buffer = generar_excel_grupo(evaluacion, grupo)
        
        # Construir contenido del correo
        current_year = datetime.now().year
        
        # Parte 1: Saludo y retroalimentación personalizada (si existe)
        contenido_partes = []
        
        if retroalimentacion_personalizada and strip_tags(retroalimentacion_personalizada).strip():
            contenido_partes.append(f"""
                <div class="retroalimentacion-personalizada" style="background: #e8f5e8; border-radius: 8px; padding: 20px; margin: 25px 0; border-left: 4px solid #025a27;">
                    <h4 style="color: #025a27; margin-bottom: 15px; font-size: 18px; font-weight: 600;">
                        <i style="margin-right: 8px;">📝</i> Mensaje del Organizador
                    </h4>
                    <div style="color: #424242; font-size: 14px; line-height: 1.6;">
                        {retroalimentacion_personalizada}
                    </div>
                </div>
            """)
        
        # Parte 2: Análisis de categorías - formato profesional con tabla
        analisis_categorias_html = ""
        
        # Combinar todas las categorías
        todas_categorias = categorias_dificiles + categorias_medias
        
        if todas_categorias:
            # Calcular promedio general como calificación sobre 10
            total_porcentaje = sum(cat['porcentaje'] for cat in todas_categorias)
            promedio_porcentaje = total_porcentaje / len(todas_categorias) if todas_categorias else 0
            calificacion_promedio = promedio_porcentaje / 10  # Convertir a escala de 10
            
            # Determinar color del badge según calificación
            if calificacion_promedio >= 7:
                badge_color = "#28a745"  # Verde
            elif calificacion_promedio >= 5:
                badge_color = "#ffc107"  # Amarillo
            else:
                badge_color = "#dc3545"  # Rojo
            
            analisis_categorias_html = f"""
                <div class="analisis-categorias" style="margin: 30px 0;">
                    <h4 style="color: #025a27; margin-bottom: 20px; font-size: 20px; font-weight: 600;">
                        📊 Análisis de rendimiento
                    </h4>
                    
                    <!-- Resumen del Rendimiento -->
                    <div style="background: #e3f2fd; border-radius: 8px; padding: 20px; margin: 15px 0; border-left: 4px solid #1976d2;">
                        <h5 style="color: #1565c0; margin-bottom: 15px; font-size: 16px; font-weight: 600;">
                            📋 Resumen del rendimiento
                        </h5>
                        <p style="color: #1565c0; margin-bottom: 10px;">
                            <strong>Categorías evaluadas:</strong> {len(todas_categorias)}
                        </p>
                        <p style="color: #1565c0; margin-bottom: 0;">
                            <strong>Calificación promedio:</strong> 
                            <span style="background: {badge_color}; color: white; padding: 4px 12px; border-radius: 4px; font-weight: 600;">
                                {calificacion_promedio:.2f} / 10
                            </span>
                        </p>
                    </div>
                    
                    <!-- Tabla de Categorías -->
                    <div style="background: #f8f9fa; border-radius: 8px; padding: 20px; margin: 15px 0; border: 1px solid #dee2e6;">
                        <h5 style="color: #333; margin-bottom: 15px; font-size: 16px; font-weight: 600;">
                            📝 Rendimiento por categoría
                        </h5>
                        <table style="width: 100%; border-collapse: collapse; font-size: 14px;">
                            <thead>
                                <tr style="background: #e9ecef;">
                                    <th style="padding: 12px; text-align: left; border-bottom: 2px solid #dee2e6;">Categoría</th>
                                    <th style="padding: 12px; text-align: center; border-bottom: 2px solid #dee2e6;">Preguntas</th>
                                    <th style="padding: 12px; text-align: center; border-bottom: 2px solid #dee2e6;">% Acierto</th>
                                </tr>
                            </thead>
                            <tbody>
            """
            
            for cat in todas_categorias:
                # Determinar color del badge según porcentaje
                if cat['porcentaje'] >= 70:
                    cat_badge_color = "#28a745"  # Verde
                elif cat['porcentaje'] >= 50:
                    cat_badge_color = "#ffc107"  # Amarillo
                elif cat['porcentaje'] > 0:
                    cat_badge_color = "#dc3545"  # Rojo
                else:
                    cat_badge_color = "#6c757d"  # Gris
                
                analisis_categorias_html += f"""
                                <tr>
                                    <td style="padding: 12px; border-bottom: 1px solid #dee2e6;"><strong>{cat['nombre']}</strong></td>
                                    <td style="padding: 12px; text-align: center; border-bottom: 1px solid #dee2e6;">{cat['preguntas']}</td>
                                    <td style="padding: 12px; text-align: center; border-bottom: 1px solid #dee2e6;">
                                        <span style="background: {cat_badge_color}; color: white; padding: 4px 10px; border-radius: 4px; font-weight: 600;">
                                            {cat['porcentaje']}%
                                        </span>
                                    </td>
                                </tr>
                """
            
            analisis_categorias_html += """
                            </tbody>
                        </table>
                    </div>
                </div>
            """
        
        contenido_partes.append(analisis_categorias_html)
        
        # Parte 3: Agradecimiento
        contenido_partes.append("""
            <div class="agradecimiento" style="background: #d4edda; border-radius: 8px; padding: 20px; margin: 25px 0; border-left: 4px solid #28a745;">
                <h4 style="color: #155724; margin-bottom: 15px; font-size: 18px; font-weight: 600;">
                    🎓 Agradecimiento
                </h4>
                <p style="color: #155724; font-size: 14px; line-height: 1.6;">
                    Agradecemos sinceramente la participación de su institución en la 
                    <strong>Olimpiada Intercolegial de Matemática {}</strong>. 
                    El compromiso y dedicación de sus estudiantes es admirable.
                </p>
                <p style="color: #155724; font-size: 14px; line-height: 1.6; margin-top: 10px;">
                    Esperamos contar con su participación en futuras ediciones y seguir promoviendo 
                    el desarrollo del pensamiento lógico-matemático en nuestros jóvenes.
                </p>
                <p style="color: #155724; font-size: 14px; line-height: 1.6; margin-top: 10px; font-weight: 600;">
                    ¡Hasta la próxima olimpiada! 🏆
                </p>
            </div>
        """.format(current_year))
        
        # Construir HTML completo del correo
        html_content = f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <link href="https://fonts.googleapis.com/css2?family=Open+Sans:wght@300;400;600;700;800&display=swap" rel="stylesheet">
            <title>Retroalimentación - {evaluacion.title}</title>
            <style>
                * {{
                    margin: 0;
                    padding: 0;
                    box-sizing: border-box;
                }}
                
                body {{
                    font-family: 'Open Sans', sans-serif;
                    line-height: 1.6;
                    color: #333;
                    background: #f4f4f4;
                    padding: 20px;
                }}
                
                .email-container {{
                    max-width: 800px;
                    margin: 0 auto;
                    background: white;
                    border-radius: 8px;
                    box-shadow: 0 4px 8px rgba(0,0,0,0.1);
                    overflow: hidden;
                    border: 1px solid #ddd;
                }}
                
                .header {{
                    background: linear-gradient(135deg, #025a27 0%, #034a2a 100%);
                    color: white;
                    padding: 30px 30px;
                    text-align: center;
                    position: relative;
                }}
                
                .header::after {{
                    content: '';
                    position: absolute;
                    bottom: 0;
                    left: 0;
                    right: 0;
                    height: 3px;
                    background: linear-gradient(90deg, #ffd700 0%, #ffed4e 50%, #ffd700 100%);
                }}
                
                .header h1 {{
                    font-size: 24px;
                    font-weight: 700;
                    margin: 0 0 10px 0;
                    letter-spacing: 0.5px;
                    text-shadow: 1px 1px 2px rgba(0,0,0,0.3);
                }}
                
                .header h2 {{
                    font-size: 18px;
                    font-weight: 400;
                    margin: 0;
                    letter-spacing: 0.3px;
                    opacity: 0.95;
                    text-shadow: 1px 1px 2px rgba(0,0,0,0.3);
                }}
                
                .content {{
                    padding: 40px 30px;
                }}
                
                .greeting {{
                    font-size: 16px;
                    margin-bottom: 25px;
                    color: #555;
                    font-weight: 400;
                }}
                
                .footer {{
                    background: #f4f4f4;
                    padding: 30px;
                    text-align: center;
                    border-top: 1px solid #ddd;
                }}
                
                .footer p {{
                    color: #666;
                    font-size: 14px;
                    font-weight: 400;
                }}
                
                .footer .signature {{
                    font-weight: 600;
                    color: #025a27;
                    margin-top: 10px;
                }}
                
                @media (max-width: 600px) {{
                    body {{
                        padding: 10px;
                    }}
                    
                    .content {{
                        padding: 20px 15px;
                    }}
                    
                    .header {{
                        padding: 20px;
                    }}
                }}
            </style>
        </head>
        <body>
            <div class="email-container">
                <div class="header">
                    <h1>Universidad Técnica Estatal de Quevedo</h1>
                    <h2>Olimpiada Intercolegial de Matemática {current_year}</h2>
                </div>
                
                <div class="content">
                    <h1 style="font-size: 28px; font-weight: 700; color: #025a27; text-align: center; margin-bottom: 10px;">
                        Retroalimentación de Evaluación
                    </h1>
                    <div class="subtitle" style="font-size: 18px; color: #555; text-align: center; margin-bottom: 30px;">
                        {evaluacion.title}
                    </div>
                    
                    <div class="greeting">
                        Estimado/a <strong>{representante.NombresRepresentante}</strong>,
                    </div>
                    
                    <p style="margin-bottom: 20px;">
                        Le enviamos la retroalimentación correspondiente a la evaluación 
                        <strong>{evaluacion.title}</strong> del grupo <strong>{grupo.name}</strong> 
                        de la institución <strong>{representante.NombreColegio}</strong>.
                    </p>
                    
                    {''.join(contenido_partes)}
                    
                    <div style="background: #f8f9fa; border-radius: 8px; padding: 20px; margin: 25px 0; border: 1px solid #dee2e6;">
                        <h4 style="color: #025a27; margin-bottom: 15px; font-size: 16px; font-weight: 600;">
                            📎 Archivo Adjunto
                        </h4>
                        <p style="color: #555; font-size: 14px;">
                            Adjunto a este correo encontrará un archivo Excel con los resultados detallados 
                            de todos los participantes del grupo, incluyendo puntajes, tiempos y análisis por pregunta.
                        </p>
                    </div>
                    
                    <p style="margin-top: 25px;">
                        Si tiene alguna pregunta o necesita información adicional, no dude en contactarnos.
                    </p>
                </div>
                
                <div class="footer">
                    <p>Atentamente,</p>
                    <div class="signature">
                        Carrera de Ingeniería Mecánica<br>
                        Universidad Técnica Estatal Quevedo
                    </div>
                    <p style="margin-top: 15px; font-size: 12px; color: #999;">
                        Este es un mensaje automático, por favor no responda a este correo.<br>
                        Para soporte: <a href="mailto:olimpiadasmecanicauteq@gmail.com" style="color: #025a27;">olimpiadasmecanicauteq@gmail.com</a>
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Crear correo
        asunto = f'Retroalimentación - {evaluacion.title} - {grupo.name}'
        
        email = EmailMessage(
            subject=asunto,
            body=strip_tags(html_content),  # Versión texto plano
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[correo_destino]
        )
        
        # Añadir versión HTML
        email.content_subtype = "html"
        email.body = html_content
        
        # Adjuntar Excel
        nombre_archivo = f'Resultados_{grupo.name}_{evaluacion.title.replace(" ", "_")}.xlsx'
        email.attach(nombre_archivo, excel_buffer.getvalue(), 
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Enviar correo
        email.send()
        
        return JsonResponse({
            'success': True,
            'message': f'Retroalimentación enviada exitosamente a {correo_destino}'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        import traceback
        print(f"Error al enviar retroalimentación: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Error al enviar retroalimentación: {str(e)}'
        }, status=500)


def generar_excel_grupo(evaluacion, grupo):
    """
    Genera un archivo Excel con los resultados del grupo especificado
    Usa la misma estructura completa que exportar_resultados (3 hojas)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.utils.html import strip_tags
    from io import BytesIO
    from datetime import datetime
    from django.db.models import Avg, Max, Min
    import html
    
    # Obtener participantes del grupo
    participantes_filtrados = grupo.participantes.all()
    
    # Filtrar resultados por participantes del grupo
    resultados_completados = ResultadoEvaluacion.objects.filter(
        evaluacion=evaluacion, 
        completada=True,
        participante__in=participantes_filtrados
    )
    todos_resultados = ResultadoEvaluacion.objects.filter(
        evaluacion=evaluacion,
        participante__in=participantes_filtrados
    )
    
    # Crear un nuevo workbook
    wb = Workbook()
    
    # === HOJA 1: RESUMEN GENERAL ===
    ws_resumen = wb.active
    ws_resumen.title = "Resumen General"
    
    # Configurar estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=14)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Título principal
    titulo_reporte = f"REPORTE DE RESULTADOS - {evaluacion.title} - {grupo.name}"
    
    ws_resumen.merge_cells('A1:F1')
    ws_resumen['A1'] = titulo_reporte
    ws_resumen['A1'].font = title_font
    ws_resumen['A1'].alignment = Alignment(horizontal='center')
    
    # Información general
    ws_resumen['A3'] = "Información General"
    ws_resumen['A3'].font = header_font
    ws_resumen['A3'].fill = header_fill
    
    ws_resumen['A4'] = "Título:"
    ws_resumen['B4'] = evaluacion.title
    ws_resumen['A5'] = "Etapa:"
    ws_resumen['B5'] = f"Etapa {evaluacion.etapa}"
    ws_resumen['A6'] = "Fecha de inicio:"
    ws_resumen['B6'] = evaluacion.start_time.strftime("%d/%m/%Y %H:%M")
    ws_resumen['A7'] = "Fecha de fin:"
    ws_resumen['B7'] = evaluacion.end_time.strftime("%d/%m/%Y %H:%M")
    ws_resumen['A8'] = "Duración:"
    ws_resumen['B8'] = f"{evaluacion.duration_minutes} minutos"
    ws_resumen['A9'] = "Grupo:"
    ws_resumen['B9'] = grupo.name
    
    # Estadísticas de participación
    participantes_completaron = resultados_completados.values_list('participante', flat=True).distinct().count()
    participantes_con_resultados = todos_resultados.values_list('participante', flat=True).distinct().count()
    participantes_en_progreso = todos_resultados.filter(completada=False).values_list('participante', flat=True).distinct().count()
    
    ws_resumen['D3'] = "Estadísticas de Participación"
    ws_resumen['D3'].font = header_font
    ws_resumen['D3'].fill = header_fill
    
    ws_resumen['D4'] = "Participantes que completaron:"
    ws_resumen['E4'] = participantes_completaron
    ws_resumen['D5'] = "Participantes en progreso:"
    ws_resumen['E5'] = participantes_en_progreso
    ws_resumen['D6'] = "Total con intentos:"
    ws_resumen['E6'] = participantes_con_resultados
    
    # Estadísticas de rendimiento
    if resultados_completados.exists():
        stats = resultados_completados.aggregate(
            promedio=Avg('puntos_obtenidos'),
            maximo=Max('puntos_obtenidos'),
            minimo=Min('puntos_obtenidos'),
            tiempo_promedio=Avg('tiempo_utilizado')
        )
        
        ws_resumen['A11'] = "Estadísticas de Rendimiento"
        ws_resumen['A11'].font = header_font
        ws_resumen['A11'].fill = header_fill
        
        ws_resumen['A12'] = "Promedio:"
        ws_resumen['B12'] = f"{stats['promedio']:.2f}/10" if stats['promedio'] else "N/A"
        ws_resumen['A13'] = "Mejor puntaje:"
        ws_resumen['B13'] = f"{stats['maximo']:.2f}/10" if stats['maximo'] else "N/A"
        ws_resumen['A14'] = "Peor puntaje:"
        ws_resumen['B14'] = f"{stats['minimo']:.2f}/10" if stats['minimo'] else "N/A"
        ws_resumen['A15'] = "Tiempo promedio:"
        ws_resumen['B15'] = f"{int(stats['tiempo_promedio'])} min" if stats['tiempo_promedio'] else "N/A"
    
    # === HOJA 2: RESULTADOS DETALLADOS ===
    ws_resultados = wb.create_sheet("Resultados Detallados")
    
    # Encabezados
    headers = ['#', 'Participante', 'Cédula', 'Grupo', 'Puntaje', 'Porcentaje', 'Tiempo (min)', 
              'Intento', 'Fecha Inicio', 'Fecha Fin', 'Estado']
    
    for col, header in enumerate(headers, 1):
        cell = ws_resultados.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Datos de resultados
    resultados = todos_resultados.order_by('-puntos_obtenidos', 'tiempo_utilizado')
    
    for row, resultado in enumerate(resultados, 2):
        ws_resultados.cell(row=row, column=1, value=row-1)
        ws_resultados.cell(row=row, column=2, value=resultado.participante.NombresCompletos)
        ws_resultados.cell(row=row, column=3, value=resultado.participante.cedula)
        ws_resultados.cell(row=row, column=4, value=grupo.name)
        ws_resultados.cell(row=row, column=5, value=f"{resultado.puntos_obtenidos:.2f}")
        ws_resultados.cell(row=row, column=6, value=f"{resultado.get_puntaje_porcentaje():.1f}%")
        ws_resultados.cell(row=row, column=7, value=resultado.tiempo_utilizado)
        ws_resultados.cell(row=row, column=8, value=resultado.numero_intento)
        ws_resultados.cell(row=row, column=9, value=resultado.fecha_inicio.strftime("%d/%m/%Y %H:%M") if resultado.fecha_inicio else "N/A")
        ws_resultados.cell(row=row, column=10, value=resultado.fecha_fin.strftime("%d/%m/%Y %H:%M") if resultado.fecha_fin else "N/A")
        ws_resultados.cell(row=row, column=11, value="Completado" if resultado.completada else "En progreso")
        
        # Aplicar bordes
        for col_idx in range(1, len(headers) + 1):
            ws_resultados.cell(row=row, column=col_idx).border = border
    
    # === HOJA 3: ANÁLISIS POR PREGUNTA ===
    ws_preguntas = wb.create_sheet("Análisis por Pregunta")
    
    # Encabezados
    headers_preguntas = ['Pregunta', 'Texto', 'Correctas', 'Incorrectas', 'Sin Responder', '% Acierto', 'Dificultad']
    
    for col, header in enumerate(headers_preguntas, 1):
        cell = ws_preguntas.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Datos por pregunta
    # Crear cache de opciones correctas por pregunta
    preguntas_opciones_correctas = {}
    for pregunta in evaluacion.preguntas.prefetch_related('opciones'):
        opciones_correctas = {}
        for opcion in pregunta.opciones.all():
            opciones_correctas[opcion.id] = opcion.is_correct
        preguntas_opciones_correctas[pregunta.id] = opciones_correctas
    
    for row, pregunta in enumerate(evaluacion.preguntas.all(), 2):
        correctas = 0
        incorrectas = 0
        sin_responder = 0
        
        # Obtener las opciones correctas para esta pregunta desde el cache
        opciones_correctas_pregunta = preguntas_opciones_correctas.get(pregunta.id, {})
        
        for resultado in resultados_completados:
            respuestas = resultado.respuestas_guardadas
            
            # Validar que respuestas no sea None o vacío
            if not respuestas or not isinstance(respuestas, dict):
                sin_responder += 1
                continue
            
            # Buscar la respuesta para esta pregunta con el formato correcto
            # Las claves se guardan como "pregunta_407", "pregunta_410", etc.
            opcion_id = None
            pregunta_key = f"pregunta_{pregunta.id}"
            
            if pregunta_key in respuestas:
                opcion_id = respuestas[pregunta_key]
            
            if opcion_id is not None:
                try:
                    # Normalizar opcion_id a entero
                    if isinstance(opcion_id, str) and opcion_id.isdigit():
                        opcion_id = int(opcion_id)
                    elif not isinstance(opcion_id, int):
                        sin_responder += 1
                        continue
                    
                    # Verificar si la respuesta es correcta usando el cache
                    if opcion_id in opciones_correctas_pregunta:
                        if opciones_correctas_pregunta[opcion_id]:
                            correctas += 1
                        else:
                            incorrectas += 1
                    else:
                        # La opción no existe, contar como sin responder
                        sin_responder += 1
                        
                except (ValueError, TypeError):
                    sin_responder += 1
            else:
                sin_responder += 1
        
        total_respuestas = correctas + incorrectas  # Solo respuestas dadas (no incluir sin_responder)
        porcentaje = (correctas / total_respuestas * 100) if total_respuestas > 0 else 0
        
        # Determinar dificultad - Si nadie respondió, mostrar "Sin datos"
        if correctas == 0 and incorrectas == 0:
            dificultad = 'Sin datos'
        elif porcentaje > 70:
            dificultad = 'Fácil'
        elif porcentaje > 40:
            dificultad = 'Media'
        else:
            dificultad = 'Difícil'
        
        # Limpiar el texto HTML de la pregunta y decodificar entidades HTML
        texto_sin_html = strip_tags(pregunta.text)
        texto_limpio = html.unescape(texto_sin_html).strip()
        # Limitar el largo del texto para que no desborde la celda
        texto_pregunta = texto_limpio[:150] + "..." if len(texto_limpio) > 150 else texto_limpio
        
        ws_preguntas.cell(row=row, column=1, value=f"P{row-1}")
        ws_preguntas.cell(row=row, column=2, value=texto_pregunta)
        ws_preguntas.cell(row=row, column=3, value=correctas)
        ws_preguntas.cell(row=row, column=4, value=incorrectas)
        ws_preguntas.cell(row=row, column=5, value=sin_responder)
        ws_preguntas.cell(row=row, column=6, value=f"{porcentaje:.1f}%")
        ws_preguntas.cell(row=row, column=7, value=dificultad)
        
        # Aplicar bordes
        for col in range(1, len(headers_preguntas) + 1):
            ws_preguntas.cell(row=row, column=col).border = border
    
    # Ajustar ancho de columnas
    for ws in [ws_resumen, ws_resultados, ws_preguntas]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


@login_required
def descargar_retroalimentacion_pdf(request, pk):
    """
    Genera y descarga un PDF con la retroalimentación del grupo
    (similar al correo pero sin el archivo Excel adjunto)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    # Verificar permisos
    if not (request.user.is_superuser or hasattr(request.user, 'adminprofile')):
        return JsonResponse({'success': False, 'error': 'No tienes permisos para esta acción'}, status=403)
    
    try:
        import json
        from django.utils.html import strip_tags
        from io import BytesIO
        from datetime import datetime
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
        from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
        import html as html_module
        import os
        
        # Obtener evaluación
        evaluacion = get_object_or_404(Evaluacion, pk=pk)
        
        # Parsear datos del request
        data = json.loads(request.body)
        grupo_id = data.get('grupo_id')
        retroalimentacion_personalizada = data.get('retroalimentacion_personalizada', '')
        categorias_dificiles = data.get('categorias_dificiles', [])
        categorias_medias = data.get('categorias_medias', [])
        
        if not grupo_id:
            return JsonResponse({'success': False, 'error': 'No se especificó el grupo'}, status=400)
        
        # Obtener grupo y representante
        grupo = get_object_or_404(GrupoParticipantes, pk=grupo_id)
        
        if not grupo.representante:
            return JsonResponse({
                'success': False, 
                'error': 'El grupo no tiene un representante asignado'
            }, status=400)
        
        representante = grupo.representante
        
        # Crear buffer para PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                               rightMargin=72, leftMargin=72,
                               topMargin=72, bottomMargin=18)
        
        # Estilos
        styles = getSampleStyleSheet()
        story = []
        
        # Estilo personalizado para título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#025a27'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        # Estilo para subtítulo
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=14,
            textColor=colors.HexColor('#555555'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        # Estilo para encabezado de sección
        heading_style = ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#025a27'),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        )
        
        # Estilo para texto normal
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=12,
            alignment=TA_JUSTIFY
        )
        
        # Estilo para texto en caja
        box_style = ParagraphStyle(
            'BoxText',
            parent=styles['Normal'],
            fontSize=10,
            leftIndent=20,
            rightIndent=20,
            spaceAfter=6
        )
        
        # Construir contenido del PDF
        current_year = datetime.now().year
        
        # Encabezado con fondo verde
        header_data = [
            [Paragraph('<b><font size=16>Universidad Técnica Estatal de Quevedo</font></b>', 
                      ParagraphStyle('HeaderTitle', parent=subtitle_style, textColor=colors.white, fontSize=16, alignment=TA_CENTER))],
            [Paragraph(f'<font size=12>Olimpiada Intercolegial de Matemática {current_year}</font>', 
                      ParagraphStyle('HeaderSubtitle', parent=normal_style, textColor=colors.white, fontSize=12, alignment=TA_CENTER))]
        ]
        header_table = Table(header_data, colWidths=[6.5*inch])
        header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#025a27')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
            ('TOPPADDING', (0, 0), (-1, -1), 15),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 20))
        
        # Cargar logos usando la misma lógica que exportar_ranking_pdf
        logo_mecanica = None
        logo_uteq = None
        
        # Cargar Logo Mecánica
        logo_mecanica_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logoMecanica.png')
        if os.path.exists(logo_mecanica_path):
            try:
                logo_mecanica = Image(logo_mecanica_path, width=1.6*inch, height=1.2*inch)
            except Exception as e:
                print(f"Error al cargar logo Mecánica: {e}")
        
        # Cargar Logo UTEQ
        logo_uteq_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo-uteq.png')
        if os.path.exists(logo_uteq_path):
            try:
                logo_uteq = Image(logo_uteq_path, width=1.6*inch, height=1.2*inch)
            except Exception as e:
                print(f"Error al cargar logo UTEQ: {e}")
        
        # Tabla de logos centrados
        if logo_mecanica or logo_uteq:
            logo_row = []
            
            if logo_mecanica:
                logo_row.append(logo_mecanica)
            else:
                logo_row.append(Paragraph("", styles['Normal']))
            
            logo_row.append(Paragraph("", styles['Normal']))  # Espacio central
            
            if logo_uteq:
                logo_row.append(logo_uteq)
            else:
                logo_row.append(Paragraph("", styles['Normal']))
            
            logo_table = Table([logo_row], colWidths=[2.4*inch, 1.2*inch, 2.4*inch])
            logo_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('ALIGN', (2, 0), (2, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, 0), 0),
                ('RIGHTPADDING', (0, 0), (-1, 0), 0),
                ('TOPPADDING', (0, 0), (-1, 0), 5),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
            ]))
            
            story.append(logo_table)
            story.append(Spacer(1, 20))
        
        # Título
        story.append(Paragraph('Retroalimentación de Evaluación', title_style))
        story.append(Paragraph(evaluacion.title, subtitle_style))
        story.append(Spacer(1, 20))
        
        # Saludo
        story.append(Paragraph(f'Estimado/a <b>{representante.NombresRepresentante}</b>,', normal_style))
        story.append(Paragraph(
            f'Le enviamos la retroalimentación correspondiente a la evaluación <b>{evaluacion.title}</b> '
            f'del grupo <b>{grupo.name}</b> de la institución <b>{representante.NombreColegio}</b>.',
            normal_style
        ))
        story.append(Spacer(1, 20))
        
        # Parte 1: Retroalimentación personalizada (si existe)
        if retroalimentacion_personalizada and strip_tags(retroalimentacion_personalizada).strip():
            story.append(Paragraph('Mensaje de las autoridades organizadoras', heading_style))
            
            # Limpiar HTML y entidades
            texto_limpio = strip_tags(retroalimentacion_personalizada)
            texto_limpio = html_module.unescape(texto_limpio)
            
            # Dividir en párrafos
            parrafos = texto_limpio.split('\n')
            for parrafo in parrafos:
                if parrafo.strip():
                    story.append(Paragraph(parrafo.strip(), box_style))
            
            story.append(Spacer(1, 20))
        
        # Parte 2: Análisis de categorías
        if categorias_dificiles or categorias_medias:
            story.append(Paragraph('Análisis de Rendimiento por Categorías', heading_style))
            story.append(Spacer(1, 10))
            
            # Categorías críticas
            if categorias_dificiles:
                story.append(Paragraph('Categorías que Requieren Mayor Atención (< 50% acierto)', 
                                      ParagraphStyle('Warning', parent=heading_style, fontSize=12, 
                                                    textColor=colors.HexColor('#856404'))))
                story.append(Paragraph('Las siguientes categorías necesitan ser reforzadas:', box_style))
                
                # Crear tabla de categorías
                cat_data = [['Categoría', 'Acierto', 'Preguntas']]
                for cat in categorias_dificiles:
                    cat_data.append([
                        cat['nombre'],
                        f"{cat['porcentaje']}%",
                        str(cat['preguntas'])
                    ])
                
                cat_table = Table(cat_data, colWidths=[3.5*inch, 1*inch, 1*inch])
                cat_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ffc107')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('TOPPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fff3cd')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#856404')),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))
                story.append(cat_table)
                story.append(Spacer(1, 10))
                story.append(Paragraph(
                    '<b>Recomendación:</b> Revisar fundamentos teóricos y proporcionar ejercicios de refuerzo específicos en estas áreas.',
                    box_style
                ))
                story.append(Spacer(1, 15))
            
            # Categorías en desarrollo
            if categorias_medias:
                story.append(Paragraph('Categorías en Desarrollo (50-80% acierto)', 
                                      ParagraphStyle('Info', parent=heading_style, fontSize=12, 
                                                    textColor=colors.HexColor('#0c5460'))))
                story.append(Paragraph('Estas categorías muestran progreso pero necesitan más práctica:', box_style))
                
                # Crear tabla de categorías
                cat_data = [['Categoría', 'Acierto', 'Preguntas']]
                for cat in categorias_medias:
                    cat_data.append([
                        cat['nombre'],
                        f"{cat['porcentaje']}%",
                        str(cat['preguntas'])
                    ])
                
                cat_table = Table(cat_data, colWidths=[3.5*inch, 1*inch, 1*inch])
                cat_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#17a2b8')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('TOPPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#d1ecf1')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#0c5460')),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))
                story.append(cat_table)
                story.append(Spacer(1, 10))
                story.append(Paragraph(
                    '<b>Recomendación:</b> Continuar práctica con ejercicios intermedios y ejemplos aplicados.',
                    box_style
                ))
                story.append(Spacer(1, 20))
        
        # Parte 3: Agradecimiento
        story.append(Paragraph('Agradecimiento', heading_style))
        story.append(Paragraph(
            f'Agradecemos sinceramente la participación de su institución en la '
            f'<b>Olimpiada Intercolegial de Matemática {current_year}</b>. '
            f'El compromiso y dedicación de sus estudiantes es admirable.',
            normal_style
        ))
        story.append(Paragraph(
            'Esperamos contar con su participación en futuras ediciones y seguir promoviendo '
            'el desarrollo del pensamiento lógico-matemático en nuestros jóvenes.',
            normal_style
        ))
        story.append(Paragraph('¡Hasta la próxima olimpiada!', normal_style))
        story.append(Spacer(1, 20))
        
        # Cierre
        story.append(Paragraph(
            'Si tiene alguna pregunta o necesita información adicional, no dude en contactarnos.',
            normal_style
        ))
        story.append(Spacer(1, 30))
        
        # Footer
        footer_data = [
            [Paragraph('Atentamente,', ParagraphStyle('Footer', parent=normal_style, alignment=TA_CENTER))],
            [Paragraph('<b>Carrera de Ingeniería Mecánica</b><br/>Universidad Técnica Estatal Quevedo', 
                      ParagraphStyle('FooterBold', parent=normal_style, alignment=TA_CENTER, 
                                    textColor=colors.HexColor('#025a27'), fontName='Helvetica-Bold'))],
            [Paragraph('<font size=9>Para soporte: olimpiadasmecanicauteq@gmail.com</font>', 
                      ParagraphStyle('FooterSmall', parent=normal_style, alignment=TA_CENTER, 
                                    textColor=colors.HexColor('#999999')))]
        ]
        footer_table = Table(footer_data, colWidths=[6.5*inch])
        footer_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(footer_table)
        
        # Construir PDF
        doc.build(story)
        
        # Preparar respuesta
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        nombre_archivo = f'Retroalimentacion_{grupo.name}_{evaluacion.title.replace(" ", "_")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        
        return response
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        import traceback
        print(f"Error al generar PDF: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Error al generar PDF: {str(e)}'
        }, status=500)


# ==============================================================================
# VISTAS DE GESTIÓN ORGANIZACIONAL (FACULTADES, CARRERAS, CONCURSOS Y TEMAS)
# ==============================================================================

@login_required
@superuser_required
def gestionar_facultades(request):
    """Permite registrar y gestionar Facultades"""
    from .models import Facultad
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        siglas = request.POST.get('siglas', '').strip()
        if nombre:
            facultad, created = Facultad.objects.get_or_create(nombre=nombre, defaults={'siglas': siglas})
            if not created:
                facultad.siglas = siglas
                facultad.save()
            messages.success(request, f'Facultad "{nombre}" guardada correctamente.')
        else:
            messages.error(request, 'El nombre de la facultad es obligatorio.')
        return redirect('quizzes:gestionar_facultades')
    
    facultades = Facultad.objects.prefetch_related('carreras').all()
    return render(request, 'quizzes/gestionar_facultades.html', {'facultades': facultades})


@login_required
@superuser_required
def gestionar_carreras(request):
    """Permite registrar y gestionar Carreras vinculadas a Facultades"""
    from .models import Facultad, Carrera
    if request.method == 'POST':
        facultad_id = request.POST.get('facultad_id')
        nombre = request.POST.get('nombre', '').strip()
        codigo = request.POST.get('codigo', '').strip()
        if facultad_id and nombre:
            try:
                facultad = Facultad.objects.get(id=facultad_id)
                Carrera.objects.create(facultad=facultad, nombre=nombre, codigo=codigo)
                messages.success(request, f'Carrera "{nombre}" creada correctamente.')
            except Exception as e:
                messages.error(request, f'Error al crear carrera: {str(e)}')
        else:
            messages.error(request, 'Debe seleccionar una facultad e ingresar el nombre de la carrera.')
        return redirect('quizzes:gestionar_carreras')
    
    facultades = Facultad.objects.filter(activa=True)
    carreras = Carrera.objects.select_related('facultad').all()
    return render(request, 'quizzes/gestionar_carreras.html', {'facultades': facultades, 'carreras': carreras})


@login_required
@full_access_required
def gestionar_concursos(request):
    """Permite crear y gestionar Concursos u Olimpiadas por Carrera"""
    from .models import Concurso, Carrera
    user = request.user
    
    admin_profile = getattr(user, 'adminprofile', None)
    if not user.is_superuser and admin_profile and admin_profile.carrera:
        carreras = Carrera.objects.filter(id=admin_profile.carrera.id)
    else:
        carreras = Carrera.objects.filter(activa=True)
        
    if request.method == 'POST':
        carrera_id = request.POST.get('carrera_id')
        nombre = request.POST.get('nombre', '').strip()
        num_etapas = int(request.POST.get('num_etapas', 2))
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        estado = request.POST.get('estado', 'BORRADOR')
        
        if carrera_id and nombre and fecha_inicio and fecha_fin:
            try:
                carrera = Carrera.objects.get(id=carrera_id)
                concurso = Concurso.objects.create(
                    carrera=carrera,
                    nombre=nombre,
                    num_etapas=num_etapas,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,
                    estado=estado
                )
                request.session['active_concurso_id'] = concurso.id
                messages.success(request, f'Concurso "{nombre}" creado exitosamente.')
            except ValidationError as e:
                error_msg = extract_validation_error_message(e)
                messages.error(request, f'Error al crear concurso: {error_msg}')
            except Exception as e:
                messages.error(request, f'Error al crear concurso: {str(e)}')
        else:
            messages.error(request, 'Faltan campos obligatorios para el concurso.')
        return redirect('quizzes:gestionar_concursos')
        
    concursos = Concurso.objects.select_related('carrera', 'carrera__facultad').all()
    
    # Agregar contador de registros vinculados a cada concurso para deshabilitar campos en la plantilla
    for conc in concursos:
        conc.tiene_registros = (
            conc.participantes.exists() or
            conc.representantes.exists() or
            conc.grupos.exists() or
            conc.evaluaciones.exists()
        )
        
    return render(request, 'quizzes/gestionar_concursos.html', {
        'concursos': concursos, 
        'carreras': carreras,
    })


@login_required
@full_access_required
def editar_concurso(request, concurso_id):
    """Permite editar los datos de un concurso existente aplicando reglas de integridad"""
    from .models import Concurso, Carrera, Evaluacion
    from django.utils import timezone

    concurso = get_object_or_404(Concurso, id=concurso_id)

    if request.method == 'POST':
        # 1. Restricción de Evaluación Activa en Ejecución
        now = timezone.now()
        evaluaciones_activas = Evaluacion.objects.filter(
            concurso=concurso,
            start_time__lte=now,
            end_time__gte=now
        )
        if evaluaciones_activas.exists():
            messages.error(request, '⚠️ No se puede editar el concurso porque tiene una evaluación en curso activa en este momento.')
            return redirect(request.META.get('HTTP_REFERER', 'quizzes:gestionar_concursos'))

        nombre = request.POST.get('nombre', '').strip()
        carrera_id = request.POST.get('carrera_id')
        num_etapas = request.POST.get('num_etapas')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        estado = request.POST.get('estado')

        if not (nombre and fecha_inicio and fecha_fin):
            messages.error(request, 'Faltan campos obligatorios para actualizar el concurso.')
            return redirect(request.META.get('HTTP_REFERER', 'quizzes:gestionar_concursos'))

        try:
            if carrera_id:
                nueva_carrera = get_object_or_404(Carrera, id=carrera_id)
                concurso.carrera = nueva_carrera

            concurso.nombre = nombre
            if num_etapas:
                concurso.num_etapas = int(num_etapas)
            concurso.fecha_inicio = fecha_inicio
            concurso.fecha_fin = fecha_fin
            if estado:
                concurso.estado = estado

            concurso.save()
            messages.success(request, f'Concurso "{concurso.nombre}" actualizado exitosamente.')
        except ValidationError as e:
            error_msg = extract_validation_error_message(e)
            messages.error(request, f'Error al actualizar concurso: {error_msg}')
        except Exception as e:
            messages.error(request, f'Error al actualizar concurso: {str(e)}')

    return redirect(request.META.get('HTTP_REFERER', 'quizzes:gestionar_concursos'))


@login_required
def cambiar_contexto_activo(request):
    """
    Vista para cambiar el Concurso y/o Carrera activos en la sesión del usuario (Selector Navbar).
    """
    if request.method in ['POST', 'GET']:
        concurso_id = request.POST.get('concurso_id') or request.GET.get('concurso_id')
        carrera_id = request.POST.get('carrera_id') or request.GET.get('carrera_id')
        
        if carrera_id is not None and request.user.is_superuser:
            if str(carrera_id) == 'ALL':
                request.session['active_carrera_id'] = 'ALL'
                request.session['active_concurso_id'] = 'ALL'
            else:
                try:
                    request.session['active_carrera_id'] = int(carrera_id)
                    request.session['active_concurso_id'] = 'ALL'
                except ValueError:
                    pass

        if concurso_id is not None:
            if str(concurso_id) == 'ALL':
                request.session['active_concurso_id'] = 'ALL'
            else:
                try:
                    request.session['active_concurso_id'] = int(concurso_id)
                except ValueError:
                    pass
                    
        next_url = request.META.get('HTTP_REFERER') or reverse('quizzes:dashboard')
        return redirect(next_url)


@login_required
@full_access_required
def gestionar_unidades_temas(request):
    """Gestión de Unidades Temáticas y Temas académicos"""
    from .models import UnidadTematica, Tema
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        if tipo == 'unidad':
            numero = int(request.POST.get('numero', 1))
            nombre = request.POST.get('nombre', '').strip()
            desc = request.POST.get('descripcion', '').strip()
            if nombre:
                UnidadTematica.objects.get_or_create(numero=numero, defaults={'nombre': nombre, 'descripcion': desc})
                messages.success(request, f'Unidad Temática {numero} guardada.')
        elif tipo == 'tema':
            unidad_id = request.POST.get('unidad_id')
            nombre = request.POST.get('nombre', '').strip()
            desc = request.POST.get('descripcion', '').strip()
            if unidad_id and nombre:
                u_obj = UnidadTematica.objects.get(id=unidad_id)
                Tema.objects.get_or_create(unidad=u_obj, nombre=nombre, defaults={'descripcion': desc})
                messages.success(request, f'Tema "{nombre}" guardado en la Unidad {u_obj.numero}.')
        return redirect('quizzes:gestionar_unidades_temas')
        
    unidades = UnidadTematica.objects.prefetch_related('temas').all()
    return render(request, 'quizzes/gestionar_unidades_temas.html', {'unidades': unidades})
